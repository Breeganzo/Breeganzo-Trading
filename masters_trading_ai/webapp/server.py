"""
Masters AI Trading Bot — Flask Web Application
================================================
Groww-inspired stock cards with live prices, ML predictions,
expected vs actual tracking, and alpha calculation.

Run:  python webapp/server.py
Open: http://localhost:5001
"""

import sys
import os
import json
import csv
import time
import smtplib
import logging
import threading
import resource
from email.message import EmailMessage
from collections import deque, defaultdict
from uuid import uuid4
from io import StringIO
from pathlib import Path
from datetime import datetime, date, timedelta
from zoneinfo import ZoneInfo

from dotenv import load_dotenv

# Load .env from project root BEFORE any other imports that need env vars
PROJECT_ROOT = Path(__file__).resolve().parent.parent
load_dotenv(PROJECT_ROOT / ".env")

# ── Validate critical environment variables at startup ──
_REQUIRED_ENV = []  # Add keys here if they should be mandatory
_OPTIONAL_ENV = {
    "GROQ_API_KEY": "AI explanations disabled",
    "FLASK_SECRET_KEY": "Using dev fallback",
}
for _key in _REQUIRED_ENV:
    if not os.environ.get(_key):
        raise EnvironmentError(
            f"Missing required env var: {_key}. Copy .env.example → .env and fill in values."
        )
for _key, _msg in _OPTIONAL_ENV.items():
    if not os.environ.get(_key):
        print(f"⚠️  {_key} not set — {_msg}. See .env.example")

# ── CRITICAL: import torch FIRST to avoid segfault with statsmodels C extensions ──
import torch  # noqa: F401

from flask import Flask, jsonify, render_template, request, make_response
from flask_cors import CORS
import numpy as np
import pandas as pd

try:
    from scipy.optimize import linprog
except Exception:
    linprog = None

# Add project root to path
sys.path.insert(0, str(PROJECT_ROOT))

from src.inference.predictor import LivePredictor, get_market_status, get_intraday_data
from src.tracking.prediction_logger import PredictionLogger
from src.tracking.performance_reporter import PerformanceReporter
from src.backtest.costs import GrowwCostCalculator

from webapp.groq_explainer import (
    explain_fundamental,
    explain_greek,
    explain_indicator,
    get_groq_strategy,
    get_combined_strategy,
    get_stock_overview,
    get_news_sentiment,
    explain_risk_term,
    get_groq_price_forecast,
    explain_model,
    stock_chat_response,
    portfolio_profit_suggestion,
    suggest_ticker_shortlist,
    review_trade_plan,
    ai_risk_assessment,
    get_groq_system_status,
    set_groq_request_endpoint,
    reset_groq_request_endpoint,
)
from webapp.prediction_tracker import PredictionTracker

import yaml

# ── Logging setup ──────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger("trading_bot")

# ---------------------------------------------------------------------------
# App setup
# ---------------------------------------------------------------------------
app = Flask(
    __name__,
    template_folder=str(Path(__file__).parent / "templates"),
    static_folder=str(Path(__file__).parent / "static"),
)
app.secret_key = os.environ.get("FLASK_SECRET_KEY", "dev-fallback-not-for-production")
CORS(app)


# Custom JSON encoder for numpy types
class NumpyEncoder(json.JSONEncoder):
    def default(self, obj):
        if isinstance(obj, (np.integer,)):
            return int(obj)
        if isinstance(obj, (np.floating,)):
            return float(obj)
        if isinstance(obj, (np.bool_,)):
            return bool(obj)
        if isinstance(obj, np.ndarray):
            return obj.tolist()
        return super().default(obj)


# Flask 3.0+ removed json_encoder; use json_provider_class instead
from flask.json.provider import DefaultJSONProvider


class NumpyJSONProvider(DefaultJSONProvider):
    def default(self, obj):
        if isinstance(obj, (np.integer,)):
            return int(obj)
        if isinstance(obj, (np.floating,)):
            return float(obj)
        if isinstance(obj, (np.bool_,)):
            return bool(obj)
        if isinstance(obj, np.ndarray):
            return obj.tolist()
        return super().default(obj)


app.json_provider_class = NumpyJSONProvider
app.json = NumpyJSONProvider(app)


@app.before_request
def _bind_groq_endpoint_scope():
    endpoint_name = request.endpoint or request.path or "unknown_endpoint"
    request._groq_endpoint_token = set_groq_request_endpoint(endpoint_name)  # type: ignore[attr-defined]


@app.teardown_request
def _clear_groq_endpoint_scope(_exc=None):
    token = getattr(request, "_groq_endpoint_token", None)
    if token is not None:
        reset_groq_request_endpoint(token)
    return None


IST = ZoneInfo("Asia/Kolkata")
CONFIG_DIR = PROJECT_ROOT / "config"
CACHE_DIR = PROJECT_ROOT / "cache"
PREDICTION_LOG_DIR = CACHE_DIR / "prediction_log"
PORTFOLIO_FILE = CACHE_DIR / "portfolio.json"
PORTFOLIO_TRADES_FILE = CACHE_DIR / "portfolio_trades.json"
PORTFOLIO_SIM_FILE = CACHE_DIR / "portfolio_sim.json"
SIMULATED_TRADE_LOG_FILE = PREDICTION_LOG_DIR / "simulated_trades.jsonl"
SIMULATED_TRADE_EXCEL_FILE = PREDICTION_LOG_DIR / "simulated_trades.xlsx"
SIMULATED_TRADE_CSV_FILE = PREDICTION_LOG_DIR / "simulated_trades.csv"
DAILY_TRADE_REPORT_DIR = PREDICTION_LOG_DIR / "daily_reports"
DELISTED_TICKERS_FILE = CACHE_DIR / "delisted_tickers.csv"
PREMARKET_OUTLOOK_FILE = CACHE_DIR / "premarket_outlook.json"
PREDICTION_SNAPSHOTS_FILE = CACHE_DIR / "prediction_snapshots.json"
GROQ_FORECAST_CACHE_FILE = CACHE_DIR / "groq_forecasts.json"
SNAPSHOT_SCHEMA_VERSION = 1
CACHE_DIR.mkdir(parents=True, exist_ok=True)
PREDICTION_LOG_DIR.mkdir(parents=True, exist_ok=True)
DAILY_TRADE_REPORT_DIR.mkdir(parents=True, exist_ok=True)
SIMULATION_DEFAULT_CASH = float(os.environ.get("SIMULATION_DEFAULT_CASH", "40000"))
SIMULATION_DEFAULT_CASH = max(1000.0, SIMULATION_DEFAULT_CASH)
AUTO_SIMULATION_ENABLED = os.environ.get("AUTO_SIMULATION_ENABLED", "true").lower() in (
    "1",
    "true",
    "yes",
)
AUTO_SIMULATION_AUTO_BUY = os.environ.get(
    "AUTO_SIMULATION_AUTO_BUY",
    "true",
).lower() in ("1", "true", "yes")
AUTO_SIMULATION_INTERVAL_SEC = max(
    1,
    int(os.environ.get("AUTO_SIMULATION_INTERVAL_SEC", "1")),
)
YF_MAX_PARALLEL = max(1, int(os.environ.get("YF_MAX_PARALLEL", "2")))
YF_CALL_TIMEOUT_SEC = max(3.0, float(os.environ.get("YF_CALL_TIMEOUT_SEC", "15")))
FD_SOFT_GUARD_RATIO = float(os.environ.get("FD_SOFT_GUARD_RATIO", "0.85"))
FD_TARGET_SOFT_LIMIT = max(256, int(os.environ.get("FD_TARGET_SOFT_LIMIT", "4096")))
_yf_download_semaphore = threading.BoundedSemaphore(YF_MAX_PARALLEL)

# Trading-desk risk management defaults (simulation only).
RISK_PER_TRADE = float(os.environ.get("RISK_PER_TRADE", "0.01"))
MAX_DAILY_LOSS = float(os.environ.get("MAX_DAILY_LOSS", "0.05"))
MAX_POSITION_SIZE = float(os.environ.get("MAX_POSITION_SIZE", "0.40"))
ATR_PERIOD = int(os.environ.get("ATR_PERIOD", "14"))
STOP_LOSS_HIGH_CONF = float(os.environ.get("STOP_LOSS_HIGH_CONF", "0.02"))
STOP_LOSS_MED_CONF = float(os.environ.get("STOP_LOSS_MED_CONF", "0.025"))
STOP_LOSS_LOW_CONF = float(os.environ.get("STOP_LOSS_LOW_CONF", "0.03"))
STOP_LOSS_HIGH_CONF_SENT = float(os.environ.get("STOP_LOSS_HIGH_CONF_SENT", "0.025"))
STOP_LOSS_MED_CONF_SENT = float(os.environ.get("STOP_LOSS_MED_CONF_SENT", "0.03"))
STOP_LOSS_LOW_CONF_SENT = float(os.environ.get("STOP_LOSS_LOW_CONF_SENT", "0.035"))
ATR_MULT_HIGH_CONF = float(os.environ.get("ATR_MULT_HIGH_CONF", "1.5"))
ATR_MULT_MED_CONF = float(os.environ.get("ATR_MULT_MED_CONF", "2.0"))
ATR_MULT_LOW_CONF = float(os.environ.get("ATR_MULT_LOW_CONF", "2.5"))
INDIAN_MARKET_BUFFER = float(os.environ.get("INDIAN_MARKET_BUFFER", "0.005"))
HIGH_CONFIDENCE_THRESHOLD = float(os.environ.get("HIGH_CONFIDENCE_THRESHOLD", "0.80"))
MEDIUM_CONFIDENCE_THRESHOLD = float(
    os.environ.get("MEDIUM_CONFIDENCE_THRESHOLD", "0.60")
)
AI_STRATEGY_GAP_WARN = float(os.environ.get("AI_STRATEGY_GAP_WARN", "0.10"))
AI_STRATEGY_GAP_BLOCK = float(os.environ.get("AI_STRATEGY_GAP_BLOCK", "0.50"))
ADVISOR_RISK_AVERSION = float(os.environ.get("ADVISOR_RISK_AVERSION", "0.35"))
ADVISOR_MAX_PORTFOLIO_RISK = float(os.environ.get("ADVISOR_MAX_PORTFOLIO_RISK", "0.12"))
ADVISOR_SENTIMENT_INFLUENCE = float(
    os.environ.get("ADVISOR_SENTIMENT_INFLUENCE", "0.10")
)
ADVISOR_MAX_PER_SECTOR = max(1, int(os.environ.get("ADVISOR_MAX_PER_SECTOR", "3")))
ADVISOR_BUY_TRIGGER_BUFFER = float(
    os.environ.get("ADVISOR_BUY_TRIGGER_BUFFER", "0.002")
)
ADVISOR_ENTRY_RANGE_PCT = float(os.environ.get("ADVISOR_ENTRY_RANGE_PCT", "0.003"))
ADVISOR_ENTRY_RANGE_MIN_PCT = float(
    os.environ.get("ADVISOR_ENTRY_RANGE_MIN_PCT", "0.001")
)
ADVISOR_TARGET_STEP_MULTIPLIER = float(
    os.environ.get("ADVISOR_TARGET_STEP_MULTIPLIER", "1.0")
)
ADVISOR_TARGET_STEP_MIN_PCT = float(
    os.environ.get("ADVISOR_TARGET_STEP_MIN_PCT", "0.005")
)
SIM_MAX_PRICE_DEVIATION_FROM_STRATEGY = float(
    os.environ.get("SIM_MAX_PRICE_DEVIATION_FROM_STRATEGY", "0.35")
)
SIM_MAX_PRICE_DEVIATION_FROM_ENTRY = float(
    os.environ.get("SIM_MAX_PRICE_DEVIATION_FROM_ENTRY", "0.35")
)
PORTFOLIO_MAX_TRADE_QTY = int(os.environ.get("PORTFOLIO_MAX_TRADE_QTY", "25000"))
PORTFOLIO_MAX_TRADE_PRICE = float(os.environ.get("PORTFOLIO_MAX_TRADE_PRICE", "200000"))
SIM_PRICE_HARD_MAX = float(os.environ.get("SIM_PRICE_HARD_MAX", "1000000"))
SIM_QTY_HARD_MAX = int(os.environ.get("SIM_QTY_HARD_MAX", "100000"))
TRADE_DAILY_REPORT_HOUR_IST = int(
    np.clip(int(os.environ.get("TRADE_DAILY_REPORT_HOUR_IST", "15")), 0, 23)
)
TRADE_DAILY_REPORT_MINUTE_IST = int(
    np.clip(int(os.environ.get("TRADE_DAILY_REPORT_MINUTE_IST", "45")), 0, 59)
)
TRADE_EMAIL_ENABLED = os.environ.get("TRADE_EMAIL_ENABLED", "false").lower() in (
    "1",
    "true",
    "yes",
)
TRADE_EMAIL_TO = os.environ.get("TRADE_EMAIL_TO", "anthonybreeganzo07@gmail.com")
SMTP_HOST = os.environ.get("SMTP_HOST", "").strip()
try:
    SMTP_PORT = int(os.environ.get("SMTP_PORT", "587"))
except Exception:
    SMTP_PORT = 587
SMTP_USER = os.environ.get("SMTP_USER", "").strip()
SMTP_PASSWORD = os.environ.get("SMTP_PASSWORD", "").strip()
SMTP_FROM = os.environ.get("SMTP_FROM", "").strip() or SMTP_USER
SMTP_USE_SSL = os.environ.get("SMTP_USE_SSL", "false").lower() in ("1", "true", "yes")
SMTP_STARTTLS = os.environ.get("SMTP_STARTTLS", "true").lower() in (
    "1",
    "true",
    "yes",
)
SIM_TRADE_CACHE_MAX_ROWS = max(
    100,
    int(os.environ.get("SIM_TRADE_CACHE_MAX_ROWS", "5000")),
)


def _raise_nofile_soft_limit() -> None:
    """
    Best-effort increase of soft open-files limit for long-running local sessions.
    Helps avoid OSError: [Errno 24] Too many open files under heavy refresh/load.
    """
    try:
        soft, hard = resource.getrlimit(resource.RLIMIT_NOFILE)
        if soft <= 0:
            return
        target = min(max(soft, FD_TARGET_SOFT_LIMIT), hard)
        if target > soft:
            resource.setrlimit(resource.RLIMIT_NOFILE, (target, hard))
            log.info("Raised RLIMIT_NOFILE soft limit: %s -> %s", soft, target)
    except Exception:
        # Non-fatal on environments where setrlimit is restricted.
        pass


def _fd_pressure_high() -> bool:
    """Return True when process file descriptors are near the soft limit."""
    try:
        soft, _ = resource.getrlimit(resource.RLIMIT_NOFILE)
        if soft <= 0:
            return False
        used = len(os.listdir("/dev/fd"))
        ratio = max(0.5, min(FD_SOFT_GUARD_RATIO, 0.98))
        return used >= int(float(soft) * ratio)
    except Exception:
        return False


def _safe_yf_download(*args, **kwargs) -> pd.DataFrame:
    """
    Guarded yfinance download:
    - bounded concurrent calls
    - yfinance internal threads disabled
    - backpressure when FD usage is high
    """
    try:
        import yfinance as yf
    except Exception:
        return pd.DataFrame()

    if _fd_pressure_high():
        log.warning("Skipping yfinance download under high FD pressure")
        return pd.DataFrame()

    if not _yf_download_semaphore.acquire(timeout=YF_CALL_TIMEOUT_SEC):
        log.warning("Skipping yfinance download due to semaphore timeout")
        return pd.DataFrame()
    try:
        kwargs.setdefault("progress", False)
        kwargs.setdefault("threads", False)
        kwargs.setdefault("timeout", YF_CALL_TIMEOUT_SEC)
        try:
            out = yf.download(*args, **kwargs)
        except TypeError:
            kwargs.pop("timeout", None)
            out = yf.download(*args, **kwargs)
        return out if isinstance(out, pd.DataFrame) else pd.DataFrame()
    except Exception as exc:
        log.warning("yfinance download failed: %s", exc)
        return pd.DataFrame()
    finally:
        _yf_download_semaphore.release()


_raise_nofile_soft_limit()


try:
    import openpyxl  # noqa: F401

    OPENPYXL_AVAILABLE = True
except Exception:
    OPENPYXL_AVAILABLE = False


def _load_settings() -> dict:
    """Load config/settings.yaml if available."""
    settings_path = CONFIG_DIR / "settings.yaml"
    if not settings_path.exists():
        return {}
    try:
        with open(settings_path) as f:
            data = yaml.safe_load(f) or {}
        return data if isinstance(data, dict) else {}
    except Exception as e:
        log.warning("Failed to load settings.yaml: %s", e)
        return {}


_settings = _load_settings()
_webapp_settings = (
    _settings.get("webapp", {}) if isinstance(_settings.get("webapp", {}), dict) else {}
)
_market_settings = (
    _settings.get("market", {}) if isinstance(_settings.get("market", {}), dict) else {}
)
PREMARKET_MAX_BUFFER_MINUTES = int(
    os.environ.get(
        "PREMARKET_MAX_BUFFER_MINUTES",
        _webapp_settings.get("premarket_max_buffer_minutes", 30),
    )
)
PREMARKET_MAX_BUFFER_MINUTES = int(np.clip(PREMARKET_MAX_BUFFER_MINUTES, 0, 180))
PREMARKET_DEFAULT_TICKERS = int(_webapp_settings.get("premarket_default_tickers", 10))
PREMARKET_DEFAULT_TICKERS = max(1, PREMARKET_DEFAULT_TICKERS)
NEXT_DAY_PREDICTION_HOUR = int(
    os.environ.get(
        "NEXT_DAY_PREDICTION_HOUR",
        _webapp_settings.get("next_day_prediction_hour", 15),
    )
)
NEXT_DAY_PREDICTION_MINUTE = int(
    os.environ.get(
        "NEXT_DAY_PREDICTION_MINUTE",
        _webapp_settings.get("next_day_prediction_minute", 45),
    )
)
NEXT_DAY_PREDICTION_HOUR = int(np.clip(NEXT_DAY_PREDICTION_HOUR, 0, 23))
NEXT_DAY_PREDICTION_MINUTE = int(np.clip(NEXT_DAY_PREDICTION_MINUTE, 0, 59))
PREMARKET_WINDOW_START_HOUR = int(
    os.environ.get(
        "PREMARKET_WINDOW_START_HOUR",
        _webapp_settings.get("premarket_window_start_hour", 9),
    )
)
PREMARKET_WINDOW_START_MINUTE = int(
    os.environ.get(
        "PREMARKET_WINDOW_START_MINUTE",
        _webapp_settings.get("premarket_window_start_minute", 15),
    )
)
MARKET_LOCK_START_HOUR = int(
    os.environ.get(
        "MARKET_LOCK_START_HOUR",
        _webapp_settings.get("market_lock_start_hour", 9),
    )
)
MARKET_LOCK_START_MINUTE = int(
    os.environ.get(
        "MARKET_LOCK_START_MINUTE",
        _webapp_settings.get("market_lock_start_minute", 30),
    )
)
MARKET_LOCK_END_HOUR = int(
    os.environ.get(
        "MARKET_LOCK_END_HOUR",
        _webapp_settings.get("market_lock_end_hour", 15),
    )
)
MARKET_LOCK_END_MINUTE = int(
    os.environ.get(
        "MARKET_LOCK_END_MINUTE",
        _webapp_settings.get("market_lock_end_minute", 30),
    )
)
PRECOMPUTE_OPEN_HOUR = int(
    os.environ.get(
        "PRECOMPUTE_OPEN_HOUR", _webapp_settings.get("precompute_open_hour", 9)
    )
)
PRECOMPUTE_OPEN_MINUTE = int(
    os.environ.get(
        "PRECOMPUTE_OPEN_MINUTE", _webapp_settings.get("precompute_open_minute", 28)
    )
)
PRECOMPUTE_CLOSE_HOUR = int(
    os.environ.get(
        "PRECOMPUTE_CLOSE_HOUR", _webapp_settings.get("precompute_close_hour", 15)
    )
)
PRECOMPUTE_CLOSE_MINUTE = int(
    os.environ.get(
        "PRECOMPUTE_CLOSE_MINUTE", _webapp_settings.get("precompute_close_minute", 31)
    )
)
PRECOMPUTE_OPEN_HOUR = int(np.clip(PRECOMPUTE_OPEN_HOUR, 0, 23))
PRECOMPUTE_OPEN_MINUTE = int(np.clip(PRECOMPUTE_OPEN_MINUTE, 0, 59))
PRECOMPUTE_CLOSE_HOUR = int(np.clip(PRECOMPUTE_CLOSE_HOUR, 0, 23))
PRECOMPUTE_CLOSE_MINUTE = int(np.clip(PRECOMPUTE_CLOSE_MINUTE, 0, 59))
ALPHA_RISK_FREE_ANNUAL = float(
    os.environ.get(
        "ALPHA_RISK_FREE_ANNUAL", _market_settings.get("risk_free_rate", 0.0)
    )
)
ALPHA_DEFAULT_BETA = float(
    os.environ.get("ALPHA_DEFAULT_BETA", _market_settings.get("default_beta", 1.0))
)

# Thread-safe lock for file I/O
_log_lock = threading.Lock()
_portfolio_lock = threading.Lock()
_portfolio_sim_lock = threading.Lock()
_delisted_lock = threading.Lock()
_sim_trade_cache_lock = threading.Lock()
_daily_trade_report_lock = threading.Lock()

# ---------------------------------------------------------------------------
# Global state (loaded once at startup)
# ---------------------------------------------------------------------------
predictor: LivePredictor | None = None
logger = PredictionLogger()
tickers_by_sector: dict[str, list[str]] = {}
all_tickers: list[str] = []
ticker_names: dict[str, str] = {}  # RELIANCE.NS → Reliance Industries
ticker_to_sector: dict[str, str] = {}
models_loaded = False
load_error = ""
models_loading = False
models_load_started_at: float | None = None
models_load_completed_at: float | None = None
_groq_forecast_cache: dict[str, dict] = {}
_groq_forecast_cache_time: dict[str, float] = {}
_groq_forecast_cache_lock = threading.Lock()
_groq_forecast_cache_loaded = False
GROQ_FORECAST_TTL = 900  # 15m
NEWS_CONTEXT_TTL = max(60, int(os.environ.get("NEWS_CONTEXT_TTL", "300")))
_news_context_cache: dict[str, dict] = {}
_news_context_cache_lock = threading.Lock()
_sim_trade_event_cache: deque = deque(maxlen=SIM_TRADE_CACHE_MAX_ROWS)
_daily_trade_report_dates_done: set[str] = set()
try:
    _groww_cost_calculator = GrowwCostCalculator()
except Exception as _cost_exc:
    _groww_cost_calculator = None
    log.warning("Groww transaction cost model unavailable: %s", _cost_exc)


def _clean_name(ticker: str) -> str:
    """Convert RELIANCE.NS → Reliance"""
    name = ticker.replace(".NS", "").replace(".BO", "")
    # Make it title-case with common replacements
    name = name.replace("_", " ")
    return name


def load_tickers():
    """Load ticker universe from config."""
    global tickers_by_sector, all_tickers, ticker_names, ticker_to_sector
    cfg_path = CONFIG_DIR / "tickers.yaml"
    with open(cfg_path) as f:
        data = yaml.safe_load(f)

    tradeable_sectors = [
        "large_cap",
        "banking",
        "mid_cap",
        "high_volatility",
        "commodities",
    ]
    for sec in tradeable_sectors:
        syms = data.get(sec, [])
        if isinstance(syms, list):
            tickers_by_sector[sec] = syms
            all_tickers.extend(syms)
            for sym in syms:
                ticker_to_sector[str(sym).strip().upper()] = sec

    all_tickers_set = sorted(set(all_tickers))
    all_tickers.clear()
    all_tickers.extend(all_tickers_set)

    for t in all_tickers:
        ticker_names[t] = _clean_name(t)


def load_models_background():
    """Load ML models in background thread so server starts fast."""
    global predictor, models_loaded, load_error, models_loading, models_load_started_at, models_load_completed_at
    try:
        models_loading = True
        models_load_started_at = time.time()
        models_load_completed_at = None
        log.info("Loading ML models...")
        t0 = time.time()
        predictor = LivePredictor()
        predictor.load_models()
        elapsed = time.time() - t0
        models_loaded = True
        models_loading = False
        models_load_completed_at = time.time()
        log.info(f"All models loaded in {elapsed:.1f}s")
    except Exception as e:
        models_loading = False
        load_error = str(e)
        log.error(f"Model loading failed: {e}")


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
SECTOR_DISPLAY = {
    "large_cap": "Large Cap",
    "banking": "Banking & Finance",
    "mid_cap": "Mid Cap Growth",
    "high_volatility": "High Volatility",
    "commodities": "Commodities",
}


def _log_prediction(ticker: str, pred: dict):
    """Log prediction to daily JSON file for end-of-day comparison."""
    now_ist = datetime.now(IST)
    now_iso = now_ist.isoformat()
    today_str = now_ist.strftime("%Y-%m-%d")
    log_file = PREDICTION_LOG_DIR / f"{today_str}.json"
    predicted_return_pct = float(pred.get("predicted_return", 0) or 0)
    current_price = float(pred.get("current_price", 0) or 0)
    open_price = float(
        _opening_prices.get(ticker, {}).get("open", current_price) or current_price
    )
    strategy_price_at_open = (
        round(open_price * (1 + predicted_return_pct / 100.0), 2)
        if open_price > 0
        else float(pred.get("predicted_price", 0) or 0)
    )
    premarket_row = _get_premarket_row_for_ticker(ticker, today_str)
    ai_meta = _resolve_ai_forecast_price(
        ticker,
        open_price=float(open_price or 0),
        strategy_price=float(strategy_price_at_open or 0),
        current_price=float(current_price or 0),
        allow_generate=False,
    )
    ai_last_prediction = float(
        premarket_row.get("ai_predicted_price") or ai_meta.get("price") or 0
    )
    ai_source = str(
        premarket_row.get("ai_source") or ai_meta.get("source", "none") or "none"
    )
    strategy_direction = (
        "UP"
        if strategy_price_at_open > open_price
        else "DOWN" if strategy_price_at_open < open_price else "FLAT"
    )
    ai_direction = (
        (
            "UP"
            if ai_last_prediction > current_price
            else "DOWN" if ai_last_prediction < current_price else "FLAT"
        )
        if ai_last_prediction > 0
        else "N/A"
    )
    strategy_predicted_at_open = _normalize_open_window_timestamp(
        premarket_row.get("strategy_predicted_at_open")
        or premarket_row.get("captured_at"),
        date_hint=today_str,
        default_offset_minutes=5,
    )
    ai_predicted_at_open = _normalize_open_window_timestamp(
        premarket_row.get("ai_predicted_at_open") or premarket_row.get("captured_at"),
        date_hint=today_str,
        default_offset_minutes=7,
    )

    with _log_lock:
        existing = {}
        if log_file.exists():
            try:
                with open(log_file) as f:
                    existing = json.load(f)
            except (json.JSONDecodeError, OSError):
                existing = {}

        existing[ticker] = {
            "predicted_return": pred.get("predicted_return", 0),
            "predicted_price": pred.get("predicted_price", 0),
            "current_price": pred.get("current_price", 0),
            "open_price": round(open_price, 2) if open_price > 0 else 0.0,
            "strategy_price_at_open": strategy_price_at_open,
            "ai_last_prediction": (
                round(ai_last_prediction, 2) if ai_last_prediction > 0 else None
            ),
            "ai_source": ai_source if ai_last_prediction > 0 else "none",
            "strategy_direction_at_open": strategy_direction,
            "ai_direction_last": ai_direction,
            "strategy_vs_ai_direction": (
                strategy_direction == ai_direction if ai_last_prediction > 0 else None
            ),
            "direction_comparison": None,
            "strategy_predicted_at_open": strategy_predicted_at_open,
            "ai_predicted_at_open": ai_predicted_at_open,
            "ai_last_prediction_at": now_iso,
            "signal": pred.get("signal", "HOLD"),
            "confidence": pred.get("confidence", 50),
            "model_predictions": pred.get("model_predictions", {}),
            "timestamp": now_iso,
        }

        with open(log_file, "w") as f:
            json.dump(existing, f, indent=2, default=str)

    # Also record to prediction tracker for hit/miss tracking
    try:
        tracked = dict(pred)
        tracked["open_price"] = round(open_price, 2) if open_price > 0 else 0.0
        tracked["strategy_price_at_open"] = strategy_price_at_open
        tracked["ai_last_prediction"] = (
            round(ai_last_prediction, 2) if ai_last_prediction > 0 else 0.0
        )
        tracked["ai_source"] = ai_source if ai_last_prediction > 0 else "none"
        tracked["strategy_predicted_at_open"] = strategy_predicted_at_open
        tracked["ai_predicted_at_open"] = ai_predicted_at_open
        tracked["ai_last_prediction_at"] = now_iso
        tracked["strategy_direction_at_open"] = strategy_direction
        tracked["ai_direction_last"] = ai_direction
        tracked["snapshot_type"] = _prediction_window_type(now_ist)
        tracked["strategy_vs_ai_direction"] = (
            strategy_direction == ai_direction if ai_last_prediction > 0 else None
        )
        tracked["direction_comparison"] = None
        PredictionTracker.record_prediction(ticker, tracked)
    except Exception as e:
        log.warning(f"Prediction tracking failed for {ticker}: {e}")


def _read_portfolio() -> list[dict]:
    """Read user portfolio entries from cache file."""
    with _portfolio_lock:
        if not PORTFOLIO_FILE.exists():
            return []
        try:
            data = json.loads(PORTFOLIO_FILE.read_text())
            if isinstance(data, list):
                return data
        except Exception:
            pass
        return []


def _write_portfolio(entries: list[dict]) -> None:
    """Persist user portfolio entries to cache file."""
    with _portfolio_lock:
        PORTFOLIO_FILE.parent.mkdir(parents=True, exist_ok=True)
        PORTFOLIO_FILE.write_text(json.dumps(entries, indent=2))


def _read_portfolio_trades() -> list[dict]:
    with _portfolio_lock:
        if not PORTFOLIO_TRADES_FILE.exists():
            return []
        try:
            data = json.loads(PORTFOLIO_TRADES_FILE.read_text())
            if isinstance(data, list):
                return data
        except Exception:
            pass
        return []


def _write_portfolio_trades(entries: list[dict]) -> None:
    with _portfolio_lock:
        PORTFOLIO_TRADES_FILE.parent.mkdir(parents=True, exist_ok=True)
        PORTFOLIO_TRADES_FILE.write_text(json.dumps(entries, indent=2))


def _load_delisted_registry_unlocked() -> dict[str, dict]:
    registry: dict[str, dict] = {}
    if not DELISTED_TICKERS_FILE.exists():
        return registry
    try:
        with open(DELISTED_TICKERS_FILE, newline="") as f:
            reader = csv.DictReader(f)
            for row in reader:
                ticker = str(row.get("ticker", "")).strip().upper()
                if not ticker:
                    continue
                try:
                    hit_count = int(float(row.get("hit_count", 0) or 0))
                except Exception:
                    hit_count = 0
                registry[ticker] = {
                    "ticker": ticker,
                    "first_seen": row.get("first_seen", ""),
                    "last_seen": row.get("last_seen", ""),
                    "hit_count": hit_count,
                    "last_reason": row.get("last_reason", ""),
                    "last_source": row.get("last_source", ""),
                    "status": row.get("status", "watchlist"),
                }
    except Exception as e:
        log.warning(f"Failed to read delisted ticker registry: {e}")
    return registry


def _save_delisted_registry_unlocked(registry: dict[str, dict]) -> None:
    DELISTED_TICKERS_FILE.parent.mkdir(parents=True, exist_ok=True)
    fieldnames = [
        "ticker",
        "first_seen",
        "last_seen",
        "hit_count",
        "last_reason",
        "last_source",
        "status",
    ]
    with open(DELISTED_TICKERS_FILE, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for ticker in sorted(registry):
            row = registry[ticker]
            writer.writerow(
                {
                    "ticker": row.get("ticker", ticker),
                    "first_seen": row.get("first_seen", ""),
                    "last_seen": row.get("last_seen", ""),
                    "hit_count": int(row.get("hit_count", 0) or 0),
                    "last_reason": row.get("last_reason", ""),
                    "last_source": row.get("last_source", ""),
                    "status": row.get("status", "watchlist"),
                }
            )


def _record_unavailable_ticker(
    ticker: str, reason: str, source: str = "yfinance"
) -> None:
    """Track tickers that repeatedly fail to return valid price data."""
    t = str(ticker or "").strip().upper()
    if not t:
        return
    now = datetime.now(IST).strftime("%Y-%m-%d %H:%M:%S")
    with _delisted_lock:
        registry = _load_delisted_registry_unlocked()
        row = registry.get(t)
        if row is None:
            row = {
                "ticker": t,
                "first_seen": now,
                "last_seen": now,
                "hit_count": 1,
                "last_reason": reason,
                "last_source": source,
                "status": "watchlist",
            }
        else:
            row["last_seen"] = now
            row["hit_count"] = int(row.get("hit_count", 0) or 0) + 1
            row["last_reason"] = reason
            row["last_source"] = source
        if int(row.get("hit_count", 0) or 0) >= 3:
            row["status"] = "delisted_candidate"
        elif row.get("status") != "recovered":
            row["status"] = "watchlist"
        registry[t] = row
        _save_delisted_registry_unlocked(registry)


def _mark_ticker_recovered(ticker: str, source: str = "yfinance") -> None:
    """Mark previously unavailable ticker as recovered once data is back."""
    t = str(ticker or "").strip().upper()
    if not t:
        return
    now = datetime.now(IST).strftime("%Y-%m-%d %H:%M:%S")
    with _delisted_lock:
        registry = _load_delisted_registry_unlocked()
        row = registry.get(t)
        if not row:
            return
        row["last_seen"] = now
        row["last_source"] = source
        row["last_reason"] = "price_available"
        row["status"] = "recovered"
        registry[t] = row
        _save_delisted_registry_unlocked(registry)


def _prune_delisted_registry() -> dict:
    """
    Remove noisy batch artifacts from delisted registry.
    Keep:
      - symbols not in configured universe (e.g., accidental dummy entries),
      - entries created from explicit single-ticker failures.
    """
    with _delisted_lock:
        registry = _load_delisted_registry_unlocked()
        if not registry:
            return {"changed": False, "removed": 0, "remaining": 0}

        allowed_universe = set(all_tickers)
        kept: dict[str, dict] = {}
        removed = 0
        for ticker, row in registry.items():
            src = str(row.get("last_source", ""))
            if ticker not in allowed_universe and not ticker.startswith("^"):
                kept[ticker] = row
                continue
            if src == "yfinance_single":
                kept[ticker] = row
                continue
            removed += 1

        changed = len(kept) != len(registry)
        if changed:
            _save_delisted_registry_unlocked(kept)
        return {"changed": changed, "removed": removed, "remaining": len(kept)}


def _is_tradeable_ticker(ticker: str) -> bool:
    t = str(ticker or "").strip().upper()
    if not t:
        return False
    if t in set(all_tickers) or t in ticker_names:
        return True
    # Test/dev fallback before ticker universe is loaded.
    if not all_tickers and (t.endswith(".NS") or t.endswith(".BO")) and len(t) >= 4:
        return True
    return False


def _new_trade_entry(ticker: str, side: str, qty: float, price: float) -> dict:
    return {
        "id": uuid4().hex[:12],
        "ticker": ticker,
        "name": ticker_names.get(ticker, _clean_name(ticker)),
        "side": side,
        "quantity": round(float(qty), 4),
        "price": round(float(price), 2),
        "timestamp": datetime.now(IST).isoformat(),
    }


def _ensure_trade_ids(trades: list[dict]) -> bool:
    """Backfill IDs for older trade rows. Returns True if modified."""
    changed = False
    seen: set[str] = set()
    for tr in trades:
        tid = str(tr.get("id", "")).strip()
        if not tid or tid in seen:
            tr["id"] = uuid4().hex[:12]
            tid = tr["id"]
            changed = True
        seen.add(tid)
    return changed


def _validate_trade_sequence(trades: list[dict]) -> tuple[bool, str]:
    """Simple balance validation: cumulative SELL qty cannot exceed BUY qty per ticker."""
    qty_open: dict[str, float] = {}
    for tr in trades:
        ticker = str(tr.get("ticker", "")).strip().upper()
        side = str(tr.get("side", "")).strip().upper()
        if not ticker or side not in {"BUY", "SELL"}:
            return False, "Invalid trade row"
        try:
            qty = float(tr.get("quantity", 0) or 0)
            price = float(tr.get("price", 0) or 0)
        except Exception:
            return False, f"Invalid quantity/price for {ticker}"
        if qty <= 0 or price <= 0:
            return False, f"Non-positive quantity/price for {ticker}"
        if side == "BUY":
            qty_open[ticker] = qty_open.get(ticker, 0.0) + qty
        else:
            have = qty_open.get(ticker, 0.0)
            if have + 1e-9 < qty:
                return False, f"SELL exceeds open quantity for {ticker}"
            qty_open[ticker] = have - qty
    return True, ""


def _simulation_state_template(cash: float = SIMULATION_DEFAULT_CASH) -> dict:
    start_cash = max(1000.0, float(cash or SIMULATION_DEFAULT_CASH))
    now_iso = datetime.now(IST).isoformat()
    tracker = _daily_tracker_for_state(
        {"initial_cash": start_cash},
        now_ts=now_iso,
    )
    return {
        "schema_version": 1,
        "initial_cash": round(start_cash, 2),
        "cash": round(start_cash, 2),
        "open_positions": {},
        "closed_trades": [],
        "trade_history": [],
        "daily_loss_tracker": tracker,
        "session_started_at": now_iso,
        "updated_at": now_iso,
    }


def _latest_sim_prices_from_ledgers() -> dict[str, dict]:
    """
    Recover latest known simulated trade prices per ticker from persisted ledgers.
    Priority:
      1) JSONL ledger (authoritative event stream)
      2) Excel ledger (fallback when JSONL is unavailable)
    """
    latest: dict[str, dict] = {}

    def _upsert(ticker: str, price: float, ts_value, source: str) -> None:
        t = str(ticker or "").strip().upper()
        px = _safe_float(price)
        if not t or px <= 0:
            return
        ts_dt = _parse_iso_with_ist(ts_value if isinstance(ts_value, str) else None)
        if ts_dt is None:
            if isinstance(ts_value, datetime):
                ts_dt = ts_value if ts_value.tzinfo else ts_value.replace(tzinfo=IST)
                ts_dt = ts_dt.astimezone(IST)
            else:
                ts_dt = datetime.now(IST)
        prev = latest.get(t)
        prev_ts = _parse_iso_with_ist(prev.get("timestamp")) if prev else None
        if prev is None or prev_ts is None or ts_dt >= prev_ts:
            latest[t] = {
                "price": round(px, 2),
                "timestamp": ts_dt.isoformat(),
                "source": source,
            }

    if SIMULATED_TRADE_LOG_FILE.exists():
        try:
            with open(SIMULATED_TRADE_LOG_FILE) as f:
                for line in f:
                    line = line.strip()
                    if not line:
                        continue
                    try:
                        raw = json.loads(line)
                    except Exception:
                        continue
                    row = _normalize_sim_trade_event(raw)
                    action = str(row.get("action", "")).upper()
                    if action not in {"BUY", "SELL"}:
                        continue
                    _upsert(
                        row.get("ticker"),
                        row.get("price"),
                        row.get("timestamp"),
                        "jsonl_ledger",
                    )
        except Exception:
            pass

    if OPENPYXL_AVAILABLE and SIMULATED_TRADE_EXCEL_FILE.exists():
        try:
            from openpyxl import load_workbook

            wb = load_workbook(
                SIMULATED_TRADE_EXCEL_FILE, read_only=True, data_only=True
            )
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                rows = ws.iter_rows(min_row=1, max_row=1, values_only=True)
                header_row = next(rows, None)
                if not header_row:
                    continue
                headers = [str(v or "").strip().lower() for v in header_row]
                try:
                    i_ticker = headers.index("ticker")
                    i_price = headers.index("price")
                except ValueError:
                    continue
                i_action = headers.index("action") if "action" in headers else None
                i_ts = headers.index("timestamp") if "timestamp" in headers else None
                i_date = headers.index("date_ist") if "date_ist" in headers else None
                i_time = headers.index("time_ist") if "time_ist" in headers else None

                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not row:
                        continue
                    action = (
                        str(row[i_action]).strip().upper()
                        if i_action is not None and i_action < len(row)
                        else ""
                    )
                    if action and action not in {"BUY", "SELL"}:
                        continue
                    ticker = row[i_ticker] if i_ticker < len(row) else ""
                    price = row[i_price] if i_price < len(row) else 0
                    ts_val = row[i_ts] if i_ts is not None and i_ts < len(row) else None
                    if (
                        ts_val in (None, "")
                        and i_date is not None
                        and i_time is not None
                    ):
                        d_val = row[i_date] if i_date < len(row) else ""
                        t_val = row[i_time] if i_time < len(row) else ""
                        ts_val = f"{d_val}T{t_val}"
                    _upsert(ticker, price, ts_val, "excel_ledger")
            wb.close()
        except Exception:
            pass

    return latest


def _read_portfolio_sim_state() -> dict:
    with _portfolio_sim_lock:
        if not PORTFOLIO_SIM_FILE.exists():
            state = _simulation_state_template()
            PORTFOLIO_SIM_FILE.parent.mkdir(parents=True, exist_ok=True)
            PORTFOLIO_SIM_FILE.write_text(json.dumps(state, indent=2, default=str))
            return state
        try:
            state = json.loads(PORTFOLIO_SIM_FILE.read_text())
            if isinstance(state, dict):
                state.setdefault("schema_version", 1)
                state.setdefault("initial_cash", SIMULATION_DEFAULT_CASH)
                state.setdefault(
                    "cash", state.get("initial_cash", SIMULATION_DEFAULT_CASH)
                )
                state.setdefault("open_positions", {})
                state.setdefault("closed_trades", [])
                state.setdefault("trade_history", [])
                state.setdefault(
                    "session_started_at",
                    datetime.now(IST).isoformat(),
                )
                state["daily_loss_tracker"] = _daily_tracker_for_state(state)
                open_positions = dict(state.get("open_positions", {}))
                cleaned_open: dict[str, dict] = {}
                need_price_backfill = False
                for ticker, pos in open_positions.items():
                    t = str(ticker or "").strip().upper()
                    if not t or not isinstance(pos, dict):
                        continue
                    qty = _safe_float(pos.get("quantity"))
                    avg = _safe_float(pos.get("avg_entry_price"))
                    if qty <= 0 or avg <= 0:
                        continue
                    if _safe_float(pos.get("last_price")) <= 0:
                        need_price_backfill = True
                    cleaned_open[t] = dict(pos)
                latest_prices = (
                    _latest_sim_prices_from_ledgers()
                    if cleaned_open and need_price_backfill
                    else {}
                )
                for t, pos in cleaned_open.items():
                    avg = _safe_float(pos.get("avg_entry_price"))
                    lp = _safe_float(pos.get("last_price"))
                    if lp <= 0:
                        from_logs = latest_prices.get(t, {})
                        lp = _safe_float(from_logs.get("price"))
                        if lp > 0:
                            pos["last_price_source"] = str(
                                from_logs.get("source", "ledger")
                            )
                            pos["last_price_at"] = str(
                                from_logs.get("timestamp")
                                or datetime.now(IST).isoformat()
                            )
                    if lp <= 0:
                        lp = avg
                        pos.setdefault("last_price_source", "entry_fallback")
                        pos.setdefault("last_price_at", datetime.now(IST).isoformat())
                    pos["last_price"] = round(lp, 2)
                    pos["quantity"] = round(_safe_float(pos.get("quantity")), 4)
                    pos["avg_entry_price"] = round(avg, 2)
                state["open_positions"] = cleaned_open
                return state
        except Exception:
            pass
        return _simulation_state_template()


def _write_portfolio_sim_state(state: dict) -> None:
    with _portfolio_sim_lock:
        payload = dict(state or {})
        payload["daily_loss_tracker"] = _daily_tracker_for_state(payload)
        payload["updated_at"] = datetime.now(IST).isoformat()
        open_positions = dict(payload.get("open_positions", {}))
        cleaned_open: dict[str, dict] = {}
        for ticker, pos in open_positions.items():
            t = str(ticker or "").strip().upper()
            if not t or not isinstance(pos, dict):
                continue
            qty = _safe_float(pos.get("quantity"))
            avg = _safe_float(pos.get("avg_entry_price"))
            if qty <= 0 or avg <= 0:
                continue
            row = dict(pos)
            row["quantity"] = round(qty, 4)
            row["avg_entry_price"] = round(avg, 2)
            last_px = _safe_float(row.get("last_price")) or avg
            row["last_price"] = round(last_px, 2)
            row["last_price_at"] = str(
                row.get("last_price_at") or payload["updated_at"]
            )
            cleaned_open[t] = row
        payload["open_positions"] = cleaned_open
        PORTFOLIO_SIM_FILE.parent.mkdir(parents=True, exist_ok=True)
        PORTFOLIO_SIM_FILE.write_text(json.dumps(payload, indent=2, default=str))


def _normalize_sim_trade_event(event: dict) -> dict:
    """Normalize one simulated-trade event for log/cache/export consistency."""
    row = dict(event or {})
    ts = str(row.get("timestamp") or datetime.now(IST).isoformat())
    try:
        ts_dt = datetime.fromisoformat(ts)
        if ts_dt.tzinfo is None:
            ts_dt = ts_dt.replace(tzinfo=IST)
        ts_dt = ts_dt.astimezone(IST)
    except Exception:
        ts_dt = datetime.now(IST)
        ts = ts_dt.isoformat()

    action = str(row.get("action", "")).strip().upper()
    ticker = str(row.get("ticker", "")).strip().upper()
    qty = _safe_float(row.get("quantity"))
    price = _safe_float(row.get("price"))
    notional = _safe_float(row.get("notional"))
    fee = max(0.0, _safe_float(row.get("fee")))
    total_paid = round(notional + fee, 2) if action == "BUY" else 0.0
    total_received = round(max(0.0, notional - fee), 2) if action == "SELL" else 0.0
    net_cash_impact = round(total_received - total_paid, 2)

    normalized = {
        **row,
        "timestamp": ts_dt.isoformat(),
        "date_ist": ts_dt.strftime("%Y-%m-%d"),
        "time_ist": ts_dt.strftime("%H:%M:%S"),
        "action": action,
        "ticker": ticker,
        "quantity": round(qty, 4) if qty > 0 else 0.0,
        "price": round(price, 2) if price > 0 else 0.0,
        "notional": round(notional, 2) if notional > 0 else 0.0,
        "fee": round(fee, 2),
        "total_paid": total_paid,
        "total_received": total_received,
        "net_cash_impact": net_cash_impact,
        "source": str(
            row.get("source", "strategy_simulation") or "strategy_simulation"
        ),
    }
    return normalized


def _append_trade_event_to_csv(event: dict) -> None:
    fieldnames = [
        "id",
        "timestamp",
        "date_ist",
        "time_ist",
        "action",
        "ticker",
        "quantity",
        "price",
        "notional",
        "fee",
        "total_paid",
        "total_received",
        "net_cash_impact",
        "trade_type",
        "reason",
        "realized_pnl",
        "cash_after",
        "source",
        "auto",
    ]
    SIMULATED_TRADE_CSV_FILE.parent.mkdir(parents=True, exist_ok=True)
    exists = SIMULATED_TRADE_CSV_FILE.exists()
    with open(SIMULATED_TRADE_CSV_FILE, "a", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        if not exists:
            writer.writeheader()
        writer.writerow({k: event.get(k) for k in fieldnames})


def _append_trade_event_to_excel(event: dict) -> None:
    """Append one event to .xlsx ledger using one sheet per date (best effort)."""
    if not OPENPYXL_AVAILABLE:
        return
    try:
        from openpyxl import Workbook, load_workbook

        fieldnames = [
            "id",
            "timestamp",
            "date_ist",
            "time_ist",
            "action",
            "ticker",
            "quantity",
            "price",
            "notional",
            "fee",
            "total_paid",
            "total_received",
            "net_cash_impact",
            "trade_type",
            "reason",
            "realized_pnl",
            "cash_after",
            "source",
            "auto",
        ]
        sheet_name = str(
            event.get("date_ist") or datetime.now(IST).strftime("%Y-%m-%d")
        )[:31]

        SIMULATED_TRADE_EXCEL_FILE.parent.mkdir(parents=True, exist_ok=True)
        if SIMULATED_TRADE_EXCEL_FILE.exists():
            wb = load_workbook(SIMULATED_TRADE_EXCEL_FILE)
        else:
            wb = Workbook()

        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.create_sheet(title=sheet_name)
            ws.append(fieldnames)

        if ws.max_row < 1:
            ws.append(fieldnames)
        elif ws.max_row == 1:
            first = [str(c.value or "") for c in ws[1]]
            if first != fieldnames:
                ws.append(fieldnames)

        ws.append([event.get(k) for k in fieldnames])

        if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
            default_ws = wb["Sheet"]
            if (
                default_ws.max_row == 1
                and default_ws.max_column == 1
                and default_ws["A1"].value is None
            ):
                wb.remove(default_ws)

        wb.save(SIMULATED_TRADE_EXCEL_FILE)
    except Exception as exc:
        log.warning("Failed to update Excel trade ledger: %s", exc)


def _smtp_ready() -> bool:
    return bool(
        TRADE_EMAIL_ENABLED
        and SMTP_HOST
        and SMTP_PORT > 0
        and SMTP_FROM
        and TRADE_EMAIL_TO
    )


def _send_email(
    subject: str, body: str, recipient: str | None = None
) -> tuple[bool, str]:
    to_addr = (recipient or TRADE_EMAIL_TO or "").strip()
    if not to_addr:
        return False, "No recipient configured"
    if not _smtp_ready():
        return False, "SMTP is not configured"
    try:
        msg = EmailMessage()
        msg["Subject"] = subject
        msg["From"] = SMTP_FROM
        msg["To"] = to_addr
        msg.set_content(body)

        if SMTP_USE_SSL:
            with smtplib.SMTP_SSL(SMTP_HOST, SMTP_PORT, timeout=30) as smtp:
                if SMTP_USER and SMTP_PASSWORD:
                    smtp.login(SMTP_USER, SMTP_PASSWORD)
                smtp.send_message(msg)
        else:
            with smtplib.SMTP(SMTP_HOST, SMTP_PORT, timeout=30) as smtp:
                if SMTP_STARTTLS:
                    smtp.starttls()
                if SMTP_USER and SMTP_PASSWORD:
                    smtp.login(SMTP_USER, SMTP_PASSWORD)
                smtp.send_message(msg)
        return True, "sent"
    except Exception as exc:
        return False, str(exc)


def _send_trade_notification_async(event: dict) -> None:
    if str(event.get("action", "")).upper() not in {"BUY", "SELL"}:
        return
    if not TRADE_EMAIL_ENABLED:
        return

    def _send():
        subject = (
            f"[Trading Bot] {event.get('action')} {event.get('ticker')} "
            f"qty={event.get('quantity')} @ ₹{event.get('price')}"
        )
        body = (
            f"Trade executed (simulation).\n\n"
            f"Timestamp (IST): {event.get('timestamp')}\n"
            f"Action: {event.get('action')}\n"
            f"Ticker: {event.get('ticker')}\n"
            f"Quantity: {event.get('quantity')}\n"
            f"Price: ₹{event.get('price')}\n"
            f"Notional: ₹{event.get('notional')}\n"
            f"Fee: ₹{event.get('fee')}\n"
            f"Total Paid: ₹{event.get('total_paid')}\n"
            f"Total Received: ₹{event.get('total_received')}\n"
            f"Reason: {event.get('reason')}\n"
            f"Source: {event.get('source')}\n"
            f"Cash After: ₹{event.get('cash_after')}\n"
        )
        ok, msg = _send_email(subject, body)
        if not ok:
            log.warning("Trade notification email failed: %s", msg)

    threading.Thread(target=_send, daemon=True).start()


def _read_sim_trade_events(date_str: str | None = None) -> list[dict]:
    rows: list[dict] = []
    if not SIMULATED_TRADE_LOG_FILE.exists():
        return rows
    try:
        with open(SIMULATED_TRADE_LOG_FILE) as f:
            for line in f:
                line = line.strip()
                if not line:
                    continue
                try:
                    raw = json.loads(line)
                except Exception:
                    continue
                row = _normalize_sim_trade_event(raw)
                if date_str and str(row.get("date_ist")) != str(date_str):
                    continue
                rows.append(row)
    except Exception as exc:
        log.warning("Failed reading simulated trade log: %s", exc)
    return rows


def _warm_sim_trade_cache_from_log() -> None:
    """Best-effort bootstrapping of in-memory trade cache from JSONL ledger."""
    rows = _read_sim_trade_events()
    if not rows:
        return
    tail = rows[-SIM_TRADE_CACHE_MAX_ROWS:]
    with _sim_trade_cache_lock:
        _sim_trade_event_cache.clear()
        for row in tail:
            _sim_trade_event_cache.append(row)


def _parse_iso_with_ist(value: str | None) -> datetime | None:
    try:
        if not value:
            return None
        dt = datetime.fromisoformat(str(value))
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=IST)
        return dt.astimezone(IST)
    except Exception:
        return None


def _build_transaction_summary(
    date_str: str | None = None,
    *,
    since_ts: str | None = None,
) -> dict:
    target_date = str(date_str or datetime.now(IST).strftime("%Y-%m-%d"))
    rows = _read_sim_trade_events(target_date)
    since_dt = _parse_iso_with_ist(since_ts)
    if since_dt is not None:
        rows = [
            r
            for r in rows
            if (
                _parse_iso_with_ist(r.get("timestamp"))
                or datetime.min.replace(tzinfo=IST)
            )
            >= since_dt
        ]
    with _sim_trade_cache_lock:
        cache_rows = [
            r for r in list(_sim_trade_event_cache) if r.get("date_ist") == target_date
        ]
    trade_rows = [r for r in rows if r.get("action") in {"BUY", "SELL"}]
    buy_rows = [r for r in trade_rows if r.get("action") == "BUY"]
    sell_rows = [r for r in trade_rows if r.get("action") == "SELL"]
    buy_notional = sum(_safe_float(r.get("notional")) for r in buy_rows)
    sell_notional = sum(_safe_float(r.get("notional")) for r in sell_rows)
    buy_fees = sum(_safe_float(r.get("fee")) for r in buy_rows)
    sell_fees = sum(_safe_float(r.get("fee")) for r in sell_rows)
    total_fees = buy_fees + sell_fees
    total_paid_buy = buy_notional + buy_fees
    total_received_sell = max(0.0, sell_notional - sell_fees)
    realized_pnl = sum(_safe_float(r.get("realized_pnl")) for r in sell_rows)

    return {
        "date": target_date,
        "total_events": len(rows),
        "total_transactions": len(trade_rows),
        "buy_transactions": len(buy_rows),
        "sell_transactions": len(sell_rows),
        "buy_notional": round(buy_notional, 2),
        "sell_notional": round(sell_notional, 2),
        "buy_transaction_cost": round(buy_fees, 2),
        "sell_transaction_cost": round(sell_fees, 2),
        "total_transaction_cost": round(total_fees, 2),
        "total_paid_for_buys": round(total_paid_buy, 2),
        "total_received_from_sells": round(total_received_sell, 2),
        "net_cash_flow": round(total_received_sell - total_paid_buy, 2),
        "realized_pnl": round(realized_pnl, 2),
        "excel_ledger_file": str(SIMULATED_TRADE_EXCEL_FILE),
        "csv_ledger_file": str(SIMULATED_TRADE_CSV_FILE),
        "in_memory_cache_rows": len(cache_rows),
        "events": trade_rows[-100:],
    }


def _minutes_between(start_iso: str | None, end_iso: str | None) -> int | None:
    try:
        if not start_iso or not end_iso:
            return None
        start_dt = datetime.fromisoformat(str(start_iso))
        end_dt = datetime.fromisoformat(str(end_iso))
        if start_dt.tzinfo is None:
            start_dt = start_dt.replace(tzinfo=IST)
        if end_dt.tzinfo is None:
            end_dt = end_dt.replace(tzinfo=IST)
        return max(0, int((end_dt - start_dt).total_seconds() // 60))
    except Exception:
        return None


def _build_sim_trade_ledger(
    date_str: str | None = None,
    *,
    since_ts: str | None = None,
    limit: int = 500,
) -> dict:
    """
    Build human-readable trade ledger rows from BUY/SELL events.
    Rows are either:
    - SOLD: buy lot matched against sell execution (FIFO)
    - HOLD: buy lot still open
    """
    target_date = str(date_str or datetime.now(IST).strftime("%Y-%m-%d"))
    events = _read_sim_trade_events(target_date)
    since_dt = _parse_iso_with_ist(since_ts)
    if since_dt is not None:
        events = [
            e
            for e in events
            if (
                _parse_iso_with_ist(e.get("timestamp"))
                or datetime.min.replace(tzinfo=IST)
            )
            >= since_dt
        ]
    trade_events = [
        e
        for e in events
        if str(e.get("action", "")).upper() in {"BUY", "SELL"}
        and str(e.get("ticker", "")).strip()
        and _safe_float(e.get("quantity")) > 0
    ]

    buy_lots: dict[str, list[dict]] = defaultdict(list)
    ledger_rows: list[dict] = []
    eps = 1e-9

    for ev in trade_events:
        action = str(ev.get("action", "")).upper()
        ticker = str(ev.get("ticker", "")).upper()
        qty = _safe_float(ev.get("quantity"))
        price = _safe_float(ev.get("price"))
        fee = max(0.0, _safe_float(ev.get("fee")))
        if qty <= 0 or price <= 0:
            continue

        if action == "BUY":
            buy_lots[ticker].append(
                {
                    "id": ev.get("id"),
                    "timestamp": ev.get("timestamp"),
                    "date_ist": ev.get("date_ist"),
                    "time_ist": ev.get("time_ist"),
                    "quantity": qty,
                    "remaining_qty": qty,
                    "price": price,
                    "fee_total": fee,
                    "reason": str(ev.get("reason", "buy")),
                    "source": str(ev.get("source", "strategy_simulation")),
                    "stop_loss_price": _safe_float(ev.get("stop_loss_price")),
                    "target_price": _safe_float(ev.get("target_price")),
                    "entry_range_low": _safe_float(ev.get("entry_range_low")),
                    "entry_range_high": _safe_float(ev.get("entry_range_high")),
                }
            )
            continue

        # SELL (FIFO match against open buy lots)
        sell_qty_total = qty
        sell_qty_remaining = qty
        sell_realized_total = _safe_float(ev.get("realized_pnl"))
        lots = buy_lots.get(ticker, [])

        while sell_qty_remaining > eps and lots:
            lot = lots[0]
            lot_rem = _safe_float(lot.get("remaining_qty"))
            if lot_rem <= eps:
                lots.pop(0)
                continue
            matched_qty = min(lot_rem, sell_qty_remaining)
            if matched_qty <= eps:
                break

            buy_fee_alloc = _safe_float(lot.get("fee_total")) * (
                matched_qty / max(_safe_float(lot.get("quantity")), eps)
            )
            sell_fee_alloc = fee * (matched_qty / max(sell_qty_total, eps))
            buy_notional = matched_qty * _safe_float(lot.get("price"))
            sell_notional = matched_qty * price

            realized_alloc = sell_realized_total * (
                matched_qty / max(sell_qty_total, eps)
            )
            if abs(realized_alloc) <= eps:
                realized_alloc = (
                    (price - _safe_float(lot.get("price"))) * matched_qty
                    - buy_fee_alloc
                    - sell_fee_alloc
                )

            buy_ts = lot.get("timestamp")
            sell_ts = ev.get("timestamp")
            hold_minutes = _minutes_between(buy_ts, sell_ts)

            ledger_rows.append(
                {
                    "ticker": ticker,
                    "name": ticker_names.get(ticker, _clean_name(ticker)),
                    "status": "SOLD",
                    "quantity": round(matched_qty, 4),
                    "buy_trade_id": lot.get("id"),
                    "buy_timestamp": buy_ts,
                    "buy_date_ist": lot.get("date_ist"),
                    "buy_time_ist": lot.get("time_ist"),
                    "buy_price": round(_safe_float(lot.get("price")), 2),
                    "buy_fee": round(buy_fee_alloc, 2),
                    "buy_total_paid": round(buy_notional + buy_fee_alloc, 2),
                    "sell_trade_id": ev.get("id"),
                    "sell_timestamp": sell_ts,
                    "sell_date_ist": ev.get("date_ist"),
                    "sell_time_ist": ev.get("time_ist"),
                    "sell_price": round(price, 2),
                    "sell_fee": round(sell_fee_alloc, 2),
                    "sell_total_received": round(
                        max(0.0, sell_notional - sell_fee_alloc), 2
                    ),
                    "realized_pnl": round(realized_alloc, 2),
                    "unrealized_pnl": None,
                    "hold_minutes": hold_minutes,
                    "sell_reason": str(ev.get("reason", "")),
                    "source": str(
                        ev.get("source", lot.get("source", "strategy_simulation"))
                    ),
                    "stop_loss_price": _display_price_or_none(
                        ev.get("stop_loss_price") or lot.get("stop_loss_price")
                    ),
                    "target_price": _display_price_or_none(
                        ev.get("target_price") or lot.get("target_price")
                    ),
                    "entry_range_low": _display_price_or_none(
                        lot.get("entry_range_low")
                    ),
                    "entry_range_high": _display_price_or_none(
                        lot.get("entry_range_high")
                    ),
                }
            )

            lot["remaining_qty"] = round(lot_rem - matched_qty, 6)
            sell_qty_remaining = round(sell_qty_remaining - matched_qty, 6)
            if lot["remaining_qty"] <= eps:
                lots.pop(0)

    # Remaining lots are HOLD rows (still open)
    state = _read_portfolio_sim_state()
    open_positions_state = {
        str(t).upper(): dict(v or {})
        for t, v in dict(state.get("open_positions", {})).items()
    }
    now_iso = datetime.now(IST).isoformat()

    hold_tickers = [
        t
        for t, lots in buy_lots.items()
        if any(_safe_float(lot.get("remaining_qty")) > eps for lot in lots)
    ]
    live_hold = _get_live_prices_batch(hold_tickers) if hold_tickers else {}

    for ticker, lots in buy_lots.items():
        pos = open_positions_state.get(ticker, {})
        current_price = _safe_float(live_hold.get(ticker, {}).get("price"))
        if current_price <= 0:
            current_price = _safe_float(pos.get("last_price"))
        if current_price <= 0:
            current_price = _safe_float(pos.get("avg_entry_price"))
        for lot in lots:
            rem_qty = _safe_float(lot.get("remaining_qty"))
            if rem_qty <= eps:
                continue
            lot_qty = max(_safe_float(lot.get("quantity")), eps)
            lot_price = _safe_float(lot.get("price"))
            buy_fee_alloc = _safe_float(lot.get("fee_total")) * (rem_qty / lot_qty)
            notional = rem_qty * lot_price
            unrealized = None
            if current_price > 0:
                unrealized = (current_price - lot_price) * rem_qty - buy_fee_alloc

            ledger_rows.append(
                {
                    "ticker": ticker,
                    "name": ticker_names.get(ticker, _clean_name(ticker)),
                    "status": "HOLD",
                    "quantity": round(rem_qty, 4),
                    "buy_trade_id": lot.get("id"),
                    "buy_timestamp": lot.get("timestamp"),
                    "buy_date_ist": lot.get("date_ist"),
                    "buy_time_ist": lot.get("time_ist"),
                    "buy_price": round(lot_price, 2),
                    "buy_fee": round(buy_fee_alloc, 2),
                    "buy_total_paid": round(notional + buy_fee_alloc, 2),
                    "sell_trade_id": None,
                    "sell_timestamp": None,
                    "sell_date_ist": None,
                    "sell_time_ist": None,
                    "sell_price": None,
                    "sell_fee": None,
                    "sell_total_received": None,
                    "realized_pnl": None,
                    "unrealized_pnl": (
                        round(unrealized, 2) if unrealized is not None else None
                    ),
                    "hold_minutes": _minutes_between(lot.get("timestamp"), now_iso),
                    "sell_reason": None,
                    "source": str(lot.get("source", "strategy_simulation")),
                    "current_price": _display_price_or_none(current_price),
                    "stop_loss_price": _display_price_or_none(
                        pos.get("stop_loss_price") or lot.get("stop_loss_price")
                    ),
                    "target_price": _display_price_or_none(
                        pos.get("target_price") or lot.get("target_price")
                    ),
                    "entry_range_low": _display_price_or_none(
                        pos.get("entry_range_low") or lot.get("entry_range_low")
                    ),
                    "entry_range_high": _display_price_or_none(
                        pos.get("entry_range_high") or lot.get("entry_range_high")
                    ),
                }
            )

    ledger_rows.sort(
        key=lambda row: str(row.get("buy_timestamp") or ""),
        reverse=True,
    )
    if limit > 0:
        ledger_rows = ledger_rows[: max(1, min(limit, 1000))]

    sold_count = sum(1 for r in ledger_rows if r.get("status") == "SOLD")
    hold_count = sum(1 for r in ledger_rows if r.get("status") == "HOLD")
    realized_total = sum(
        _safe_float(r.get("realized_pnl"))
        for r in ledger_rows
        if r.get("status") == "SOLD"
    )
    unrealized_total = sum(
        _safe_float(r.get("unrealized_pnl"))
        for r in ledger_rows
        if r.get("status") == "HOLD"
    )

    return {
        "date": target_date,
        "count": len(ledger_rows),
        "sold_count": sold_count,
        "hold_count": hold_count,
        "realized_pnl_total": round(realized_total, 2),
        "unrealized_pnl_total": round(unrealized_total, 2),
        "rows": ledger_rows,
    }


def _write_daily_trade_report(date_str: str | None = None) -> dict:
    target_date = str(date_str or datetime.now(IST).strftime("%Y-%m-%d"))
    summary = _build_transaction_summary(target_date)
    report_payload = {
        "generated_at": datetime.now(IST).isoformat(),
        "report_type": "simulation_transaction_summary",
        "schema_version": 1,
        **summary,
    }
    DAILY_TRADE_REPORT_DIR.mkdir(parents=True, exist_ok=True)
    report_file = DAILY_TRADE_REPORT_DIR / f"{target_date}.json"
    report_file.write_text(json.dumps(report_payload, indent=2, default=str))
    report_payload["report_file"] = str(report_file)
    return report_payload


def _log_simulated_trade(event: dict) -> None:
    row = _normalize_sim_trade_event(event)
    try:
        with _log_lock:
            SIMULATED_TRADE_LOG_FILE.parent.mkdir(parents=True, exist_ok=True)
            with open(SIMULATED_TRADE_LOG_FILE, "a") as f:
                f.write(json.dumps(row, default=str) + "\n")
            _append_trade_event_to_csv(row)
            _append_trade_event_to_excel(row)
    except Exception as exc:
        log.warning("Failed to append simulated trade log: %s", exc)
        return

    with _sim_trade_cache_lock:
        _sim_trade_event_cache.append(row)
    _sync_portfolio_from_simulated_trade(row)
    _send_trade_notification_async(row)


def _sync_portfolio_from_simulated_trade(event: dict) -> None:
    """
    Mirror simulation BUY/SELL events into portfolio trade ledger so
    portfolio pages reflect strategy automation without manual re-entry.
    """
    action = str(event.get("action", "")).strip().upper()
    if action not in {"BUY", "SELL"}:
        return
    ticker = str(event.get("ticker", "")).strip().upper()
    qty = _safe_float(event.get("quantity"))
    price = _safe_float(event.get("price"))
    if (
        not ticker
        or qty <= 0
        or price <= 0
        or qty > PORTFOLIO_MAX_TRADE_QTY
        or price > PORTFOLIO_MAX_TRADE_PRICE
    ):
        if ticker and (
            qty > PORTFOLIO_MAX_TRADE_QTY or price > PORTFOLIO_MAX_TRADE_PRICE
        ):
            log.warning(
                "Skipping portfolio sync for %s due extreme qty/price (qty=%s price=%s)",
                ticker,
                qty,
                price,
            )
        return

    try:
        trades = _read_portfolio_trades()
        sim_trade_id = str(event.get("id", "")).strip()
        if sim_trade_id and any(
            str(tr.get("sim_trade_id", "")).strip() == sim_trade_id for tr in trades
        ):
            return

        tr = _new_trade_entry(ticker=ticker, side=action, qty=qty, price=price)
        tr["timestamp"] = str(event.get("timestamp") or tr["timestamp"])
        tr["source"] = (
            "strategy_simulation_auto"
            if bool(event.get("auto", False))
            else "strategy_simulation"
        )
        if sim_trade_id:
            tr["sim_trade_id"] = sim_trade_id
        trades.append(tr)

        valid, err = _validate_trade_sequence(trades)
        if not valid:
            log.warning(
                "Skipping portfolio sync for simulated trade %s (%s): %s",
                sim_trade_id or ticker,
                action,
                err,
            )
            return

        _write_portfolio_trades(trades)
        summary = _portfolio_summary_from_trades(trades, include_live_prices=False)
        _write_portfolio(
            [
                {
                    "ticker": row["ticker"],
                    "name": row["name"],
                    "quantity": row["quantity"],
                    "entry_price": row["avg_buy_price"],
                    "updated_at": datetime.now(IST).isoformat(),
                }
                for row in summary["positions"]
            ]
        )
    except Exception as exc:
        log.warning("Failed portfolio sync from simulated trade: %s", exc)


def _is_simulation_portfolio_trade(tr: dict) -> bool:
    src = str(tr.get("source", "")).strip().lower()
    if src.startswith("strategy_simulation"):
        return True
    if str(tr.get("sim_trade_id", "")).strip():
        return True
    return False


def _clear_simulation_history(*, clear_portfolio_sim_trades: bool = True) -> dict:
    removed_files: list[str] = []
    with _log_lock:
        for path in (
            SIMULATED_TRADE_LOG_FILE,
            SIMULATED_TRADE_CSV_FILE,
            SIMULATED_TRADE_EXCEL_FILE,
        ):
            try:
                if path.exists():
                    path.unlink()
                    removed_files.append(str(path))
            except Exception as exc:
                log.warning("Failed removing simulation ledger file %s: %s", path, exc)

    with _sim_trade_cache_lock:
        removed_cache = len(_sim_trade_event_cache)
        _sim_trade_event_cache.clear()

    removed_portfolio_trades = 0
    if clear_portfolio_sim_trades:
        try:
            trades = _read_portfolio_trades()
            kept = [tr for tr in trades if not _is_simulation_portfolio_trade(tr)]
            removed_portfolio_trades = max(0, len(trades) - len(kept))
            if removed_portfolio_trades > 0:
                _write_portfolio_trades(kept)
                summary = _portfolio_summary_from_trades(
                    kept, include_live_prices=False
                )
                _write_portfolio(
                    [
                        {
                            "ticker": row["ticker"],
                            "name": row["name"],
                            "quantity": row["quantity"],
                            "entry_price": row["avg_buy_price"],
                            "updated_at": datetime.now(IST).isoformat(),
                        }
                        for row in summary["positions"]
                    ]
                )
        except Exception as exc:
            log.warning("Failed pruning portfolio simulation trades: %s", exc)

    return {
        "removed_files": removed_files,
        "removed_cache_rows": removed_cache,
        "removed_portfolio_sim_trades": removed_portfolio_trades,
    }


def _infer_trade_type_for_exit(entry_ts: str, exit_ts: str) -> str:
    try:
        in_dt = datetime.fromisoformat(str(entry_ts))
        out_dt = datetime.fromisoformat(str(exit_ts))
        if in_dt.tzinfo is None:
            in_dt = in_dt.replace(tzinfo=IST)
        if out_dt.tzinfo is None:
            out_dt = out_dt.replace(tzinfo=IST)
        return (
            "equity_intraday"
            if in_dt.astimezone(IST).date() == out_dt.astimezone(IST).date()
            else "equity_delivery"
        )
    except Exception:
        return "equity_delivery"


def _estimate_entry_fee(buy_value: float, trade_type: str = "equity_delivery") -> float:
    value = max(0.0, float(buy_value or 0))
    if value <= 0:
        return 0.0
    if _groww_cost_calculator is None:
        return round(value * 0.0008, 2)
    try:
        return round(float(_groww_cost_calculator.buy_cost(value, trade_type).total), 2)
    except Exception:
        return round(value * 0.0008, 2)


def _estimate_exit_fee(sell_value: float, trade_type: str = "equity_delivery") -> float:
    value = max(0.0, float(sell_value or 0))
    if value <= 0:
        return 0.0
    if _groww_cost_calculator is None:
        return round(value * 0.0008, 2)
    try:
        return round(
            float(_groww_cost_calculator.sell_cost(value, trade_type).total), 2
        )
    except Exception:
        return round(value * 0.0008, 2)


def _round_to_tick(price: float, tick: float = 0.05) -> float:
    p = float(price or 0)
    if p <= 0:
        return 0.0
    return round(round(p / tick) * tick, 2)


def _entry_price_range(strategy_price: float) -> tuple[float, float]:
    sp = max(0.0, float(strategy_price or 0))
    if sp <= 0:
        return 0.0, 0.0
    tol_pct = max(ADVISOR_ENTRY_RANGE_MIN_PCT, ADVISOR_ENTRY_RANGE_PCT)
    low = _round_to_tick(sp * (1 - tol_pct))
    high = _round_to_tick(sp * (1 + tol_pct))
    if low <= 0:
        low = _round_to_tick(sp)
    if high < low:
        high = low
    return low, high


def _within_entry_range(price: float, low: float, high: float) -> bool:
    px = float(price or 0)
    lo = float(low or 0)
    hi = float(high or 0)
    if px <= 0 or lo <= 0 or hi <= 0:
        return False
    if hi < lo:
        lo, hi = hi, lo
    return lo - 1e-9 <= px <= hi + 1e-9


def _compute_dynamic_stop_loss_pct(
    *,
    confidence_pct: float,
    atr_pct: float,
    uses_sentiment: bool,
) -> tuple[float, bool]:
    """
    Returns (stop_loss_pct, skip_trade) where stop_loss_pct is decimal (e.g., 0.03).
    """
    conf = float(np.clip((confidence_pct or 0) / 100.0, 0.0, 1.0))
    atr_dec = max(0.0, float(atr_pct or 0) / 100.0)

    if uses_sentiment:
        if conf >= HIGH_CONFIDENCE_THRESHOLD:
            base = STOP_LOSS_HIGH_CONF_SENT
            atr_mult = ATR_MULT_HIGH_CONF
        elif conf >= MEDIUM_CONFIDENCE_THRESHOLD:
            base = STOP_LOSS_MED_CONF_SENT
            atr_mult = ATR_MULT_MED_CONF
        else:
            base = STOP_LOSS_LOW_CONF_SENT
            atr_mult = ATR_MULT_LOW_CONF
    else:
        if conf >= HIGH_CONFIDENCE_THRESHOLD:
            base = STOP_LOSS_HIGH_CONF
            atr_mult = ATR_MULT_HIGH_CONF
        elif conf >= MEDIUM_CONFIDENCE_THRESHOLD:
            base = STOP_LOSS_MED_CONF
            atr_mult = ATR_MULT_MED_CONF
        else:
            base = STOP_LOSS_LOW_CONF
            atr_mult = ATR_MULT_LOW_CONF

    sl = max(base, atr_mult * atr_dec) + INDIAN_MARKET_BUFFER
    sl = float(np.clip(sl, 0.005, 0.25))
    skip_trade = conf < 0.45
    return sl, skip_trade


def _position_size_for_risk(
    *,
    account_balance: float,
    entry_price: float,
    stop_loss_pct: float,
) -> int:
    bal = max(0.0, float(account_balance or 0))
    ep = max(0.0, float(entry_price or 0))
    sl = max(0.0001, float(stop_loss_pct or 0))
    if bal <= 0 or ep <= 0:
        return 0

    risk_amount = bal * max(0.0, min(0.2, RISK_PER_TRADE))
    qty_by_risk = int(risk_amount // (ep * sl))
    max_notional = bal * max(0.05, min(1.0, MAX_POSITION_SIZE))
    qty_by_position = int(max_notional // ep)
    return max(0, min(qty_by_risk, qty_by_position))


def _daily_tracker_for_state(state: dict, now_ts: str | None = None) -> dict:
    tracker = dict(state.get("daily_loss_tracker", {}))
    now_dt = datetime.fromisoformat(now_ts) if now_ts else datetime.now(IST)
    if now_dt.tzinfo is None:
        now_dt = now_dt.replace(tzinfo=IST)
    today = now_dt.astimezone(IST).strftime("%Y-%m-%d")
    if str(tracker.get("date", "")) != today:
        tracker = {
            "date": today,
            "realized_loss": 0.0,
            "trading_paused": False,
            "paused_reason": "",
            "max_daily_loss_amount": round(
                float(state.get("initial_cash", SIMULATION_DEFAULT_CASH) or 0)
                * MAX_DAILY_LOSS,
                2,
            ),
        }
    return tracker


def _next_recommended_sell_time(
    current_price: float, stop_loss_price: float, target_price: float
) -> str:
    cp = float(current_price or 0)
    sl = float(stop_loss_price or 0)
    tp = float(target_price or 0)
    now = datetime.now(IST)
    if cp <= 0 or sl <= 0 or tp <= 0:
        return "N/A"
    distance = min(abs(cp - sl), abs(tp - cp)) / max(cp, 1e-9)
    if distance <= 0.003:
        return (now + timedelta(minutes=5)).isoformat()
    if distance <= 0.01:
        return (now + timedelta(minutes=15)).isoformat()
    return (now + timedelta(minutes=45)).isoformat()


def _is_realistic_price(price: float) -> bool:
    p = float(price or 0)
    return np.isfinite(p) and 0 < p <= SIM_PRICE_HARD_MAX


def _strategy_action_from_signal(signal: str, predicted_return_pct: float = 0.0) -> str:
    sig = str(signal or "").strip().upper()
    if sig in {"STRONG_BUY", "BUY"}:
        return "BUY"
    if sig in {"STRONG_SELL", "SELL"}:
        return "SELL"
    if predicted_return_pct > 0.2:
        return "BUY"
    if predicted_return_pct < -0.2:
        return "SELL"
    return "HOLD"


def _has_relevant_today_sentiment(row: dict) -> bool:
    articles = int(_safe_float(row.get("sentiment_articles")))
    if articles <= 0:
        return False
    weights = row.get("sentiment_decay_weights", {})
    if isinstance(weights, dict):
        day0 = _safe_float(weights.get("day0", weights.get("0")))
        if day0 > 0:
            return True
        for k, v in weights.items():
            key = str(k).strip().lower()
            if key in {"day0", "0"} and _safe_float(v) > 0:
                return True
    ts = _parse_iso_datetime(row.get("timestamp") or row.get("captured_at"))
    if ts is not None:
        return ts.astimezone(IST).date() == datetime.now(IST).date()
    return True


def _sentiment_adjusted_action(strategy_action: str, row: dict) -> tuple[str, str]:
    """
    Strategy-first decisioning:
      - use strategy signal as base
      - only adjust when same-day sentiment is relevant
    """
    base = str(strategy_action or "HOLD").upper()
    if base not in {"BUY", "SELL", "HOLD"}:
        base = "HOLD"

    if not _has_relevant_today_sentiment(row):
        return base, "strategy_only"

    sent_raw = _safe_float(row.get("weighted_sentiment_raw"))
    # Keep sentiment conservative: only severe negative news can override strategy.
    if sent_raw <= -0.55:
        if base == "BUY":
            return "HOLD", "sentiment_blocked_buy"
        if base == "HOLD":
            return "SELL", "sentiment_downgrade_to_sell"
    elif sent_raw <= -0.30:
        if base == "BUY":
            return "HOLD", "sentiment_downgrade_to_hold"

    return base, "strategy_plus_sentiment"


def _max_affordable_qty(
    available_cash: float,
    entry_price: float,
    trade_type: str = "equity_delivery",
) -> int:
    cash = max(0.0, float(available_cash or 0))
    px = max(0.0, float(entry_price or 0))
    if cash <= 0 or px <= 0:
        return 0
    hi = int(min(SIM_QTY_HARD_MAX, cash // px))
    if hi <= 0:
        return 0
    lo, best = 1, 0
    while lo <= hi:
        mid = (lo + hi) // 2
        notional = mid * px
        fee = _estimate_entry_fee(notional, trade_type=trade_type)
        total = notional + fee
        if total <= cash + 1e-9:
            best = mid
            lo = mid + 1
        else:
            hi = mid - 1
    return max(0, min(best, SIM_QTY_HARD_MAX))


def _build_strategy_buy_candidates(
    rows: list[dict],
    *,
    live_prices: dict[str, dict] | None = None,
    exclude_tickers: set[str] | None = None,
    allow_soft_buy: bool = False,
) -> tuple[list[dict], list[str]]:
    """
    Build strategy-driven BUY candidates with sentiment adjustment metadata.
    """
    candidates: list[dict] = []
    warnings: list[str] = []
    live_prices = live_prices or {}
    exclude = {str(t).upper() for t in (exclude_tickers or set())}

    for raw in rows:
        if not isinstance(raw, dict):
            continue
        ticker = str(raw.get("ticker", "")).strip().upper()
        if not ticker or ticker in exclude:
            continue
        if not _is_tradeable_ticker(ticker):
            continue

        cp = _safe_float(live_prices.get(ticker, {}).get("price"))
        if cp <= 0:
            cp = _safe_float(raw.get("current_price"))
        strategy_price = _safe_float(
            raw.get("strategy_price_at_open")
            or raw.get("target_price")
            or raw.get("predicted_price")
        )
        if strategy_price <= 0:
            strategy_price = cp

        if not _is_realistic_price(cp) or not _is_realistic_price(strategy_price):
            continue

        pred_ret_pct = _normalize_predicted_return_pct(
            raw.get(
                "predicted_return",
                raw.get(
                    "predicted_return_pct",
                    _safe_float(raw.get("predicted_return_decimal")) * 100.0,
                ),
            )
        )
        strategy_signal = _strategy_action_from_signal(
            raw.get("signal"),
            predicted_return_pct=pred_ret_pct,
        )
        effective_action, action_source = _sentiment_adjusted_action(
            strategy_signal,
            raw,
        )
        confidence_pct = float(np.clip(_safe_float(raw.get("confidence", 0)), 0, 100))
        soft_candidate = False
        if effective_action != "BUY":
            if (
                not allow_soft_buy
                or effective_action == "SELL"
                or pred_ret_pct <= -0.10
                or confidence_pct < 35
            ):
                continue
            soft_candidate = True

        model_agreement = float(
            np.clip(_safe_float(raw.get("model_agreement", 0)), 0, 100)
        )
        atr_pct = float(np.clip(abs(_safe_float(raw.get("atr_pct", 0))), 0, 25))
        liq_factor = _safe_float(raw.get("liquidity_factor"))
        if liq_factor <= 0:
            liq_factor = 1.0
        liq_factor = float(np.clip(liq_factor, 0.5, 1.5))
        rr = _safe_float(raw.get("risk_reward"))
        rr = float(np.clip(rr if rr > 0 else 1.3, 0.5, 5.0))
        uses_sentiment = (
            _has_relevant_today_sentiment(raw)
            and abs(_safe_float(raw.get("weighted_sentiment_raw"))) > 0.0
        )
        stop_loss_pct, skip_trade = _compute_dynamic_stop_loss_pct(
            confidence_pct=confidence_pct,
            atr_pct=atr_pct,
            uses_sentiment=uses_sentiment,
        )
        if skip_trade:
            if allow_soft_buy and confidence_pct >= 30 and pred_ret_pct >= 0:
                soft_candidate = True
                warnings.append(
                    f"{ticker} kept as soft candidate: low confidence {confidence_pct:.1f}% under risk policy."
                )
            else:
                warnings.append(
                    f"{ticker} skipped: low confidence {confidence_pct:.1f}% under risk policy."
                )
                continue

        stop_loss = _round_to_tick(strategy_price * (1 - stop_loss_pct))
        if stop_loss <= 0 or stop_loss >= strategy_price:
            stop_loss = _round_to_tick(strategy_price * (1 - STOP_LOSS_MED_CONF))
        target_delta = max(
            abs(strategy_price - stop_loss) * max(rr, 1.0),
            strategy_price * max(pred_ret_pct / 100.0, 0.008),
        )
        target_price = _round_to_tick(strategy_price + target_delta)
        entry_low = _safe_float(raw.get("entry_range_low"))
        entry_high = _safe_float(raw.get("entry_range_high"))
        if entry_low <= 0 or entry_high <= 0:
            entry_low, entry_high = _entry_price_range(strategy_price)
        if entry_high < entry_low:
            entry_low, entry_high = entry_high, entry_low
        # Keep strategy entry anchored inside its configured range for UI/API consistency.
        strategy_px_rounded = round(strategy_price, 2)
        if entry_low <= 0 or entry_low > strategy_px_rounded:
            entry_low = _round_to_tick(strategy_px_rounded)
        if entry_high <= 0 or entry_high < strategy_px_rounded:
            entry_high = _round_to_tick(strategy_px_rounded)
        if entry_high < entry_low:
            entry_high = entry_low

        predicted_edge = max(0.0, pred_ret_pct) / 100.0
        sentiment_edge = (
            float(np.clip(_safe_float(raw.get("weighted_sentiment_raw")), -1, 1))
            if _has_relevant_today_sentiment(raw)
            else 0.0
        )
        strategy_edge_component = (
            predicted_edge
            * (confidence_pct / 100.0)
            * (model_agreement / 100.0)
            * liq_factor
        )
        # Keep sentiment on the same return scale as strategy edge.
        sentiment_component = sentiment_edge * max(predicted_edge, 0.01)
        sentiment_influence = float(np.clip(ADVISOR_SENTIMENT_INFLUENCE, 0.0, 1.0))
        expected_edge = (
            1.0 - sentiment_influence
        ) * strategy_edge_component + sentiment_influence * sentiment_component
        if soft_candidate:
            expected_edge *= 0.90
        risk_metric = max(stop_loss_pct, atr_pct / 100.0, 0.01)
        utility = expected_edge - ADVISOR_RISK_AVERSION * risk_metric

        if utility <= (-0.06 if soft_candidate else -0.02):
            continue

        if abs(pred_ret_pct) > 4.5:
            warnings.append(
                f"{ticker} skipped: projected return {pred_ret_pct:.2f}% above conservative cap."
            )
            continue

        candidates.append(
            {
                "ticker": ticker,
                "name": ticker_names.get(ticker, _clean_name(ticker)),
                "sector": str(ticker_to_sector.get(ticker, "other")),
                "source": "strategy",
                "current_price": round(cp, 2),
                "strategy_price_at_open": round(strategy_price, 2),
                "predicted_return_pct": round(pred_ret_pct, 3),
                "confidence": round(confidence_pct, 1),
                "model_agreement": round(model_agreement, 1),
                "liquidity_factor": round(liq_factor, 3),
                "atr_pct": round(atr_pct, 3),
                "risk_reward": round(rr, 3),
                "stop_loss_pct": round(stop_loss_pct, 6),
                "stop_loss_price": round(stop_loss, 2),
                "target_price": round(target_price, 2),
                "entry_range_low": round(entry_low, 2),
                "entry_range_high": round(entry_high, 2),
                "expected_edge": float(expected_edge),
                "strategy_edge_component": float(strategy_edge_component),
                "sentiment_edge_component": float(sentiment_component),
                "sentiment_influence": round(sentiment_influence, 4),
                "risk_metric": float(risk_metric),
                "objective_utility": float(utility),
                "signal": str(raw.get("signal", "") or strategy_signal),
                "strategy_action": strategy_signal,
                "effective_action": effective_action,
                "action_source": action_source,
                "strategy_price_source": str(
                    raw.get("strategy_price_source") or "strategy_model"
                ),
                "soft_candidate": bool(soft_candidate),
                "weighted_sentiment_raw": round(
                    _safe_float(raw.get("weighted_sentiment_raw")),
                    4,
                ),
                "sentiment_articles": int(_safe_float(raw.get("sentiment_articles"))),
                "sentiment_weight_used": round(
                    _safe_float(raw.get("sentiment_weight_used")),
                    4,
                ),
                "avg_volume_30d": (
                    round(_safe_float(raw.get("avg_volume_30d")), 2)
                    if _safe_float(raw.get("avg_volume_30d")) > 0
                    else None
                ),
                "strategy_generated_at": raw.get("timestamp")
                or raw.get("captured_at")
                or datetime.now(IST).isoformat(),
            }
        )

    candidates.sort(
        key=lambda c: (
            c["objective_utility"],
            c["confidence"],
            c["predicted_return_pct"],
        ),
        reverse=True,
    )
    return candidates, sorted(set(warnings))


def _optimize_candidate_allocations(
    candidates: list[dict],
    *,
    budget: float,
    trade_type: str,
    max_positions: int,
) -> list[dict]:
    """
    Risk-constrained portfolio optimization over strategy BUY candidates.
    Objective: maximize expected utility under budget, risk, and concentration caps.
    """
    if not candidates:
        return []
    budget = max(0.0, float(budget or 0))
    if budget <= 0:
        return []

    n = len(candidates)
    max_positions = max(1, min(max_positions, n))
    utilities = np.array(
        [max(-1.0, float(c["objective_utility"])) for c in candidates], dtype=float
    )
    risks = np.array(
        [max(0.001, float(c["risk_metric"])) for c in candidates], dtype=float
    )
    per_name_cap = np.array(
        [
            min(budget * max(0.1, min(MAX_POSITION_SIZE, 1.0)), budget)
            for _ in candidates
        ],
        dtype=float,
    )
    allocations = np.zeros(n, dtype=float)

    positive = utilities > 0
    if linprog is not None and positive.any():
        c_vec = -utilities
        A_ub = np.vstack([np.ones(n, dtype=float), risks])
        b_ub = np.array(
            [
                budget,
                budget * max(0.02, min(ADVISOR_MAX_PORTFOLIO_RISK, 1.0)),
            ],
            dtype=float,
        )
        bounds = [(0.0, float(per_name_cap[i])) for i in range(n)]
        try:
            res = linprog(c_vec, A_ub=A_ub, b_ub=b_ub, bounds=bounds, method="highs")
            if res.success and res.x is not None:
                allocations = np.maximum(0.0, np.asarray(res.x, dtype=float))
        except Exception:
            allocations = np.zeros(n, dtype=float)

    if allocations.sum() <= 0:
        # Fallback: normalized utility-weighted allocation under concentration cap.
        util = np.maximum(utilities, 0.0)
        if util.sum() <= 0:
            util = np.ones(n, dtype=float)
        weights = util / util.sum()
        allocations = np.minimum(weights * budget, per_name_cap)

    order = sorted(
        range(n),
        key=lambda i: (
            utilities[i],
            candidates[i]["confidence"],
            candidates[i]["predicted_return_pct"],
        ),
        reverse=True,
    )
    picks: list[dict] = []
    remaining = budget
    sector_cap = max(1, min(ADVISOR_MAX_PER_SECTOR, max_positions))
    sector_counts: dict[str, int] = defaultdict(int)
    for idx in order:
        if len(picks) >= max_positions or remaining <= 0:
            break
        c = candidates[idx]
        sector_key = str(c.get("sector") or "other")
        if sector_counts.get(sector_key, 0) >= sector_cap:
            continue
        px = float(c["strategy_price_at_open"])
        if not _is_realistic_price(px):
            continue

        alloc = min(float(allocations[idx]), remaining)
        qty_by_alloc = int(alloc // px) if alloc > 0 else 0
        qty_by_cash = _max_affordable_qty(remaining, px, trade_type=trade_type)
        qty = min(max(qty_by_alloc, 0), qty_by_cash)
        if qty <= 0 and qty_by_cash > 0 and alloc > 0:
            qty = 1
        qty = max(0, min(int(qty), SIM_QTY_HARD_MAX))
        if qty <= 0:
            continue

        while qty > 0:
            notional = qty * px
            fee = _estimate_entry_fee(notional, trade_type=trade_type)
            total = notional + fee
            if total <= remaining + 1e-9:
                break
            qty -= 1
        if qty <= 0:
            continue

        notional = round(qty * px, 2)
        fee = round(_estimate_entry_fee(notional, trade_type=trade_type), 2)
        est_trade_cost = round(notional + fee, 2)
        if est_trade_cost <= 0 or est_trade_cost > remaining + 1e-9:
            continue

        remaining = round(remaining - est_trade_cost, 2)
        picks.append(
            {
                **c,
                "suggested_qty": int(qty),
                "estimated_notional": round(notional, 2),
                "estimated_fee": round(fee, 2),
                "est_trade_cost": round(est_trade_cost, 2),
                "next_recommended_sell_time": _next_recommended_sell_time(
                    c["current_price"],
                    c["stop_loss_price"],
                    c["target_price"],
                ),
                "edge_score": round(c["objective_utility"], 6),
            }
        )
        sector_counts[sector_key] = sector_counts.get(sector_key, 0) + 1

    # If optimizer is too sparse, fill remaining slots with minimum 1-qty positions.
    picked_tickers = {str(r.get("ticker", "")).upper() for r in picks}
    if len(picks) < max_positions and remaining > 0:
        for idx in order:
            if len(picks) >= max_positions or remaining <= 0:
                break
            c = candidates[idx]
            ticker = str(c.get("ticker", "")).upper()
            if not ticker or ticker in picked_tickers:
                continue
            sector_key = str(c.get("sector") or "other")
            if sector_counts.get(sector_key, 0) >= sector_cap:
                continue
            px = float(c["strategy_price_at_open"])
            if not _is_realistic_price(px):
                continue
            qty_cash = _max_affordable_qty(remaining, px, trade_type=trade_type)
            qty = 1 if qty_cash >= 1 else 0
            if qty <= 0:
                continue
            notional = round(qty * px, 2)
            fee = round(_estimate_entry_fee(notional, trade_type=trade_type), 2)
            est_trade_cost = round(notional + fee, 2)
            if est_trade_cost <= 0 or est_trade_cost > remaining + 1e-9:
                continue
            remaining = round(remaining - est_trade_cost, 2)
            picks.append(
                {
                    **c,
                    "suggested_qty": int(qty),
                    "estimated_notional": round(notional, 2),
                    "estimated_fee": round(fee, 2),
                    "est_trade_cost": round(est_trade_cost, 2),
                    "next_recommended_sell_time": _next_recommended_sell_time(
                        c["current_price"],
                        c["stop_loss_price"],
                        c["target_price"],
                    ),
                    "edge_score": round(c["objective_utility"], 6),
                }
            )
            picked_tickers.add(ticker)
            sector_counts[sector_key] = sector_counts.get(sector_key, 0) + 1

    return picks


def _run_sim_auto_check(
    state: dict,
    *,
    auto_buy_enabled: bool,
    now_iso: str | None = None,
    source: str = "strategy_simulation",
) -> dict:
    """
    Execute one simulation auto-check cycle:
      - auto-sell on stop-loss / target hit
      - optional auto-buy on strategy-entry threshold
    """
    state = dict(state or {})
    now_iso = str(now_iso or datetime.now(IST).isoformat())
    state["daily_loss_tracker"] = _daily_tracker_for_state(state, now_ts=now_iso)

    open_positions = dict(state.get("open_positions", {}))
    state["open_positions"] = dict(open_positions)
    tickers = sorted(open_positions.keys())
    live = _get_live_prices_batch(tickers) if tickers else {}
    snapshot = _get_prediction_snapshot(use_latest_stored=True)
    snap_rows = snapshot.get("items", []) if isinstance(snapshot, dict) else []
    snapshot_by_ticker = {}
    for row in snap_rows if isinstance(snap_rows, list) else []:
        if not isinstance(row, dict):
            continue
        ticker = str(row.get("ticker", "")).strip().upper()
        if ticker:
            snapshot_by_ticker[ticker] = row

    events = []
    trailing_updates = []

    for ticker in tickers:
        pos = dict(open_positions.get(ticker, {}))
        qty = _safe_float(pos.get("quantity"))
        if qty <= 0:
            continue
        cp = _safe_float(live.get(ticker, {}).get("price"))
        if not _is_realistic_price(cp):
            continue
        pos["last_price"] = round(cp, 2)
        pos["last_price_at"] = now_iso
        pos["last_price_source"] = "live_quote"
        open_positions[ticker] = pos
        state.setdefault("open_positions", {})[ticker] = pos
        stop_loss = _safe_float(pos.get("stop_loss_price"))
        target = _safe_float(pos.get("target_price"))
        entry = _safe_float(pos.get("avg_entry_price"))
        if entry > 0:
            entry_dev = abs(cp - entry) / max(entry, 1e-9)
            if entry_dev > max(0.05, SIM_MAX_PRICE_DEVIATION_FROM_ENTRY):
                log.warning(
                    "AUTO_CHECK skipped %s due price-entry deviation %.2f%% (cp=%.2f entry=%.2f)",
                    ticker,
                    entry_dev * 100.0,
                    cp,
                    entry,
                )
                continue
        snap_row = snapshot_by_ticker.get(ticker, {})
        pred_ret_pct = _normalize_predicted_return_pct(
            snap_row.get(
                "predicted_return_pct",
                _safe_float(snap_row.get("predicted_return_decimal")) * 100.0,
            )
        )
        strategy_signal = _strategy_action_from_signal(
            snap_row.get("signal"),
            predicted_return_pct=pred_ret_pct,
        )
        effective_action, action_source = _sentiment_adjusted_action(
            strategy_signal,
            snap_row,
        )
        pos["strategy_signal"] = strategy_signal
        pos["strategy_recommendation"] = effective_action
        pos["strategy_recommendation_source"] = action_source

        if entry > 0:
            highest = max(_safe_float(pos.get("highest_price")), cp, entry)
            pos["highest_price"] = round(highest, 2)
            initial_dist = max(entry - stop_loss, entry * 0.01)
            # Optional trailing stop: activate once move >= 1.5x initial stop distance.
            if initial_dist > 0 and cp >= entry + 1.5 * initial_dist:
                proposed_trail = _round_to_tick(highest * 0.985)
                prev_trail = _safe_float(pos.get("trail_stop_price"))
                new_trail = max(proposed_trail, prev_trail, stop_loss)
                if new_trail > stop_loss:
                    pos["trail_stop_price"] = round(new_trail, 2)
                    pos["stop_loss_price"] = round(new_trail, 2)
                    stop_loss = float(new_trail)
                    trailing_updates.append(
                        {
                            "ticker": ticker,
                            "new_stop_loss": round(new_trail, 2),
                            "price": round(cp, 2),
                        }
                    )
            open_positions[ticker] = pos
            state.setdefault("open_positions", {})[ticker] = pos

        reason = None
        if stop_loss > 0 and cp <= stop_loss:
            reason = "auto_stop_loss"
        elif effective_action == "SELL":
            reason = "auto_strategy_sell_signal"
        elif target > 0 and cp >= target:
            # Step target forward and lift stop-loss to preserve gains.
            base_step = max(
                target - max(stop_loss, 0.0),
                entry * max(ADVISOR_TARGET_STEP_MIN_PCT, 0.001),
                0.05,
            )
            step = base_step * max(0.5, ADVISOR_TARGET_STEP_MULTIPLIER)
            new_stop = _round_to_tick(max(stop_loss, target))
            new_target = _round_to_tick(target + step)
            if new_target > target and new_stop >= target:
                pos["stop_loss_price"] = round(new_stop, 2)
                pos["trail_stop_price"] = round(
                    max(_safe_float(pos.get("trail_stop_price")), new_stop), 2
                )
                pos["target_price"] = round(new_target, 2)
                pos["highest_price"] = round(
                    max(_safe_float(pos.get("highest_price")), cp), 2
                )
                open_positions[ticker] = pos
                state.setdefault("open_positions", {})[ticker] = pos
                trailing_updates.append(
                    {
                        "ticker": ticker,
                        "new_stop_loss": round(new_stop, 2),
                        "new_target_price": round(new_target, 2),
                        "price": round(cp, 2),
                        "event": "target_step_up",
                    }
                )
                continue
            reason = "auto_target_hit"
        if not reason:
            continue

        state, event, err = _execute_sim_sell(
            state,
            ticker=ticker,
            quantity=qty,
            price=cp,
            reason=reason,
            timestamp=now_iso,
        )
        if err:
            log.warning("AUTO_CHECK skipped %s: %s", ticker, err)
            continue
        if event:
            events.append(event)
            _log_simulated_trade(
                {
                    **event,
                    "source": source,
                    "auto": True,
                }
            )

    auto_buy_events = []
    tracker = _daily_tracker_for_state(state, now_ts=now_iso)
    if (
        auto_buy_enabled
        and not tracker.get("trading_paused")
        and models_loaded
        and predictor is not None
    ):
        try:
            open_now = {
                str(t).upper()
                for t in dict(state.get("open_positions", {})).keys()
                if str(t).strip()
            }
            with _daily_analysis_lock:
                cached_rows = list((_daily_analysis_cache or {}).get("all_stocks", []))
            base_rows = [
                r
                for r in cached_rows
                if isinstance(r, dict)
                and str(r.get("ticker", "")).strip()
                and _safe_float(r.get("current_price")) > 0
            ]
            if not base_rows:
                base_rows = [r for r in snap_rows if isinstance(r, dict)]
            candidates, _warnings = _build_strategy_buy_candidates(
                base_rows,
                live_prices={},
                exclude_tickers=open_now,
            )
            if not candidates:
                candidates, _warnings = _build_strategy_buy_candidates(
                    base_rows,
                    live_prices={},
                    exclude_tickers=open_now,
                    allow_soft_buy=True,
                )
            max_new_positions = max(1, 10 - len(open_now))
            optimized = _optimize_candidate_allocations(
                candidates,
                budget=_safe_float(state.get("cash")),
                trade_type="equity_delivery",
                max_positions=max_new_positions,
            )
            optimized_tickers = [
                str(row.get("ticker", "")).upper()
                for row in optimized
                if str(row.get("ticker", "")).strip()
            ]
            trigger_live = (
                _get_live_prices_batch(optimized_tickers) if optimized_tickers else {}
            )
            for row in optimized:
                ticker = str(row.get("ticker", "")).upper()
                if not ticker or ticker in state.get("open_positions", {}):
                    continue
                cp = _safe_float(trigger_live.get(ticker, {}).get("price"))
                if cp <= 0:
                    cp = _safe_float(row.get("current_price"))
                strategy_entry = _safe_float(row.get("strategy_price_at_open"))
                if not _is_realistic_price(cp) or not _is_realistic_price(
                    strategy_entry
                ):
                    continue
                strategy_dev = abs(cp - strategy_entry) / max(strategy_entry, 1e-9)
                if strategy_dev > max(0.05, SIM_MAX_PRICE_DEVIATION_FROM_STRATEGY):
                    log.warning(
                        "AUTO_CHECK skipped BUY %s due strategy deviation %.2f%% (cp=%.2f strategy=%.2f)",
                        ticker,
                        strategy_dev * 100.0,
                        cp,
                        strategy_entry,
                    )
                    continue
                entry_low = _safe_float(row.get("entry_range_low"))
                entry_high = _safe_float(row.get("entry_range_high"))
                if entry_low <= 0 or entry_high <= 0:
                    entry_low, entry_high = _entry_price_range(strategy_entry)
                trigger = _within_entry_range(cp, entry_low, entry_high)
                if not trigger:
                    continue
                qty_suggested = int(_safe_float(row.get("suggested_qty")))
                qty_cash = _max_affordable_qty(
                    _safe_float(state.get("cash")),
                    cp,
                    trade_type="equity_delivery",
                )
                qty = max(0, min(qty_suggested, qty_cash, SIM_QTY_HARD_MAX))
                if qty <= 0:
                    continue
                available_cash = _safe_float(state.get("cash"))
                est_notional = qty * cp
                est_fee = _estimate_entry_fee(
                    est_notional,
                    trade_type="equity_delivery",
                )
                if est_notional + est_fee > available_cash + 1e-9:
                    continue
                stop_loss = _safe_float(row.get("stop_loss_price"))
                target = _safe_float(row.get("target_price"))
                rr = max(1.0, _safe_float(row.get("risk_reward")) or 1.5)
                strategy_rec = str(
                    row.get("effective_action") or row.get("strategy_action") or "HOLD"
                ).upper()
                state, buy_event, buy_err = _execute_sim_buy(
                    state,
                    ticker=ticker,
                    quantity=qty,
                    price=cp,
                    strategy_entry_price=strategy_entry,
                    entry_range_low=entry_low,
                    entry_range_high=entry_high,
                    stop_loss_price=stop_loss,
                    target_price=target,
                    risk_reward=rr,
                    trade_type="equity_delivery",
                    timestamp=now_iso,
                    reason="auto_strategy_entry",
                    strategy_recommendation=strategy_rec,
                )
                if buy_err or not buy_event:
                    continue
                auto_buy_events.append(buy_event)
                _log_simulated_trade(
                    {
                        **buy_event,
                        "source": source,
                        "auto": True,
                    }
                )
        except Exception as exc:
            log.warning("AUTO_CHECK auto-buy stage failed: %s", exc)

    return {
        "state": state,
        "events": events,
        "auto_buy_events": auto_buy_events,
        "trailing_updates": trailing_updates,
        "triggered_count": len(events) + len(auto_buy_events),
    }


def _execute_sim_sell(
    state: dict,
    ticker: str,
    quantity: float,
    price: float,
    reason: str = "manual_sell",
    timestamp: str | None = None,
) -> tuple[dict, dict | None, str | None]:
    t = str(ticker or "").strip().upper()
    qty = float(quantity or 0)
    px = float(price or 0)
    ts = timestamp or datetime.now(IST).isoformat()
    if qty <= 0 or px <= 0:
        return state, None, "quantity and price must be > 0"
    if qty > SIM_QTY_HARD_MAX:
        return state, None, f"quantity exceeds hard cap ({SIM_QTY_HARD_MAX})"
    if not _is_realistic_price(px):
        return state, None, "price is outside allowed simulation bounds"
    open_positions = dict(state.get("open_positions", {}))
    pos = dict(open_positions.get(t, {}))
    if not pos:
        return state, None, f"No open simulated position for {t}"
    open_qty = float(pos.get("quantity", 0) or 0)
    if open_qty + 1e-9 < qty:
        return state, None, f"Cannot SELL {qty}; open simulated quantity is {open_qty}"

    entry_price = float(pos.get("avg_entry_price", 0) or 0)
    entry_fees_total = float(pos.get("total_entry_fees", 0) or 0)
    allocated_entry_fees = entry_fees_total * (qty / max(open_qty, 1e-9))
    trade_type = _infer_trade_type_for_exit(pos.get("opened_at", ts), ts)
    sell_value = round(px * qty, 2)
    sell_fee = _estimate_exit_fee(sell_value, trade_type=trade_type)
    realized_pnl = (px - entry_price) * qty - allocated_entry_fees - sell_fee

    tracker = _daily_tracker_for_state(state, now_ts=ts)
    if realized_pnl < 0:
        tracker["realized_loss"] = round(
            float(tracker.get("realized_loss", 0) or 0) + abs(realized_pnl),
            2,
        )
    max_loss_amount = float(
        tracker.get(
            "max_daily_loss_amount",
            float(state.get("initial_cash", SIMULATION_DEFAULT_CASH) or 0)
            * MAX_DAILY_LOSS,
        )
        or 0
    )
    if tracker.get("realized_loss", 0) >= max_loss_amount > 0:
        tracker["trading_paused"] = True
        tracker["paused_reason"] = "Daily loss limit reached"
        log.warning(
            "Daily loss limit reached in simulation. realized_loss=%s limit=%s",
            tracker.get("realized_loss"),
            round(max_loss_amount, 2),
        )
    state["daily_loss_tracker"] = tracker

    cash = float(
        state.get("cash", state.get("initial_cash", SIMULATION_DEFAULT_CASH)) or 0
    )
    state["cash"] = round(cash + sell_value - sell_fee, 2)

    remaining_qty = open_qty - qty
    if remaining_qty <= 1e-9:
        open_positions.pop(t, None)
    else:
        pos["quantity"] = round(remaining_qty, 4)
        pos["total_entry_fees"] = round(
            max(0.0, entry_fees_total - allocated_entry_fees), 2
        )
        pos["last_price"] = round(px, 2)
        pos["last_price_at"] = ts
        pos["last_updated"] = ts
        pos["last_price_source"] = "trade_execution"
        open_positions[t] = pos

    state["open_positions"] = open_positions
    state.setdefault("closed_trades", []).append(
        {
            "id": uuid4().hex[:12],
            "ticker": t,
            "quantity": round(qty, 4),
            "entry_price": round(entry_price, 2),
            "exit_price": round(px, 2),
            "entry_fees_allocated": round(allocated_entry_fees, 2),
            "exit_fee": round(sell_fee, 2),
            "realized_pnl": round(realized_pnl, 2),
            "trade_type": trade_type,
            "reason": reason,
            "closed_at": ts,
        }
    )
    event = {
        "id": uuid4().hex[:12],
        "timestamp": ts,
        "action": "SELL",
        "reason": reason,
        "ticker": t,
        "quantity": round(qty, 4),
        "price": round(px, 2),
        "trade_type": trade_type,
        "notional": round(sell_value, 2),
        "fee": round(sell_fee, 2),
        "realized_pnl": round(realized_pnl, 2),
        "cash_after": round(state["cash"], 2),
        "stop_loss_price": _display_price_or_none(pos.get("stop_loss_price")),
        "target_price": _display_price_or_none(pos.get("target_price")),
        "entry_range_low": _display_price_or_none(pos.get("entry_range_low")),
        "entry_range_high": _display_price_or_none(pos.get("entry_range_high")),
    }
    state.setdefault("trade_history", []).append(event)
    return state, event, None


def _execute_sim_buy(
    state: dict,
    *,
    ticker: str,
    quantity: float,
    price: float,
    strategy_entry_price: float,
    entry_range_low: float = 0.0,
    entry_range_high: float = 0.0,
    stop_loss_price: float,
    target_price: float,
    risk_reward: float,
    trade_type: str,
    timestamp: str,
    reason: str = "manual_buy",
    strategy_recommendation: str = "HOLD",
) -> tuple[dict, dict | None, str | None]:
    t = str(ticker or "").strip().upper()
    qty = float(quantity or 0)
    px = float(price or 0)
    ts = str(timestamp or datetime.now(IST).isoformat())
    if qty <= 0 or px <= 0:
        return state, None, "quantity and price must be > 0"
    if qty > SIM_QTY_HARD_MAX:
        return state, None, f"quantity exceeds hard cap ({SIM_QTY_HARD_MAX})"
    if not _is_realistic_price(px):
        return state, None, "price is outside allowed simulation bounds"

    strategy_entry = float(strategy_entry_price or 0)
    if strategy_entry <= 0:
        strategy_entry = px
    lo = float(entry_range_low or 0)
    hi = float(entry_range_high or 0)
    if lo <= 0 or hi <= 0:
        lo, hi = _entry_price_range(strategy_entry)
    if not _within_entry_range(px, lo, hi):
        return (
            state,
            None,
            f"price {px:.2f} is outside strategy entry range [{lo:.2f}, {hi:.2f}]",
        )

    tracker = _daily_tracker_for_state(state, now_ts=ts)
    if tracker.get("trading_paused"):
        return (
            state,
            None,
            str(tracker.get("paused_reason") or "Trading paused for the day"),
        )

    notional = round(px * qty, 2)
    entry_fee = _estimate_entry_fee(notional, trade_type=trade_type)
    total_debit = round(notional + entry_fee, 2)
    cash = float(
        state.get("cash", state.get("initial_cash", SIMULATION_DEFAULT_CASH)) or 0
    )
    if total_debit > cash + 1e-9:
        return (
            state,
            None,
            f"Insufficient simulated cash. Need ₹{total_debit:.2f}, available ₹{cash:.2f}",
        )

    rr_ratio = float(risk_reward or 0)
    if rr_ratio <= 0:
        rr_ratio = 1.5
    sl_price = float(stop_loss_price or 0)
    if sl_price <= 0:
        sl_price = max(0.01, px * (1 - STOP_LOSS_MED_CONF))
    sl_price = _round_to_tick(sl_price)
    tgt_price = float(target_price or 0)
    if tgt_price <= 0:
        tgt_price = px + max(abs(px - sl_price) * max(rr_ratio, 1.0), px * 0.01)
    tgt_price = _round_to_tick(tgt_price)

    open_positions = dict(state.get("open_positions", {}))
    pos = dict(open_positions.get(t, {}))
    prev_qty = _safe_float(pos.get("quantity"))
    prev_avg = _safe_float(pos.get("avg_entry_price"))
    prev_fees = _safe_float(pos.get("total_entry_fees"))
    new_qty = prev_qty + qty
    avg_entry = ((prev_avg * prev_qty) + (px * qty)) / new_qty if new_qty > 0 else px
    highest_seen = max(_safe_float(pos.get("highest_price", 0)), px, avg_entry)
    trail_stop = _safe_float(pos.get("trail_stop_price", 0))
    strategy_rec = str(strategy_recommendation or "HOLD").strip().upper()
    if strategy_rec not in {"BUY", "SELL", "HOLD"}:
        strategy_rec = "HOLD"
    pos.update(
        {
            "ticker": t,
            "name": ticker_names.get(t, _clean_name(t)),
            "quantity": round(new_qty, 4),
            "avg_entry_price": round(avg_entry, 2),
            "total_entry_fees": round(prev_fees + entry_fee, 2),
            "strategy_entry_price": round(strategy_entry, 2),
            "entry_range_low": round(lo, 2),
            "entry_range_high": round(hi, 2),
            "stop_loss_price": round(sl_price, 2),
            "target_price": round(tgt_price, 2),
            "initial_target_price": round(
                _safe_float(pos.get("initial_target_price")) or tgt_price, 2
            ),
            "trade_type_hint": trade_type,
            "opened_at": pos.get("opened_at") or ts,
            "last_buy_at": ts,
            "last_price": round(px, 2),
            "last_price_at": ts,
            "last_price_source": "trade_execution",
            "highest_price": round(highest_seen, 2),
            "trail_stop_price": round(max(trail_stop, sl_price), 2),
            "source": "strategy",
            "strategy_recommendation": strategy_rec,
            "strategy_recommendation_source": "strategy_signal",
        }
    )
    open_positions[t] = pos
    state["open_positions"] = open_positions
    state["cash"] = round(cash - total_debit, 2)
    state["daily_loss_tracker"] = tracker
    event = {
        "id": uuid4().hex[:12],
        "timestamp": ts,
        "action": "BUY",
        "reason": reason,
        "ticker": t,
        "quantity": round(qty, 4),
        "price": round(px, 2),
        "trade_type": trade_type,
        "notional": round(notional, 2),
        "fee": round(entry_fee, 2),
        "cash_after": round(state["cash"], 2),
        "strategy_entry_price": round(strategy_entry, 2),
        "entry_range_low": round(lo, 2),
        "entry_range_high": round(hi, 2),
        "stop_loss_price": round(sl_price, 2),
        "target_price": round(tgt_price, 2),
        "source": "strategy_simulation",
    }
    state.setdefault("trade_history", []).append(event)
    return state, event, None


def _sanitize_portfolio_storage() -> dict:
    """
    Remove invalid/dummy/unknown portfolio rows and backfill missing trade IDs.
    This prevents placeholders like ABC.NS/B.NS from polluting portfolio logic.
    """
    trades = _read_portfolio_trades()
    holdings = _read_portfolio()
    removed_trade_rows = 0
    removed_holding_rows = 0
    changed = False

    valid_trades: list[dict] = []
    for tr in trades:
        ticker = str(tr.get("ticker", "")).strip().upper()
        side = str(tr.get("side", "")).strip().upper()
        try:
            qty = float(tr.get("quantity", 0) or 0)
            price = float(tr.get("price", 0) or 0)
        except Exception:
            qty, price = 0, 0
        if (
            not _is_tradeable_ticker(ticker)
            or side not in {"BUY", "SELL"}
            or qty <= 0
            or price <= 0
            or qty > PORTFOLIO_MAX_TRADE_QTY
            or price > PORTFOLIO_MAX_TRADE_PRICE
        ):
            removed_trade_rows += 1
            changed = True
            continue
        tr["ticker"] = ticker
        tr["side"] = side
        tr["name"] = ticker_names.get(ticker, _clean_name(ticker))
        valid_trades.append(tr)

    if _ensure_trade_ids(valid_trades):
        changed = True

    ok, _err = _validate_trade_sequence(valid_trades)
    if not ok:
        # Keep only BUY rows if sequence got corrupted beyond recovery.
        rebuilt = [t for t in valid_trades if t.get("side") == "BUY"]
        if len(rebuilt) != len(valid_trades):
            changed = True
            removed_trade_rows += len(valid_trades) - len(rebuilt)
            valid_trades = rebuilt

    valid_holdings: list[dict] = []
    for row in holdings:
        ticker = str(row.get("ticker", "")).strip().upper()
        if not _is_tradeable_ticker(ticker):
            removed_holding_rows += 1
            changed = True
            continue
        row["ticker"] = ticker
        row["name"] = ticker_names.get(ticker, _clean_name(ticker))
        valid_holdings.append(row)

    if changed:
        _write_portfolio_trades(valid_trades)
        summary = _portfolio_summary_from_trades(valid_trades)
        _write_portfolio(
            [
                {
                    "ticker": row["ticker"],
                    "name": row["name"],
                    "quantity": row["quantity"],
                    "entry_price": row["avg_buy_price"],
                    "updated_at": datetime.now(IST).isoformat(),
                }
                for row in summary["positions"]
            ]
        )

    return {
        "changed": changed,
        "removed_trade_rows": removed_trade_rows,
        "removed_holding_rows": removed_holding_rows,
        "remaining_trades": len(valid_trades),
    }


def _portfolio_summary_from_trades(
    trades: list[dict], include_live_prices: bool = True
) -> dict:
    positions: dict[str, dict] = {}
    realized_gross_total = 0.0
    realized_cost_total = 0.0
    realized_intraday_cost_total = 0.0
    realized_delivery_cost_total = 0.0
    open_entry_fee_total = 0.0
    open_exit_fee_total = 0.0
    unrealized_after_cost_total = 0.0
    invested_total = 0.0
    current_after_cost_total = 0.0

    def _trade_day(ts: str) -> date | None:
        try:
            dt = datetime.fromisoformat(str(ts))
            if dt.tzinfo is None:
                dt = dt.replace(tzinfo=IST)
            return dt.astimezone(IST).date()
        except Exception:
            return None

    # FIFO lot accounting for accurate realized pnl
    lots: dict[str, list[dict]] = {}
    ordered_trades = sorted(trades, key=lambda t: str(t.get("timestamp", "")))
    for tr in ordered_trades:
        ticker = tr.get("ticker")
        side = str(tr.get("side", "BUY")).upper()
        qty = float(tr.get("quantity", 0) or 0)
        price = float(tr.get("price", 0) or 0)
        trade_ts = str(tr.get("timestamp", "") or datetime.now(IST).isoformat())
        trade_dt = _trade_day(trade_ts)
        if not ticker or qty <= 0 or price <= 0:
            continue
        lots.setdefault(ticker, [])
        if side == "BUY":
            lots[ticker].append(
                {
                    "qty": qty,
                    "price": price,
                    "timestamp": trade_ts,
                    "trade_date": trade_dt,
                }
            )
        elif side == "SELL":
            remaining = qty
            while remaining > 1e-9 and lots[ticker]:
                lot = lots[ticker][0]
                consume = min(remaining, lot["qty"])
                buy_value = float(lot["price"]) * consume
                sell_value = float(price) * consume
                gross_pnl = sell_value - buy_value
                realized_gross_total += gross_pnl

                trade_type = "equity_delivery"
                lot_day = lot.get("trade_date")
                if lot_day is not None and trade_dt is not None and lot_day == trade_dt:
                    trade_type = "equity_intraday"

                txn_cost = 0.0
                if _groww_cost_calculator is not None:
                    try:
                        cost = _groww_cost_calculator.round_trip_cost(
                            buy_value=buy_value,
                            sell_value=sell_value,
                            trade_type=trade_type,
                        )
                        txn_cost = float(cost.total)
                    except Exception:
                        txn_cost = 0.0
                realized_cost_total += txn_cost
                if trade_type == "equity_intraday":
                    realized_intraday_cost_total += txn_cost
                else:
                    realized_delivery_cost_total += txn_cost

                lot["qty"] -= consume
                remaining -= consume
                if lot["qty"] <= 1e-9:
                    lots[ticker].pop(0)

    now_ist = datetime.now(IST)
    today_ist = now_ist.date()
    now_iso = now_ist.isoformat()
    for ticker, ticker_lots in lots.items():
        qty = sum(l["qty"] for l in ticker_lots)
        if qty <= 1e-9:
            continue
        cost = sum(l["qty"] * l["price"] for l in ticker_lots)
        avg_price = cost / qty
        intraday_hint = bool(ticker_lots) and all(
            l.get("trade_date") == today_ist for l in ticker_lots
        )
        trade_type_hint = "equity_intraday" if intraday_hint else "equity_delivery"
        entry_fee_estimate = 0.0
        for lot in ticker_lots:
            lot_qty = float(lot.get("qty", 0) or 0)
            lot_px = float(lot.get("price", 0) or 0)
            lot_value = lot_qty * lot_px
            if lot_value <= 0:
                continue
            lot_ts = str(lot.get("timestamp", "") or now_iso)
            lot_trade_type = _infer_trade_type_for_exit(lot_ts, now_iso)
            if lot_trade_type not in {"equity_intraday", "equity_delivery"}:
                lot_trade_type = trade_type_hint
            entry_fee_estimate += _estimate_entry_fee(
                lot_value,
                trade_type=lot_trade_type,
            )
        positions[ticker] = {
            "ticker": ticker,
            "name": ticker_names.get(ticker, _clean_name(ticker)),
            "quantity": round(qty, 4),
            "avg_buy_price": round(avg_price, 2),
            "cost_value": round(cost, 2),
            "entry_fee_estimate": round(entry_fee_estimate, 2),
            "trade_type_hint": trade_type_hint,
        }

    if include_live_prices and positions:
        live = _get_live_prices_batch(list(positions.keys()))
    else:
        live = {}
    sim_last_prices: dict[str, float] = {}
    try:
        sim_open = dict(_read_portfolio_sim_state().get("open_positions", {}))
        for t, pos in sim_open.items():
            px = _safe_float((pos or {}).get("last_price"))
            if px > 0:
                sim_last_prices[str(t).upper()] = px
    except Exception:
        sim_last_prices = {}
    trade_last_prices: dict[str, float] = {}
    for tr in ordered_trades:
        t = str(tr.get("ticker", "")).upper()
        px = _safe_float(tr.get("price"))
        if t and px > 0:
            trade_last_prices[t] = px

    unrealized_total = 0.0
    for ticker, row in positions.items():
        l = live.get(ticker, {})
        curr = float(l.get("price", 0) or 0)
        if curr <= 0:
            curr = _safe_float(sim_last_prices.get(ticker))
        if curr <= 0:
            curr = _safe_float(trade_last_prices.get(ticker))
        if curr <= 0:
            curr = _safe_float(row.get("avg_buy_price"))
        row["current_price"] = round(curr, 2)
        m2m = curr * row["quantity"] if curr > 0 else 0.0
        entry_fee_estimate = _safe_float(row.get("entry_fee_estimate"))
        trade_type_hint = str(row.get("trade_type_hint") or "equity_delivery")
        exit_fee_estimate = (
            _estimate_exit_fee(m2m, trade_type=trade_type_hint) if m2m > 0 else 0.0
        )
        cost_eaten_open = entry_fee_estimate + exit_fee_estimate
        current_after_cost = max(0.0, m2m - exit_fee_estimate)
        invested_with_cost = row["cost_value"] + entry_fee_estimate
        unrealized_after_cost = current_after_cost - invested_with_cost

        row["market_value"] = round(m2m, 2)
        row["invested_value_after_cost"] = round(invested_with_cost, 2)
        row["current_value_after_cost"] = round(current_after_cost, 2)
        row["transaction_cost_eaten"] = round(cost_eaten_open, 2)
        row["exit_fee_estimate"] = round(exit_fee_estimate, 2)
        row["unrealized_pnl"] = round(m2m - row["cost_value"], 2)
        row["unrealized_pnl_after_cost"] = round(unrealized_after_cost, 2)
        row["unrealized_pnl_pct"] = (
            round(((m2m - row["cost_value"]) / row["cost_value"] * 100), 3)
            if row["cost_value"] > 0 and m2m > 0
            else 0.0
        )
        row["unrealized_pnl_after_cost_pct"] = (
            round((unrealized_after_cost / invested_with_cost * 100), 3)
            if invested_with_cost > 0 and m2m > 0
            else 0.0
        )
        unrealized_total += m2m - row["cost_value"]
        unrealized_after_cost_total += unrealized_after_cost
        invested_total += row["cost_value"]
        current_after_cost_total += current_after_cost
        open_entry_fee_total += entry_fee_estimate
        open_exit_fee_total += exit_fee_estimate

    realized_net_total = realized_gross_total - realized_cost_total
    return {
        "positions": sorted(
            positions.values(), key=lambda x: x["cost_value"], reverse=True
        ),
        "position_count": len(positions),
        "invested_value": round(invested_total, 2),
        "current_value_after_cost": round(current_after_cost_total, 2),
        "realized_pnl": round(realized_net_total, 2),
        "realized_pnl_before_cost": round(realized_gross_total, 2),
        "transaction_costs_total": round(realized_cost_total, 2),
        "transaction_costs_intraday": round(realized_intraday_cost_total, 2),
        "transaction_costs_delivery": round(realized_delivery_cost_total, 2),
        "transaction_costs_open_entry_estimate": round(open_entry_fee_total, 2),
        "transaction_costs_open_exit_estimate": round(open_exit_fee_total, 2),
        "transaction_costs_open_total_estimate": round(
            open_entry_fee_total + open_exit_fee_total, 2
        ),
        "transaction_costs_total_including_open_estimate": round(
            realized_cost_total + open_entry_fee_total + open_exit_fee_total,
            2,
        ),
        "unrealized_pnl": round(unrealized_total, 2),
        "unrealized_pnl_after_cost": round(unrealized_after_cost_total, 2),
        "total_pnl": round(realized_net_total + unrealized_total, 2),
        "total_pnl_after_cost": round(
            realized_net_total + unrealized_after_cost_total, 2
        ),
    }


def _get_live_prices_batch(tickers: list[str]) -> dict:
    """Get live prices for multiple tickers using yfinance."""
    result: dict[str, dict] = {}

    clean_tickers = [str(t).strip().upper() for t in tickers if str(t).strip()]
    if not clean_tickers:
        return result
    single_request = len(clean_tickers) == 1

    def _extract_field_series(df: pd.DataFrame, field: str, ticker: str):
        if isinstance(df.columns, pd.MultiIndex):
            lvl0 = df.columns.get_level_values(0)
            if field not in lvl0:
                return None
            obj = df[field]
            if isinstance(obj, pd.DataFrame):
                if ticker in obj.columns:
                    return obj[ticker]
                if single_request and obj.shape[1] == 1:
                    return obj.iloc[:, 0]
                return None
            return obj
        return df[field] if field in df.columns else None

    try:
        # Use unadjusted prices for live parity with broker/app quotes.
        data = _safe_yf_download(
            clean_tickers if len(clean_tickers) > 1 else clean_tickers[0],
            period="2d",
            interval="1d",
            auto_adjust=False,
        )
        if data is None or data.empty:
            if len(clean_tickers) == 1:
                _record_unavailable_ticker(
                    clean_tickers[0], "empty_batch_data", "yfinance_single"
                )
            return result

        successful: set[str] = set()
        for ticker in clean_tickers:
            try:
                close_col = _extract_field_series(data, "Close", ticker)

                if close_col is None or close_col.dropna().empty:
                    if len(clean_tickers) == 1:
                        _record_unavailable_ticker(
                            ticker, "missing_close_series", "yfinance_single"
                        )
                    continue

                close_vals = close_col.dropna()
                current = float(close_vals.iloc[-1])
                prev = float(close_vals.iloc[-2]) if len(close_vals) >= 2 else current
                change = current - prev
                change_pct = (change / prev * 100) if prev != 0 else 0

                vol_col = _extract_field_series(data, "Volume", ticker)
                high_col = _extract_field_series(data, "High", ticker)
                low_col = _extract_field_series(data, "Low", ticker)
                open_col = _extract_field_series(data, "Open", ticker)

                vol = (
                    float(vol_col.dropna().iloc[-1])
                    if vol_col is not None and not vol_col.dropna().empty
                    else 0
                )
                high = (
                    float(high_col.dropna().iloc[-1])
                    if high_col is not None and not high_col.dropna().empty
                    else current
                )
                low = (
                    float(low_col.dropna().iloc[-1])
                    if low_col is not None and not low_col.dropna().empty
                    else current
                )
                open_p = (
                    float(open_col.dropna().iloc[-1])
                    if open_col is not None and not open_col.dropna().empty
                    else current
                )

                result[ticker] = {
                    "price": round(current, 2),
                    "prev_close": round(prev, 2),
                    "change": round(change, 2),
                    "change_pct": round(change_pct, 2),
                    "volume": int(vol),
                    "high": round(high, 2),
                    "low": round(low, 2),
                    "open": round(open_p, 2),
                }
                successful.add(ticker)
                if len(clean_tickers) == 1:
                    _mark_ticker_recovered(ticker, "yfinance_single")
            except Exception as exc:
                log.debug(f"Price fetch failed for {ticker}: {exc}")
                if len(clean_tickers) == 1:
                    _record_unavailable_ticker(
                        ticker,
                        f"price_fetch_exception:{type(exc).__name__}",
                        "yfinance_single",
                    )
                continue

        if not successful and len(clean_tickers) == 1:
            _record_unavailable_ticker(
                clean_tickers[0], "all_missing_single_ticker", "yfinance_single"
            )
    except Exception as e:
        log.warning(f"Batch price fetch error: {e}")

    return result


def _get_close_prices_for_date(tickers: list[str], date_str: str) -> dict[str, float]:
    """Fetch close price for the specific trading date for each ticker."""
    try:
        day = datetime.strptime(date_str, "%Y-%m-%d").date()
    except Exception:
        return {}

    # yfinance end is exclusive; add 2 days to safely capture next trading day.
    start = day.strftime("%Y-%m-%d")
    end = (day + timedelta(days=2)).strftime("%Y-%m-%d")

    prices: dict[str, float] = {}
    if not tickers:
        return prices

    try:
        data = _safe_yf_download(
            tickers,
            start=start,
            end=end,
            interval="1d",
            auto_adjust=False,
        )
        if data is None or data.empty:
            return prices

        if len(tickers) == 1:
            ticker = tickers[0]
            close_series = data["Close"] if "Close" in data.columns else None
            if close_series is not None and not close_series.dropna().empty:
                prices[ticker] = round(float(close_series.dropna().iloc[0]), 2)
            return prices

        if isinstance(data.columns, pd.MultiIndex):
            lvl0 = data.columns.get_level_values(0)
            if "Close" in lvl0:
                close_df = data["Close"]
                for ticker in tickers:
                    if ticker in close_df.columns:
                        ser = close_df[ticker].dropna()
                        if not ser.empty:
                            prices[ticker] = round(float(ser.iloc[0]), 2)
    except Exception as e:
        log.warning(f"Date-close fetch failed for {date_str}: {e}")

    return prices


def _compute_alpha_metrics(
    portfolio_return_pct: float,
    benchmark_return_pct: float,
    *,
    beta: float | None = None,
    risk_free_annual: float | None = None,
) -> dict[str, float]:
    """
    Compute both simplified and CAPM-style daily alpha (in percentage points).
    """
    rp = float(portfolio_return_pct or 0.0)
    rm = float(benchmark_return_pct or 0.0)
    beta_used = float(beta if beta is not None else ALPHA_DEFAULT_BETA)
    if not np.isfinite(beta_used):
        beta_used = ALPHA_DEFAULT_BETA
    beta_used = float(np.clip(beta_used, -5.0, 5.0))

    rf_annual = float(
        risk_free_annual if risk_free_annual is not None else ALPHA_RISK_FREE_ANNUAL
    )
    if not np.isfinite(rf_annual):
        rf_annual = 0.0
    rf_daily_pct = (rf_annual / 252.0) * 100.0

    simplified_alpha_pct = rp - rm
    capm_alpha_pct = rp - (rf_daily_pct + beta_used * (rm - rf_daily_pct))
    return {
        "simplified_alpha_pct": float(round(simplified_alpha_pct, 6)),
        "capm_alpha_pct": float(round(capm_alpha_pct, 6)),
        "beta_used": float(round(beta_used, 6)),
        "risk_free_daily_pct": float(round(rf_daily_pct, 6)),
    }


def _get_benchmark_return_pct(date_str: str, use_eod_close: bool) -> float:
    """
    Return benchmark daily return in percent for the requested view.
    """
    bench_ticker = str(_market_settings.get("benchmark_ticker", "^NSEI") or "^NSEI")

    def _live_fallback() -> float:
        live = _get_live_prices_batch([bench_ticker]).get(bench_ticker, {})
        px = _safe_float(live.get("price"))
        prev_close = _safe_float(live.get("prev_close"))
        if px > 0 and prev_close > 0:
            return (px - prev_close) / prev_close * 100.0
        return 0.0

    if not use_eod_close:
        return _live_fallback()

    try:
        day = datetime.strptime(date_str, "%Y-%m-%d").date()
        start = day.strftime("%Y-%m-%d")
        end = (day + timedelta(days=2)).strftime("%Y-%m-%d")
        data = _safe_yf_download(
            bench_ticker,
            start=start,
            end=end,
            interval="1d",
            auto_adjust=False,
        )
        if data is None or data.empty:
            return _live_fallback()
        open_col = data["Open"] if "Open" in data.columns else None
        close_col = data["Close"] if "Close" in data.columns else None
        if open_col is None or close_col is None:
            return _live_fallback()
        open_vals = open_col.dropna()
        close_vals = close_col.dropna()
        if open_vals.empty or close_vals.empty:
            return _live_fallback()
        open_px = float(open_vals.iloc[0])
        close_px = float(close_vals.iloc[0])
        if open_px <= 0:
            return _live_fallback()
        return (close_px - open_px) / open_px * 100.0
    except Exception as exc:
        log.debug("Benchmark return fetch failed for %s: %s", date_str, exc)
        return _live_fallback()


# ---------------------------------------------------------------------------
# Routes — Pages
# ---------------------------------------------------------------------------
@app.route("/")
def index():
    """Main dashboard — Groww-style stock cards."""
    return render_template("index.html")


@app.route("/stock/<ticker>")
def stock_detail(ticker: str):
    """Individual stock detail page with live chart and predictions."""
    name = ticker_names.get(ticker, _clean_name(ticker))
    return render_template("stock.html", ticker=ticker, name=name)


@app.route("/portfolio")
def portfolio_page():
    """Portfolio page with positions and full trade ledger."""
    return render_template("portfolio.html")


@app.route("/advisor")
def advisor_page():
    """Dedicated strategy advisor page (simulation only)."""
    return render_template("advisor.html")


# ---------------------------------------------------------------------------
# Routes — API
# ---------------------------------------------------------------------------
@app.route("/api/status")
def api_status():
    """Return server + market status."""
    mkt = get_market_status()
    load_elapsed = 0.0
    if models_load_started_at:
        ref = models_load_completed_at or time.time()
        load_elapsed = max(0.0, ref - models_load_started_at)

    load_progress = predictor.get_load_status() if predictor else {}
    premarket_meta = {}
    try:
        snap = _get_prediction_snapshot(use_latest_stored=True)
        if snap:
            premarket_meta = {
                "date": snap.get("date"),
                "captured_at": snap.get("captured_at"),
                "captured_within_buffer": snap.get("captured_within_buffer"),
                "snapshot_type": snap.get("snapshot_type"),
            }
    except Exception:
        with _premarket_snapshot_lock:
            premarket_meta = (
                {
                    "date": _premarket_snapshot.get("date"),
                    "captured_at": _premarket_snapshot.get("captured_at"),
                    "captured_within_buffer": _premarket_snapshot.get(
                        "captured_within_buffer"
                    ),
                }
                if _premarket_snapshot
                else {}
            )
    snapshot_now_iso = datetime.now(IST).isoformat()
    return jsonify(
        {
            "models_loaded": models_loaded,
            "models_loading": models_loading,
            "load_error": load_error,
            "model_count": len(predictor.models) if predictor else 0,
            "model_load_elapsed_sec": round(load_elapsed, 1),
            "model_load_progress": load_progress,
            "market": {
                "status": mkt["status"],
                "description": mkt["description"],
                "next_open": mkt.get("next_open", ""),
                "ist_now": mkt["ist_now"].strftime("%d %b %Y, %I:%M %p IST"),
            },
            "ticker_count": len(all_tickers),
            "premarket_config": {
                "max_buffer_minutes": PREMARKET_MAX_BUFFER_MINUTES,
                "default_tickers": PREMARKET_DEFAULT_TICKERS,
                "premarket_window": f"{PREMARKET_WINDOW_START_HOUR:02d}:{PREMARKET_WINDOW_START_MINUTE:02d}-{MARKET_LOCK_START_HOUR:02d}:{MARKET_LOCK_START_MINUTE:02d} IST",
                "market_lock_window": f"{MARKET_LOCK_START_HOUR:02d}:{MARKET_LOCK_START_MINUTE:02d}-{MARKET_LOCK_END_HOUR:02d}:{MARKET_LOCK_END_MINUTE:02d} IST",
                "precompute_slots_ist": [
                    f"{PRECOMPUTE_OPEN_HOUR:02d}:{PRECOMPUTE_OPEN_MINUTE:02d}",
                    f"{PRECOMPUTE_CLOSE_HOUR:02d}:{PRECOMPUTE_CLOSE_MINUTE:02d}",
                ],
            },
            "premarket_snapshot": premarket_meta,
            "groq": get_groq_system_status(),
            "simulation_automation": {
                "enabled": AUTO_SIMULATION_ENABLED,
                "auto_buy_enabled": AUTO_SIMULATION_AUTO_BUY,
                "interval_sec": AUTO_SIMULATION_INTERVAL_SEC,
                "trade_window": "09:30-15:30 IST (weekdays)",
                "trade_ledger": {
                    "jsonl": str(SIMULATED_TRADE_LOG_FILE),
                    "csv": str(SIMULATED_TRADE_CSV_FILE),
                    "excel": str(SIMULATED_TRADE_EXCEL_FILE),
                },
                "daily_report_time_ist": f"{TRADE_DAILY_REPORT_HOUR_IST:02d}:{TRADE_DAILY_REPORT_MINUTE_IST:02d}",
                "email_notifications_enabled": bool(TRADE_EMAIL_ENABLED),
                "email_recipient": TRADE_EMAIL_TO,
            },
        }
    )


@app.route("/api/sectors")
def api_sectors():
    """Return tickers organized by sector."""
    result = {}
    for sec, tickers in tickers_by_sector.items():
        result[sec] = {
            "display_name": SECTOR_DISPLAY.get(sec, sec),
            "tickers": [
                {"symbol": t, "name": ticker_names.get(t, _clean_name(t))}
                for t in tickers
            ],
        }
    return jsonify(result)


@app.route("/api/groq-status")
def api_groq_status():
    """Expose Groq queue/degraded-mode state for frontend banners."""
    return jsonify(get_groq_system_status())


@app.route("/api/prices")
def api_prices():
    """Get live prices for tickers. Query: ?sector=large_cap or ?tickers=RELIANCE.NS,TCS.NS"""
    sector = request.args.get("sector")
    tickers_param = request.args.get("tickers")

    if tickers_param:
        tickers = [t.strip() for t in tickers_param.split(",")]
    elif sector and sector in tickers_by_sector:
        tickers = tickers_by_sector[sector]
    else:
        # Default: large_cap
        tickers = tickers_by_sector.get("large_cap", all_tickers[:30])

    prices = _get_live_prices_batch(tickers)

    # Attach names
    for t in prices:
        prices[t]["name"] = ticker_names.get(t, _clean_name(t))
        prices[t]["symbol"] = t

    return jsonify(prices)


@app.route("/api/predict/<ticker>")
def api_predict(ticker: str):
    """Get ML prediction for a single ticker."""
    if not models_loaded or predictor is None:
        return jsonify({"error": "Models still loading, please wait..."}), 503

    force = request.args.get("force", "").lower() in ("true", "1", "yes")
    use_latest_stored = request.args.get("use_latest_stored", "").lower() in (
        "true",
        "1",
        "yes",
    )
    use_cache_arg = request.args.get("use_cache", "").lower()
    use_cache = use_cache_arg in ("true", "1", "yes")
    if use_cache_arg in ("false", "0", "no"):
        use_cache = False
    elif not force:
        # Cache-first default for fast stock-page loads.
        use_cache = True
    if use_latest_stored and not force:
        use_cache = True
    try:
        if force:
            try:
                predictor.cache.invalidate(ticker)
            except Exception:
                pass
        pred = predictor.predict_single(ticker, use_cache=use_cache and not force)
        if pred is None:
            return jsonify({"error": f"Prediction failed for {ticker}"}), 404

        # Log for end-of-day comparison
        _log_prediction(ticker, pred)

        pred["name"] = ticker_names.get(ticker, _clean_name(ticker))
        pred["source"] = "ml"
        pred["cache_policy"] = (
            "force_refresh"
            if force
            else ("cache_first" if use_cache else "fresh_inference")
        )
        return jsonify(pred)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/strategy-price/<ticker>")
def api_strategy_price(ticker: str):
    """
    Strategy-engine price endpoint.
    Returns strategy-only fields (no Groq synthesis):
      strategy_price, rr_ratio, strategy_generated_at
    """
    if not models_loaded or predictor is None:
        return jsonify({"error": "Models still loading, please wait..."}), 503

    force = request.args.get("force", "").lower() in ("true", "1", "yes")
    use_latest_stored = request.args.get("use_latest_stored", "").lower() in (
        "true",
        "1",
        "yes",
    )
    use_cache_arg = request.args.get("use_cache", "").lower()
    use_cache = use_cache_arg in ("true", "1", "yes")
    if use_cache_arg in ("false", "0", "no"):
        use_cache = False
    elif not force:
        use_cache = True
    if use_latest_stored and not force:
        use_cache = True
    now = datetime.now(IST)
    window_type = _prediction_window_type(now)
    try:
        row = {}
        if window_type in {"premarket_open", "market_open_locked"} and not force:
            row = _get_premarket_row_for_ticker(ticker, now.strftime("%Y-%m-%d"))

        if row and _safe_float(row.get("strategy_price_at_open")) > 0:
            strategy_price = round(_safe_float(row.get("strategy_price_at_open")), 2)
            rr_ratio = _safe_float(row.get("risk_reward"))
            strategy_generated_at = (
                row.get("strategy_predicted_at_open")
                or row.get("captured_at")
                or now.isoformat()
            )
            open_price = _safe_float(row.get("open_price"))
            current_price = _safe_float(row.get("current_price"))
            predicted_return_decimal = _safe_float(row.get("predicted_return_decimal"))
            payload = {
                "ticker": ticker,
                "strategy_price": strategy_price,
                "rr_ratio": rr_ratio,
                "strategy_generated_at": strategy_generated_at,
                "predicted_return_decimal": predicted_return_decimal,
                "source": "strategy",
                "open_price": round(open_price, 2) if open_price > 0 else None,
                "current_price": round(current_price, 2) if current_price > 0 else None,
                "snapshot_type": row.get("snapshot_type", window_type),
            }
            return jsonify(payload)

        if force:
            try:
                predictor.cache.invalidate(ticker)
            except Exception:
                pass

        strategy = predictor.get_strategy_price(
            ticker,
            use_cache=use_cache and not force,
        )
        if not strategy:
            return jsonify({"error": f"Strategy price unavailable for {ticker}"}), 404

        strategy["source"] = "strategy"
        strategy["snapshot_type"] = window_type
        strategy["cache_policy"] = (
            "force_refresh"
            if force
            else ("cache_first" if use_cache else "fresh_inference")
        )
        return jsonify(strategy)
    except Exception as exc:
        return jsonify({"error": str(exc)}), 500


@app.route("/api/advisor/open-buy-list")
def api_advisor_open_buy_list():
    """
    Trading-desk advisor picks for market-open simulation.
    Returns strategy-driven BUY/HOLD ideas constrained by budget + estimated fees.
    """
    if not models_loaded or predictor is None:
        return jsonify({"error": "Models still loading..."}), 503

    state = _read_portfolio_sim_state()
    budget_arg = request.args.get("budget")
    try:
        n = int(request.args.get("n", 10))
    except Exception:
        n = 10
    n = max(1, min(n, 10))

    advisor_view = str(request.args.get("view", "buy")).strip().lower()
    if advisor_view not in {"buy", "hold"}:
        advisor_view = "buy"

    try:
        budget = (
            float(budget_arg)
            if budget_arg not in (None, "")
            else float(state.get("cash", SIMULATION_DEFAULT_CASH))
        )
    except Exception:
        budget = float(state.get("cash", SIMULATION_DEFAULT_CASH))
    budget = max(1000.0, budget)

    sectors = request.args.getlist("sectors") or [
        "large_cap",
        "banking",
        "mid_cap",
        "high_volatility",
        "commodities",
    ]
    trade_type = str(request.args.get("trade_type", "equity_delivery")).strip().lower()
    if trade_type not in {"equity_delivery", "equity_intraday"}:
        trade_type = "equity_delivery"
    force_fresh = request.args.get("force_fresh", "").lower() in ("1", "true", "yes")
    use_live_prices = request.args.get("live_prices", "").lower() in (
        "1",
        "true",
        "yes",
    )
    allow_warming = request.args.get("allow_warming", "").lower() in (
        "1",
        "true",
        "yes",
    )

    try:
        allowed_tickers = {
            t
            for sec in sectors
            for t in tickers_by_sector.get(sec, [])
            if str(t).strip()
        }
        with _daily_analysis_lock:
            cached_rows = list((_daily_analysis_cache or {}).get("all_stocks", []))

        candidate_rows: list[dict] = []
        if cached_rows:
            for row in cached_rows:
                if not isinstance(row, dict):
                    continue
                ticker = str(row.get("ticker", "")).upper()
                if allowed_tickers and ticker not in allowed_tickers:
                    continue
                if _safe_float(row.get("current_price")) <= 0:
                    continue
                candidate_rows.append(dict(row))
                if len(candidate_rows) >= 200:
                    break

        raw_holds: list[dict] = []
        raw_buys: list[dict] = []
        if not candidate_rows and allow_warming and not cached_rows and not force_fresh:
            return (
                jsonify(
                    {
                        "status": "warming",
                        "message": "Advisor cache warming. Retry shortly.",
                        "retry_after_sec": 5,
                        "count": 0,
                        "picks": [],
                        "warnings": [],
                        "source": "strategy_advisor",
                        "view": advisor_view,
                    }
                ),
                202,
            )

        if force_fresh or not cached_rows:
            grouped = predictor.predict_top_picks_grouped(sectors=sectors, top_n=50)
            raw_buys = grouped.get("top_buy", []) if isinstance(grouped, dict) else []
            raw_holds = grouped.get("top_hold", []) if isinstance(grouped, dict) else []
            raw_pool = (
                list(raw_buys) + list(raw_holds)
                if advisor_view == "buy"
                else list(raw_holds) + list(raw_buys)
            )
            seen_raw = {str(r.get("ticker", "")).upper() for r in candidate_rows}
            for row in raw_pool:
                ticker = str(row.get("ticker", "")).upper()
                if not ticker or ticker in seen_raw:
                    continue
                if _safe_float(row.get("current_price")) <= 0:
                    continue
                candidate_rows.append(dict(row))
                seen_raw.add(ticker)
                if len(candidate_rows) >= 200:
                    break

        if len(candidate_rows) < max(30, n * 5):
            seen = {str(p.get("ticker", "")).upper() for p in candidate_rows}
            try:
                scan_universe = predictor._resolve_top_pick_tickers(sectors=sectors)
            except Exception:
                scan_universe = []
            for ticker in scan_universe:
                t = str(ticker or "").upper()
                if not t or t in seen:
                    continue
                try:
                    pred = predictor.predict_single(t, use_cache=True)
                except Exception:
                    pred = None
                if not pred:
                    continue
                if _safe_float(pred.get("current_price")) <= 0:
                    continue
                candidate_rows.append(dict(pred))
                seen.add(t)
                if len(candidate_rows) >= 200:
                    break

        if not candidate_rows:
            return jsonify(
                {
                    "source": "strategy",
                    "generated_at": datetime.now(IST).isoformat(),
                    "budget": round(budget, 2),
                    "view": advisor_view,
                    "count": 0,
                    "picks": [],
                    "estimated_total_cost": 0.0,
                    "remaining_budget": round(budget, 2),
                    "error": (
                        "No strategy HOLD watchlist candidates available right now."
                        if advisor_view == "hold"
                        else "No strategy BUY candidates available right now."
                    ),
                }
            )

        warnings: list[str] = []
        filtered_rows: list[dict] = []
        window_type = _prediction_window_type(datetime.now(IST))
        today_str = datetime.now(IST).strftime("%Y-%m-%d")
        snapshot_rows: dict[str, dict] = {}
        try:
            snap = _get_prediction_snapshot(use_latest_stored=True)
            if (
                isinstance(snap, dict)
                and str(snap.get("date", "")) == today_str
                and isinstance(snap.get("items"), list)
            ):
                for snap_row in snap.get("items", []):
                    if not isinstance(snap_row, dict):
                        continue
                    t = str(snap_row.get("ticker", "")).strip().upper()
                    if t:
                        snapshot_rows[t] = dict(snap_row)
        except Exception as exc:
            log.debug("Advisor snapshot overlay unavailable: %s", exc)

        for row in candidate_rows:
            raw_row = dict(row or {})
            ticker = str(raw_row.get("ticker", "")).upper()
            snap_row = dict(snapshot_rows.get(ticker, {}) or {})
            snap_strategy = _safe_float(snap_row.get("strategy_price_at_open"))
            if snap_strategy > 0:
                raw_row["strategy_price_at_open"] = round(snap_strategy, 2)
                raw_row["strategy_generated_at"] = (
                    snap_row.get("strategy_predicted_at_open")
                    or snap_row.get("captured_at")
                    or raw_row.get("strategy_generated_at")
                )
                raw_row["strategy_price_source"] = "market_open_snapshot"
                low = _safe_float(snap_row.get("entry_range_low"))
                high = _safe_float(snap_row.get("entry_range_high"))
                if low <= 0 or high <= 0:
                    low, high = _entry_price_range(snap_strategy)
                if high < low:
                    low, high = high, low
                raw_row["entry_range_low"] = round(low, 2)
                raw_row["entry_range_high"] = round(high, 2)
            else:
                raw_row["strategy_price_source"] = (
                    "strategy_model"
                    if _safe_float(raw_row.get("strategy_price_at_open")) > 0
                    else "unavailable"
                )

            strategy_price = _safe_float(
                raw_row.get("strategy_price_at_open")
                or raw_row.get("target_price")
                or raw_row.get("predicted_price")
            )
            current_price = _safe_float(raw_row.get("current_price"))
            if strategy_price <= 0:
                warnings.append(f"{ticker} skipped: strategy price unavailable.")
                continue

            ai_price = _safe_float(raw_row.get("ai_predicted_price"))
            if ai_price > 0 and strategy_price > 0:
                ai_gap = abs(ai_price - strategy_price) / max(strategy_price, 1e-9)
                if ai_gap > AI_STRATEGY_GAP_WARN:
                    warnings.append(
                        f"{ticker} AI/strategy gap {ai_gap*100:.1f}% (ai={ai_price:.2f}, strategy={strategy_price:.2f})."
                    )
                if ai_gap > AI_STRATEGY_GAP_BLOCK:
                    warnings.append(
                        f"{ticker} skipped: AI/strategy gap exceeded block threshold ({AI_STRATEGY_GAP_BLOCK*100:.0f}%)."
                    )
                    continue
            filtered_rows.append(raw_row)

        candidate_tickers = sorted(
            {
                str(r.get("ticker", "")).upper()
                for r in filtered_rows
                if str(r.get("ticker", "")).strip()
            }
        )
        live_prices: dict[str, dict] = {}
        if candidate_tickers:
            cached_map = _price_cache or {}
            for ticker in candidate_tickers:
                cached_row = (
                    cached_map.get(ticker) if isinstance(cached_map, dict) else None
                )
                if (
                    isinstance(cached_row, dict)
                    and _safe_float(cached_row.get("price")) > 0
                ):
                    live_prices[ticker] = dict(cached_row)
            if use_live_prices:
                missing = [t for t in candidate_tickers if t not in live_prices]
                if missing:
                    live_prices.update(_get_live_prices_batch(missing))

        open_now = {
            str(t).upper()
            for t in dict(state.get("open_positions", {})).keys()
            if str(t).strip()
        }
        allow_soft_candidates = advisor_view == "hold"
        strategy_candidates, build_warnings = _build_strategy_buy_candidates(
            filtered_rows,
            live_prices=live_prices,
            exclude_tickers=open_now,
            allow_soft_buy=allow_soft_candidates,
        )
        warnings.extend(build_warnings)

        if advisor_view == "buy" and len(strategy_candidates) < n:
            soft_candidates, soft_warnings = _build_strategy_buy_candidates(
                filtered_rows,
                live_prices=live_prices,
                exclude_tickers=open_now,
                allow_soft_buy=True,
            )
            warnings.extend(soft_warnings)
            seen_soft = {str(r.get("ticker", "")).upper() for r in strategy_candidates}
            for row in soft_candidates:
                ticker = str(row.get("ticker", "")).upper()
                if ticker and ticker not in seen_soft:
                    strategy_candidates.append(row)
                    seen_soft.add(ticker)
                if len(strategy_candidates) >= 50:
                    break

        candidate_pool = list(strategy_candidates)
        if advisor_view == "hold":
            hold_only: list[dict] = []
            for row in strategy_candidates:
                eff = str(row.get("effective_action", "")).upper()
                strat = str(row.get("strategy_action", "")).upper()
                if eff == "HOLD" or strat == "HOLD" or bool(row.get("soft_candidate")):
                    hold_only.append(row)
            candidate_pool = hold_only
            if len(candidate_pool) < n:
                existing = {str(r.get("ticker", "")).upper() for r in candidate_pool}
                for raw in filtered_rows:
                    ticker = str(raw.get("ticker", "")).upper()
                    if not ticker or ticker in existing:
                        continue
                    pred_ret_pct = _normalize_predicted_return_pct(
                        raw.get("predicted_return", raw.get("predicted_return_pct", 0))
                    )
                    signal = _strategy_action_from_signal(
                        raw.get("signal"),
                        predicted_return_pct=pred_ret_pct,
                    )
                    if signal != "HOLD":
                        continue
                    cp = _safe_float(raw.get("current_price"))
                    strategy_price = _safe_float(
                        raw.get("strategy_price_at_open")
                        or raw.get("target_price")
                        or raw.get("predicted_price")
                    )
                    if strategy_price <= 0:
                        strategy_price = cp
                    if not _is_realistic_price(cp) or not _is_realistic_price(
                        strategy_price
                    ):
                        continue
                    confidence_pct = float(
                        np.clip(_safe_float(raw.get("confidence", 0)), 0, 100)
                    )
                    atr_pct = float(
                        np.clip(abs(_safe_float(raw.get("atr_pct", 0))), 0, 25)
                    )
                    rr = _safe_float(raw.get("risk_reward"))
                    rr = float(np.clip(rr if rr > 0 else 1.2, 0.5, 5.0))
                    stop_loss_pct, _ = _compute_dynamic_stop_loss_pct(
                        confidence_pct=max(confidence_pct, 45.0),
                        atr_pct=atr_pct,
                        uses_sentiment=_has_relevant_today_sentiment(raw),
                    )
                    stop_loss = _round_to_tick(strategy_price * (1 - stop_loss_pct))
                    target_delta = max(
                        abs(strategy_price - stop_loss) * max(rr, 1.0),
                        strategy_price * max(pred_ret_pct / 100.0, 0.004),
                    )
                    target_price = _round_to_tick(strategy_price + target_delta)
                    entry_low, entry_high = _entry_price_range(strategy_price)
                    utility = (max(0.0, pred_ret_pct) / 100.0) * (
                        confidence_pct / 100.0
                    ) - ADVISOR_RISK_AVERSION * max(
                        stop_loss_pct, atr_pct / 100.0, 0.01
                    )
                    candidate_pool.append(
                        {
                            "ticker": ticker,
                            "name": ticker_names.get(ticker, _clean_name(ticker)),
                            "sector": str(ticker_to_sector.get(ticker, "other")),
                            "source": "strategy",
                            "current_price": round(cp, 2),
                            "strategy_price_at_open": round(strategy_price, 2),
                            "predicted_return_pct": round(pred_ret_pct, 3),
                            "confidence": round(confidence_pct, 1),
                            "model_agreement": round(
                                float(
                                    np.clip(
                                        _safe_float(raw.get("model_agreement", 0)),
                                        0,
                                        100,
                                    )
                                ),
                                1,
                            ),
                            "liquidity_factor": round(
                                float(
                                    np.clip(
                                        _safe_float(raw.get("liquidity_factor")) or 1.0,
                                        0.5,
                                        1.5,
                                    )
                                ),
                                3,
                            ),
                            "atr_pct": round(atr_pct, 3),
                            "risk_reward": round(rr, 3),
                            "stop_loss_pct": round(stop_loss_pct, 6),
                            "stop_loss_price": round(stop_loss, 2),
                            "target_price": round(target_price, 2),
                            "entry_range_low": round(entry_low, 2),
                            "entry_range_high": round(entry_high, 2),
                            "expected_edge": round(max(0.0, pred_ret_pct) / 100.0, 6),
                            "strategy_edge_component": round(
                                max(0.0, pred_ret_pct) / 100.0,
                                6,
                            ),
                            "sentiment_edge_component": 0.0,
                            "sentiment_influence": round(
                                float(np.clip(ADVISOR_SENTIMENT_INFLUENCE, 0.0, 1.0)),
                                4,
                            ),
                            "risk_metric": max(stop_loss_pct, atr_pct / 100.0, 0.01),
                            "objective_utility": float(utility),
                            "signal": "HOLD",
                            "strategy_action": "HOLD",
                            "effective_action": "HOLD",
                            "action_source": "strategy_only",
                            "soft_candidate": True,
                            "weighted_sentiment_raw": round(
                                _safe_float(raw.get("weighted_sentiment_raw")),
                                4,
                            ),
                            "sentiment_articles": int(
                                _safe_float(raw.get("sentiment_articles"))
                            ),
                            "sentiment_weight_used": round(
                                _safe_float(raw.get("sentiment_weight_used")),
                                4,
                            ),
                            "avg_volume_30d": (
                                round(_safe_float(raw.get("avg_volume_30d")), 2)
                                if _safe_float(raw.get("avg_volume_30d")) > 0
                                else None
                            ),
                            "strategy_generated_at": raw.get("timestamp")
                            or raw.get("captured_at")
                            or datetime.now(IST).isoformat(),
                        }
                    )
                    existing.add(ticker)
                    if len(candidate_pool) >= 50:
                        break

        if not candidate_pool:
            return jsonify(
                {
                    "source": "strategy",
                    "generated_at": datetime.now(IST).isoformat(),
                    "budget": round(budget, 2),
                    "trade_type": trade_type,
                    "view": advisor_view,
                    "count": 0,
                    "picks": [],
                    "estimated_total_cost": 0.0,
                    "remaining_budget": round(budget, 2),
                    "warnings": sorted(set(warnings)),
                    "error": (
                        "No strategy HOLD watchlist candidates available under risk constraints."
                        if advisor_view == "hold"
                        else "No strategy BUY candidates available under risk constraints."
                    ),
                }
            )

        if advisor_view == "hold":
            picks = []
            hold_ranked = sorted(
                candidate_pool,
                key=lambda r: (
                    _safe_float(r.get("objective_utility")),
                    _safe_float(r.get("confidence")),
                    _safe_float(r.get("model_agreement")),
                    _safe_float(r.get("predicted_return_pct")),
                ),
                reverse=True,
            )
            per_slot_budget = max(1000.0, budget / max(1, n))
            used = set()
            for row in hold_ranked:
                ticker = str(row.get("ticker", "")).upper()
                if not ticker or ticker in used:
                    continue
                price = _safe_float(row.get("current_price")) or _safe_float(
                    row.get("strategy_price_at_open")
                )
                if price <= 0:
                    continue
                qty = _max_affordable_qty(per_slot_budget, price, trade_type=trade_type)
                if (
                    qty <= 0
                    and _max_affordable_qty(budget, price, trade_type=trade_type) > 0
                ):
                    qty = 1
                if qty <= 0:
                    continue
                notional = round(price * qty, 2)
                fee = _estimate_entry_fee(notional, trade_type=trade_type)
                total = round(notional + fee, 2)
                if total <= 0 or total > budget + 1e-9:
                    continue
                enriched = dict(row)
                enriched["suggested_qty"] = int(qty)
                enriched["estimated_notional"] = round(notional, 2)
                enriched["estimated_fee"] = round(fee, 2)
                enriched["est_trade_cost"] = round(total, 2)
                picks.append(enriched)
                used.add(ticker)
                if len(picks) >= n:
                    break
        else:
            picks = _optimize_candidate_allocations(
                candidate_pool,
                budget=budget,
                trade_type=trade_type,
                max_positions=n,
            )

        for row in picks:
            row["volatility_atr_pct"] = round(abs(_safe_float(row.get("atr_pct"))), 2)
            row["sentiment_weighted_score"] = round(
                _safe_float(row.get("weighted_sentiment_raw")), 4
            )
            row["stop_loss_pct"] = round(
                _safe_float(row.get("stop_loss_pct")) * 100.0, 3
            )
            row["strategy_price_at_open"] = _display_price_or_none(
                row.get("strategy_price_at_open")
            )
            row["current_price"] = _display_price_or_none(row.get("current_price"))
            row["entry_range_low"] = _display_price_or_none(row.get("entry_range_low"))
            row["entry_range_high"] = _display_price_or_none(
                row.get("entry_range_high")
            )
            row["buy_range_low"] = row["entry_range_low"]
            row["buy_range_high"] = row["entry_range_high"]
            row["stop_loss_price"] = _display_price_or_none(row.get("stop_loss_price"))
            row["target_price"] = _display_price_or_none(row.get("target_price"))
            row["sell_range_low"] = row["stop_loss_price"]
            row["sell_range_high"] = row["target_price"]
            row["estimated_fee"] = _display_price_or_none(row.get("estimated_fee"))
            row["estimated_notional"] = _display_price_or_none(
                row.get("estimated_notional")
            )
            row["est_trade_cost"] = _display_price_or_none(row.get("est_trade_cost"))
            qty_display = _display_qty_or_none(row.get("suggested_qty"), digits=0)
            row["suggested_qty"] = int(qty_display) if qty_display is not None else None
            row["advisor_view"] = advisor_view
            row["advisor_action"] = "WATCH" if advisor_view == "hold" else "BUY"
            row["strategy_price_source"] = str(
                row.get("strategy_price_source") or "strategy_model"
            )
            row["strategy_generated_at_display"] = _format_ist_timestamp(
                row.get("strategy_generated_at")
            )
            row["strategy_locked_for_window"] = window_type in {
                "premarket_open",
                "market_open_locked",
            }
            if _safe_float(row.get("liquidity_factor")) < 0.8:
                warnings.append(
                    f"{row.get('ticker')} has lower liquidity "
                    f"(factor={_safe_float(row.get('liquidity_factor')):.2f}, "
                    f"avg_volume_30d={_safe_float(row.get('avg_volume_30d')):.0f})."
                )

        picks = [
            row
            for row in picks
            if row.get("suggested_qty") and row.get("est_trade_cost")
        ]

        total_cost = round(sum(_safe_float(r.get("est_trade_cost")) for r in picks), 2)
        remaining_budget = round(max(0.0, budget - total_cost), 2)

        return jsonify(
            {
                "source": "strategy",
                "schema_version": 1,
                "generated_at": datetime.now(IST).isoformat(),
                "budget": round(budget, 2),
                "trade_type": trade_type,
                "view": advisor_view,
                "count": len(picks),
                "picks": picks,
                "estimated_total_cost": round(total_cost, 2),
                "remaining_budget": round(remaining_budget, 2),
                "warnings": sorted(set(warnings)),
                "simulation_cash_available": round(
                    float(state.get("cash", SIMULATION_DEFAULT_CASH) or 0), 2
                ),
            }
        )
    except Exception as exc:
        log.exception("Advisor open-buy-list failed")
        return jsonify({"error": str(exc)}), 500


@app.route("/api/simulate/portfolio")
def api_simulate_portfolio():
    """Return simulated trading-desk portfolio state."""
    state = _read_portfolio_sim_state()
    open_positions = dict(state.get("open_positions", {}))
    tickers = sorted(open_positions.keys())
    live_prices = _get_live_prices_batch(tickers) if tickers else {}

    holdings = []
    invested = 0.0
    mtm_value = 0.0
    invested_after_cost = 0.0
    current_after_cost = 0.0
    open_txn_cost_eaten = 0.0
    state_dirty = False
    fallback_prices_cache: dict[str, dict] | None = None
    for ticker in tickers:
        pos = dict(open_positions.get(ticker, {}))
        qty = _safe_float(pos.get("quantity"))
        if qty <= 0:
            continue
        entry = _safe_float(pos.get("avg_entry_price"))
        live = _safe_float(live_prices.get(ticker, {}).get("price"))
        if live <= 0:
            live = _safe_float(pos.get("last_price"))
        if live <= 0:
            if fallback_prices_cache is None:
                fallback_prices_cache = _latest_sim_prices_from_ledgers()
            live = _safe_float(fallback_prices_cache.get(ticker, {}).get("price"))
            if live > 0:
                pos["last_price"] = round(live, 2)
                pos["last_price_at"] = str(
                    fallback_prices_cache.get(ticker, {}).get("timestamp")
                    or datetime.now(IST).isoformat()
                )
                pos["last_price_source"] = str(
                    fallback_prices_cache.get(ticker, {}).get("source") or "ledger"
                )
                open_positions[ticker] = pos
                state_dirty = True
        if live > 0 and abs(_safe_float(pos.get("last_price")) - live) > 1e-9:
            pos["last_price"] = round(live, 2)
            pos["last_price_at"] = datetime.now(IST).isoformat()
            pos["last_price_source"] = "live_quote"
            open_positions[ticker] = pos
            state_dirty = True
        strategy_target = _safe_float(pos.get("target_price"))
        stop_loss = _safe_float(pos.get("stop_loss_price"))
        entry_value = qty * entry
        mark = qty * (live if live > 0 else entry)
        entry_fee_paid = _safe_float(pos.get("total_entry_fees"))
        trade_type_hint = str(pos.get("trade_type_hint") or "equity_delivery")
        exit_fee_estimate = (
            _estimate_exit_fee(mark, trade_type=trade_type_hint) if mark > 0 else 0.0
        )
        tx_cost_eaten = entry_fee_paid + exit_fee_estimate
        invested_with_cost = entry_value + entry_fee_paid
        current_value_after_cost = max(0.0, mark - exit_fee_estimate)
        invested += entry_value
        mtm_value += mark
        invested_after_cost += invested_with_cost
        current_after_cost += current_value_after_cost
        open_txn_cost_eaten += tx_cost_eaten
        holdings.append(
            {
                "ticker": ticker,
                "name": ticker_names.get(ticker, _clean_name(ticker)),
                "quantity": _display_qty_or_none(qty),
                "entry_price": _display_price_or_none(entry),
                "current_price": _display_price_or_none(live),
                "target_price": (_display_price_or_none(strategy_target)),
                "stop_loss_price": _display_price_or_none(stop_loss),
                "entry_range_low": _display_price_or_none(pos.get("entry_range_low")),
                "entry_range_high": _display_price_or_none(pos.get("entry_range_high")),
                "unrealized_pnl": round(mark - entry_value, 2),
                "transaction_cost_eaten": round(tx_cost_eaten, 2),
                "entry_fee_paid": round(entry_fee_paid, 2),
                "exit_fee_estimate": round(exit_fee_estimate, 2),
                "invested_value_after_cost": round(invested_with_cost, 2),
                "current_value_after_cost": round(current_value_after_cost, 2),
                "unrealized_pnl_after_cost": round(
                    current_value_after_cost - invested_with_cost, 2
                ),
                "next_recommended_sell_time": _next_recommended_sell_time(
                    live if live > 0 else entry, stop_loss, strategy_target
                ),
                "strategy_recommendation": str(
                    pos.get("strategy_recommendation", "HOLD")
                ).upper(),
                "strategy_recommendation_source": str(
                    pos.get("strategy_recommendation_source", "strategy_signal")
                ),
                "source": "strategy",
                "opened_at": pos.get("opened_at"),
            }
        )
    if state_dirty:
        state["open_positions"] = open_positions
        _write_portfolio_sim_state(state)
    cash = float(
        state.get("cash", state.get("initial_cash", SIMULATION_DEFAULT_CASH)) or 0
    )
    return jsonify(
        {
            "schema_version": int(state.get("schema_version", 1)),
            "source": "strategy_simulation",
            "initial_cash": round(
                float(state.get("initial_cash", SIMULATION_DEFAULT_CASH)), 2
            ),
            "cash": round(cash, 2),
            "invested_value": round(invested, 2),
            "invested_value_after_cost": round(invested_after_cost, 2),
            "mark_to_market_value": round(mtm_value, 2),
            "current_value_after_cost": round(current_after_cost, 2),
            "transaction_cost_eaten_open_positions": round(open_txn_cost_eaten, 2),
            "equity_value": round(cash + mtm_value, 2),
            "equity_value_after_cost": round(cash + current_after_cost, 2),
            "open_positions_count": len(holdings),
            "closed_trades_count": len(state.get("closed_trades", [])),
            "holdings": holdings,
            "trade_history": state.get("trade_history", [])[-50:],
            "transaction_summary": _build_transaction_summary(
                since_ts=state.get("session_started_at")
            ),
            "daily_loss_tracker": state.get("daily_loss_tracker", {}),
            "session_started_at": state.get("session_started_at"),
            "updated_at": state.get("updated_at"),
        }
    )


@app.route("/api/simulate/transactions-summary")
def api_simulate_transactions_summary():
    """Aggregate simulated BUY/SELL counts, costs and totals for a date."""
    date_str = request.args.get("date", "").strip()
    target_date = date_str or datetime.now(IST).strftime("%Y-%m-%d")
    session_scope = request.args.get("session", "").lower() in ("1", "true", "yes")
    since_ts = None
    if session_scope:
        since_ts = _read_portfolio_sim_state().get("session_started_at")
    summary = _build_transaction_summary(target_date, since_ts=since_ts)
    return jsonify(summary)


@app.route("/api/simulate/trade-ledger")
def api_simulate_trade_ledger():
    """Return BUY/SELL paired ledger rows with HOLD/SOLD status."""
    date_str = request.args.get("date", "").strip()
    target_date = date_str or datetime.now(IST).strftime("%Y-%m-%d")
    try:
        limit = int(request.args.get("limit", 300))
    except Exception:
        limit = 300
    limit = max(1, min(limit, 1000))
    session_scope = request.args.get("session", "").lower() in ("1", "true", "yes")
    since_ts = None
    if session_scope:
        since_ts = _read_portfolio_sim_state().get("session_started_at")
    payload = _build_sim_trade_ledger(target_date, since_ts=since_ts, limit=limit)
    return jsonify(payload)


@app.route("/api/simulate/daily-report")
def api_simulate_daily_report():
    """Generate or fetch the 15:45 IST style simulation summary report."""
    date_str = request.args.get("date", "").strip()
    force = request.args.get("force", "").lower() in ("1", "true", "yes")
    target_date = date_str or datetime.now(IST).strftime("%Y-%m-%d")
    report_file = DAILY_TRADE_REPORT_DIR / f"{target_date}.json"
    if report_file.exists() and not force:
        try:
            return jsonify(json.loads(report_file.read_text()))
        except Exception:
            pass
    report = _write_daily_trade_report(target_date)
    return jsonify(report)


@app.route("/api/simulate/trade", methods=["POST"])
def api_simulate_trade():
    """
    Simulated trade execution (no live orders).
    Supports BUY, SELL, AUTO_CHECK, RESET.
    """
    payload = request.get_json(silent=True) or {}
    if not payload:
        # Support query-string invocation for quick smoke checks:
        # /api/simulate/trade?action=AUTO_CHECK&auto_buy=true
        payload = dict(request.args or {})
    action = str(payload.get("action", payload.get("side", "BUY"))).strip().upper()
    if action not in {"BUY", "SELL", "AUTO_CHECK", "RESET"}:
        return jsonify({"error": "action must be BUY, SELL, AUTO_CHECK, or RESET"}), 400

    state = _read_portfolio_sim_state()
    now_iso = datetime.now(IST).isoformat()
    state["daily_loss_tracker"] = _daily_tracker_for_state(state, now_ts=now_iso)

    if action == "RESET":
        budget = _safe_float(payload.get("budget"))
        if budget <= 0:
            budget = SIMULATION_DEFAULT_CASH
        if budget <= 0:
            budget = SIMULATION_DEFAULT_CASH
        clear_history = _as_bool(payload.get("clear_history"), default=True)
        clear_portfolio_sim_trades = _as_bool(
            payload.get("clear_portfolio_sim_trades"),
            default=True,
        )
        cleanup = (
            _clear_simulation_history(
                clear_portfolio_sim_trades=clear_portfolio_sim_trades
            )
            if clear_history
            else {}
        )
        state = _simulation_state_template(cash=budget)
        _write_portfolio_sim_state(state)
        if not clear_history:
            _log_simulated_trade(
                {
                    "timestamp": now_iso,
                    "action": "RESET",
                    "budget": round(budget, 2),
                    "source": "strategy_simulation",
                }
            )
        return jsonify(
            {
                "ok": True,
                "message": "Simulation portfolio reset.",
                "state": state,
                "cleanup": cleanup,
            }
        )

    if action == "AUTO_CHECK":
        if not _is_market_trade_window():
            return jsonify(
                {
                    "ok": True,
                    "action": "AUTO_CHECK",
                    "triggered_count": 0,
                    "events": [],
                    "auto_buy_events": [],
                    "trailing_updates": [],
                    "state": state,
                    "source": "strategy_simulation",
                    "note": (
                        "Auto-check skipped: market window closed. "
                        "Auto buy/sell simulation runs only in 09:30-15:30 IST."
                    ),
                    "market_window": _prediction_window_type(),
                    "server_time_ist": datetime.now(IST).isoformat(),
                }
            )
        result = _run_sim_auto_check(
            state,
            auto_buy_enabled=_as_bool(payload.get("auto_buy"), default=False),
            now_iso=now_iso,
            source="strategy_simulation",
        )
        state = result["state"]
        _write_portfolio_sim_state(state)
        return jsonify(
            {
                "ok": True,
                "action": "AUTO_CHECK",
                "triggered_count": result["triggered_count"],
                "events": result["events"],
                "auto_buy_events": result["auto_buy_events"],
                "trailing_updates": result["trailing_updates"],
                "state": state,
                "source": "strategy_simulation",
            }
        )

    if not _is_market_trade_window():
        now_iso = datetime.now(IST).isoformat()
        return (
            jsonify(
                {
                    "error": (
                        "Simulated BUY/SELL triggers are allowed only during market hours "
                        "(09:30-15:30 IST, trading days)."
                    ),
                    "action": action,
                    "market_window": _prediction_window_type(),
                    "server_time_ist": now_iso,
                    "source": "strategy_simulation",
                }
            ),
            403,
        )

    ticker = str(payload.get("ticker", "")).strip().upper()
    if not ticker:
        return jsonify({"error": "ticker is required"}), 400
    if not _is_tradeable_ticker(ticker):
        return jsonify({"error": f"{ticker} is not in configured ticker universe"}), 400
    try:
        quantity = float(payload.get("quantity", 0) or 0)
    except Exception:
        quantity = 0.0
    if quantity <= 0:
        return jsonify({"error": "quantity must be > 0"}), 400
    if quantity > SIM_QTY_HARD_MAX:
        return (
            jsonify({"error": f"quantity exceeds hard cap ({SIM_QTY_HARD_MAX})"}),
            400,
        )

    price = _safe_float(payload.get("price"))
    if price <= 0:
        price = _safe_float(
            _get_live_prices_batch([ticker]).get(ticker, {}).get("price")
        )
    if price <= 0:
        return jsonify({"error": f"price not available for {ticker}"}), 400
    if not _is_realistic_price(price):
        return jsonify({"error": "price is outside allowed simulation bounds"}), 400

    trade_type = str(payload.get("trade_type", "equity_delivery")).strip().lower()
    if trade_type not in {"equity_delivery", "equity_intraday"}:
        trade_type = "equity_delivery"

    if action == "BUY":
        strategy_entry = _safe_float(payload.get("strategy_entry_price"))
        if strategy_entry <= 0:
            strategy_entry = price
        entry_low = _safe_float(payload.get("entry_range_low"))
        entry_high = _safe_float(payload.get("entry_range_high"))
        if entry_low <= 0 or entry_high <= 0:
            entry_low, entry_high = _entry_price_range(strategy_entry)
        confidence_pct = _safe_float(payload.get("confidence"))
        atr_pct = abs(_safe_float(payload.get("atr_pct")))
        uses_sentiment = bool(payload.get("uses_sentiment", False))
        stop_loss = _safe_float(payload.get("stop_loss_price"))
        target_price = _safe_float(payload.get("target_price"))
        rr_ratio = _safe_float(payload.get("risk_reward"))
        if rr_ratio <= 0:
            rr_ratio = 1.5
        if stop_loss <= 0:
            sl_pct, skip = _compute_dynamic_stop_loss_pct(
                confidence_pct=confidence_pct,
                atr_pct=atr_pct,
                uses_sentiment=uses_sentiment,
            )
            if skip:
                return (
                    jsonify(
                        {"error": "Trade blocked: low confidence under risk policy"}
                    ),
                    400,
                )
            stop_loss = _round_to_tick(price * (1 - sl_pct))
        if target_price <= 0:
            target_price = _round_to_tick(
                price + max(abs(price - stop_loss) * max(rr_ratio, 1.0), price * 0.01)
            )
        strategy_recommendation = (
            str(
                payload.get("strategy_recommendation")
                or _strategy_action_from_signal(
                    str(payload.get("signal", "")).strip().upper(),
                    predicted_return_pct=_safe_float(payload.get("predicted_return")),
                )
                or "HOLD"
            )
            .strip()
            .upper()
        )
        if strategy_recommendation not in {"BUY", "SELL", "HOLD"}:
            strategy_recommendation = "HOLD"

        state, event, err = _execute_sim_buy(
            state,
            ticker=ticker,
            quantity=quantity,
            price=price,
            strategy_entry_price=strategy_entry,
            entry_range_low=entry_low,
            entry_range_high=entry_high,
            stop_loss_price=stop_loss,
            target_price=target_price,
            risk_reward=rr_ratio,
            trade_type=trade_type,
            timestamp=now_iso,
            reason="manual_buy",
            strategy_recommendation=strategy_recommendation,
        )
        if err:
            return jsonify({"error": err}), 400
        if not event:
            return jsonify({"error": "Buy execution failed"}), 400
        _write_portfolio_sim_state(state)
        _log_simulated_trade(event)
        return jsonify({"ok": True, "event": event, "state": state})

    # SELL
    state, event, err = _execute_sim_sell(
        state,
        ticker=ticker,
        quantity=quantity,
        price=price,
        reason="manual_sell",
        timestamp=now_iso,
    )
    if err:
        return jsonify({"error": err}), 400
    if event:
        event["source"] = "strategy_simulation"
        _log_simulated_trade(event)
    _write_portfolio_sim_state(state)
    return jsonify({"ok": True, "event": event, "state": state})


@app.route("/api/simulate/position/<ticker>", methods=["DELETE"])
def api_simulate_position_delete(ticker: str):
    """
    Close/remove one simulated open position by ticker.
    Uses live price when available, otherwise falls back to avg entry.
    """
    t = str(ticker or "").strip().upper()
    if not t:
        return jsonify({"error": "ticker is required"}), 400

    state = _read_portfolio_sim_state()
    pos = dict(dict(state.get("open_positions", {})).get(t, {}) or {})
    if not pos:
        return jsonify({"error": f"No open simulated position for {t}"}), 404

    qty = _safe_float(pos.get("quantity"))
    if qty <= 0:
        return jsonify({"error": f"Invalid simulated quantity for {t}"}), 400

    price = _safe_float(request.args.get("price"))
    if price <= 0:
        live = _get_live_prices_batch([t]).get(t, {})
        price = _safe_float(live.get("price"))
    if price <= 0:
        price = _safe_float(pos.get("avg_entry_price"))
    if price <= 0:
        return jsonify({"error": f"Price unavailable to close {t}"}), 400

    now_iso = datetime.now(IST).isoformat()
    state, event, err = _execute_sim_sell(
        state,
        ticker=t,
        quantity=qty,
        price=price,
        reason="manual_portfolio_remove",
        timestamp=now_iso,
    )
    if err:
        return jsonify({"error": err}), 400
    if event:
        event["source"] = "strategy_simulation"
        _log_simulated_trade(event)
    _write_portfolio_sim_state(state)
    return jsonify({"ok": True, "event": event, "state": state})


@app.route("/api/top-picks")
def api_top_picks():
    """Get top picks grouped by strategy mode or AI mode."""
    if not models_loaded or predictor is None:
        return jsonify({"error": "Models still loading..."}), 503

    sectors = request.args.getlist("sectors") or ["large_cap", "banking"]
    try:
        top_n = int(request.args.get("n", 10))
    except Exception:
        top_n = 10
    top_n = max(1, min(top_n, 50))
    grouped = request.args.get("grouped", "").lower() in ("1", "true", "yes")
    force = request.args.get("force", "").lower() in ("1", "true", "yes")
    refresh_ai = request.args.get("refresh_ai", "").lower() in ("1", "true", "yes")
    mode = str(request.args.get("mode", "ai") or "ai").strip().lower()
    if mode not in {"ai", "strategy"}:
        mode = "ai"

    use_cached_analysis = request.args.get("use_cached_analysis", "false").lower() in (
        "1",
        "true",
        "yes",
    )
    ai_refresh_limit = max(0, int(os.environ.get("TOP_PICKS_AI_REFRESH_LIMIT", "6")))
    ai_refresh_count = 0

    def _ai_signal_from_return(ret_pct: float) -> str:
        if ret_pct > 0.2:
            return "BUY"
        if ret_pct < -0.2:
            return "SELL"
        return "HOLD"

    def _enrich_pick(pick: dict) -> dict:
        nonlocal ai_refresh_count
        p = dict(pick)
        ticker = str(p.get("ticker", "")).upper()
        current_price = _safe_float(p.get("current_price"))
        if current_price <= 0:
            return p

        p["name"] = ticker_names.get(ticker, _clean_name(ticker))
        strategy_ret = _normalize_predicted_return_pct(p.get("predicted_return", 0))
        strategy_px = _safe_float(p.get("target_price") or p.get("predicted_price"))
        if strategy_px <= 0 and current_price > 0:
            strategy_px = round(current_price * (1 + strategy_ret / 100.0), 2)

        ai_px_existing = _safe_float(p.get("ai_predicted_price"))
        ai_meta = {"source": str(p.get("ai_source", "none"))}
        should_try_ai = ai_px_existing <= 0 or refresh_ai
        allow_generate = (
            refresh_ai
            and ai_refresh_count < ai_refresh_limit
            and bool(os.environ.get("GROQ_API_KEY"))
        )
        if should_try_ai:
            ai_meta = _resolve_ai_forecast_price(
                ticker,
                open_price=_safe_float(p.get("open_price")) or current_price,
                strategy_price=strategy_px or current_price,
                current_price=current_price,
                allow_generate=allow_generate,
            )
            if allow_generate:
                ai_refresh_count += 1

        ai_px = _safe_float(ai_meta.get("price")) if should_try_ai else ai_px_existing
        if ai_px <= 0:
            ai_px = ai_px_existing

        ai_ret = 0.0
        ai_signal = "HOLD"
        if ai_px > 0 and current_price > 0:
            ai_ret = ((ai_px - current_price) / current_price) * 100.0
            ai_signal = _ai_signal_from_return(ai_ret)

        p["target_price"] = round(strategy_px, 2) if strategy_px > 0 else None
        p["strategy_predicted_return"] = round(strategy_ret, 3)
        p["ai_predicted_price"] = round(ai_px, 2) if ai_px > 0 else None
        p["ai_predicted_return"] = round(ai_ret, 3) if ai_px > 0 else None
        p["ai_signal"] = ai_signal
        p["ai_source"] = ai_meta.get("source", p.get("ai_source", "none"))
        p["source"] = "strategy"
        p["ai_value_source"] = "ai"
        p["ranking_source"] = mode
        return p

    def _group_from_cached_analysis() -> dict[str, list[dict]] | None:
        if not use_cached_analysis or force:
            return None
        with _daily_analysis_lock:
            cached_rows = list((_daily_analysis_cache or {}).get("all_stocks", []))
        if not cached_rows:
            return None

        allowed_tickers: set[str] | None = None
        if sectors:
            allowed_tickers = set()
            for sec in sectors:
                allowed_tickers.update(tickers_by_sector.get(sec, []))

        rows = []
        for row in cached_rows:
            ticker = str(row.get("ticker", "")).upper()
            if allowed_tickers is not None and ticker not in allowed_tickers:
                continue
            if _safe_float(row.get("current_price")) <= 0:
                continue
            r = dict(row)
            if "_score" not in r:
                r["_score"] = _safe_float(r.get("composite_score"))
            rows.append(r)

        def _rank_key(item: dict):
            return (
                _safe_float(item.get("confidence")),
                _safe_float(item.get("_score")),
            )

        buys = sorted(
            [r for r in rows if str(r.get("signal", "")).upper() == "BUY"],
            key=_rank_key,
            reverse=True,
        )[:top_n]
        sells = sorted(
            [r for r in rows if str(r.get("signal", "")).upper() == "SELL"],
            key=_rank_key,
            reverse=True,
        )[:top_n]
        holds = sorted(
            [r for r in rows if str(r.get("signal", "")).upper() == "HOLD"],
            key=_rank_key,
            reverse=True,
        )[:top_n]
        return {"top_buy": buys, "top_sell": sells, "top_hold": holds}

    try:
        if grouped:
            groups = _group_from_cached_analysis()
            if use_cached_analysis and not force and not groups:
                return (
                    jsonify(
                        {
                            "status": "warming",
                            "message": "Top picks are warming up in background. Please retry shortly.",
                            "retry_after_sec": 5,
                            "mode": mode,
                        }
                    ),
                    202,
                )
            if not groups:
                groups = predictor.predict_top_picks_grouped(
                    sectors=sectors, top_n=top_n
                )

            cleaned_groups: dict[str, list[dict]] = {
                "top_buy": [],
                "top_sell": [],
                "top_hold": [],
            }
            for key in cleaned_groups:
                raw_rows = groups.get(key, [])
                rows = [
                    _enrich_pick(p)
                    for p in raw_rows
                    if _safe_float(p.get("current_price")) > 0
                ]
                cleaned_groups[key] = rows

            if mode == "ai":
                dedup: dict[str, dict] = {}
                for bucket in ("top_buy", "top_sell", "top_hold"):
                    for row in cleaned_groups.get(bucket, []):
                        ticker = str(row.get("ticker", "")).upper()
                        if not ticker:
                            continue
                        dedup[ticker] = row
                rows = list(dedup.values())
                for row in rows:
                    ai_ret = _safe_float(row.get("ai_predicted_return"))
                    ai_sig = str(row.get("ai_signal", "HOLD")).upper()
                    row["signal"] = ai_sig
                    row["predicted_return"] = round(ai_ret, 3)
                    row["pick_source"] = "ai"

                rows.sort(
                    key=lambda r: (
                        abs(_safe_float(r.get("ai_predicted_return"))),
                        _safe_float(r.get("confidence")),
                        _safe_float(r.get("model_agreement")),
                        _safe_float(r.get("_score")),
                    ),
                    reverse=True,
                )
                ai_buys = [
                    r for r in rows if str(r.get("signal", "")).upper() == "BUY"
                ][:top_n]
                ai_sells = [
                    r for r in rows if str(r.get("signal", "")).upper() == "SELL"
                ][:top_n]
                ai_holds = [
                    r for r in rows if str(r.get("signal", "")).upper() == "HOLD"
                ][:top_n]
                return jsonify(
                    {
                        "mode": "ai",
                        "top_buy": ai_buys,
                        "top_sell": ai_sells,
                        "top_hold": ai_holds,
                    }
                )

            for bucket, signal in (
                ("top_buy", "BUY"),
                ("top_sell", "SELL"),
                ("top_hold", "HOLD"),
            ):
                for row in cleaned_groups.get(bucket, []):
                    row["signal"] = signal
                    row["pick_source"] = "strategy"
            return jsonify({"mode": "strategy", **cleaned_groups})

        picks = predictor.predict_top_picks(sectors=sectors, top_n=top_n)
        picks = [
            _enrich_pick(p) for p in picks if _safe_float(p.get("current_price")) > 0
        ]
        if mode == "ai":
            for row in picks:
                row["signal"] = str(row.get("ai_signal", "HOLD")).upper()
                row["predicted_return"] = round(
                    _safe_float(row.get("ai_predicted_return")), 3
                )
                row["pick_source"] = "ai"
            picks.sort(
                key=lambda r: (
                    abs(_safe_float(r.get("ai_predicted_return"))),
                    _safe_float(r.get("confidence")),
                    _safe_float(r.get("model_agreement")),
                    _safe_float(r.get("_score")),
                ),
                reverse=True,
            )
        else:
            for row in picks:
                row["pick_source"] = "strategy"
            picks.sort(
                key=lambda r: (
                    _safe_float(r.get("confidence")),
                    _safe_float(r.get("_score")),
                ),
                reverse=True,
            )
        return jsonify(picks[:top_n])
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/premarket-outlook")
def api_premarket_outlook():
    """
    Premarket snapshot endpoint.
    Returns:
      ticker, current_price, strategy_price_at_open, ai_predicted_price,
      strategy_direction, ai_direction, captured_at,
      strategy_predicted_at_open, ai_predicted_at_open.
    """
    if predictor is None:
        return jsonify({"error": "Predictor not initialized"}), 503

    force = request.args.get("force", "").lower() in ("1", "true", "yes")
    use_latest_stored = request.args.get("use_latest_stored", "").lower() in (
        "1",
        "true",
        "yes",
    )
    try:
        snapshot = _get_prediction_snapshot(
            force=force,
            use_latest_stored=use_latest_stored,
        )
        items = snapshot.get("items", [])
        return jsonify(
            {
                "date": snapshot.get("date"),
                "captured_at": snapshot.get("captured_at"),
                "captured_at_actual": snapshot.get("captured_at_actual"),
                "capture_cutoff": snapshot.get("capture_cutoff"),
                "captured_within_buffer": snapshot.get("captured_within_buffer", True),
                "buffer_minutes": snapshot.get(
                    "buffer_minutes", PREMARKET_MAX_BUFFER_MINUTES
                ),
                "snapshot_type": snapshot.get("snapshot_type", "premarket_open"),
                "capture_note": snapshot.get("capture_note"),
                "use_latest_stored": use_latest_stored,
                "items": items,
            }
        )
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/intraday/<ticker>")
def api_intraday(ticker: str):
    """Get intraday price data for charting."""
    period = request.args.get("period", "5d")
    interval = request.args.get("interval", "15m")

    try:
        df = get_intraday_data(ticker, period=period, interval=interval)
        if (df is None or df.empty) and interval == "1m":
            # Fallback when 1m bars are not available for the requested period window.
            df = get_intraday_data(ticker, period=period, interval="5m")
        if df is None or df.empty:
            return jsonify({"error": "No intraday data"}), 404

        records = []
        for idx, row in df.iterrows():
            records.append(
                {
                    "time": idx.isoformat() if hasattr(idx, "isoformat") else str(idx),
                    "open": round(float(row.get("Open", 0)), 2),
                    "high": round(float(row.get("High", 0)), 2),
                    "low": round(float(row.get("Low", 0)), 2),
                    "close": round(float(row.get("Close", 0)), 2),
                    "volume": int(row.get("Volume", 0)),
                }
            )
        return jsonify(records)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/history/<ticker>")
def api_history(ticker: str):
    """Get daily price history for charting."""
    period = request.args.get("period", "1y")

    try:
        df = _safe_yf_download(
            ticker,
            period=period,
            interval="1d",
            auto_adjust=False,
        )
        if df is None or df.empty:
            return jsonify({"error": "No data"}), 404
        if isinstance(df.columns, pd.MultiIndex):
            df.columns = df.columns.get_level_values(0)

        records = []
        for idx, row in df.iterrows():
            records.append(
                {
                    "time": idx.strftime("%Y-%m-%d"),
                    "open": round(float(row["Open"]), 2),
                    "high": round(float(row["High"]), 2),
                    "low": round(float(row["Low"]), 2),
                    "close": round(float(row["Close"]), 2),
                    "volume": int(row["Volume"]),
                }
            )
        return jsonify(records)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/expected-vs-actual")
def api_expected_vs_actual():
    """
    Compare today's (or a given date's) predictions vs actual closing prices.
    Returns hit rate, alpha, and per-stock results.
    """
    date_str = request.args.get("date", datetime.now(IST).strftime("%Y-%m-%d"))
    log_file = PREDICTION_LOG_DIR / f"{date_str}.json"

    if not log_file.exists():
        return jsonify(
            {"error": f"No predictions logged for {date_str}", "date": date_str}
        )

    with open(log_file) as f:
        predictions = json.load(f)

    if not predictions:
        return jsonify({"error": "Empty prediction log", "date": date_str})

    # Persist and hydrate tracker outcomes so expected-vs-actual stays retraining-ready.
    tracker_results = PredictionTracker.check_outcomes(date_str)
    tracker_by_ticker = tracker_results if isinstance(tracker_results, dict) else {}

    # Fetch actual prices
    tickers_to_check = list(predictions.keys())
    today_str = datetime.now(IST).strftime("%Y-%m-%d")
    market_status = get_market_status().get("status", "")
    use_eod_close = date_str != today_str or market_status in {"after_hours", "weekend"}
    if use_eod_close:
        actual_prices = _get_close_prices_for_date(tickers_to_check, date_str)
    else:
        actuals = _get_live_prices_batch(tickers_to_check)
        actual_prices = {k: v.get("price", 0) for k, v in actuals.items()}

    # Benchmark return for requested date/view (in %).
    benchmark_return_pct = _get_benchmark_return_pct(date_str, use_eod_close)

    def _rescale_logged_price(raw_price: float, actual_price: float) -> float:
        """
        Repair stale logged prices after corporate actions or unit drift.
        Keeps normal prices unchanged and only rescales obvious outliers.
        """
        if raw_price <= 0 or actual_price <= 0:
            return raw_price
        ratio = raw_price / actual_price
        scale_factors = (2, 3, 4, 5, 6, 8, 10, 12, 15, 20, 25, 30, 40, 50, 75, 100)
        for factor in scale_factors:
            if abs(ratio - factor) / factor <= 0.35:
                return raw_price / factor
            inv = 1.0 / factor
            if abs(ratio - inv) / inv <= 0.35:
                return raw_price * factor
        if ratio > 5.0 or ratio < 0.2:
            return actual_price
        return raw_price

    results = []
    for ticker, pred in predictions.items():
        if ticker not in actual_prices:
            continue

        tracker_row = (
            tracker_by_ticker.get(ticker, {})
            if isinstance(tracker_by_ticker, dict)
            else {}
        )

        pred_price_at_prediction = float(pred.get("current_price", 0) or 0)
        pred_return_pct = _normalize_predicted_return_pct(
            pred.get("predicted_return", 0)
        )
        actual_price = float(actual_prices[ticker] or 0)

        if pred_price_at_prediction <= 0 or actual_price <= 0:
            continue

        pred_price_at_prediction = _rescale_logged_price(
            pred_price_at_prediction, actual_price
        )

        actual_return = 0.0
        actual_return_pct = 0.0

        # Direction check uses market-open baseline vs strategy-open prediction.
        pred_dir = (
            "UP" if pred_return_pct > 0 else "DOWN" if pred_return_pct < 0 else "FLAT"
        )
        strategy_price_at_open = float(
            tracker_row.get(
                "strategy_price_at_open",
                pred.get("strategy_price_at_open", pred.get("predicted_price", 0)),
            )
            or 0
        )
        ai_last_prediction = float(
            tracker_row.get(
                "ai_last_prediction",
                pred.get("ai_last_prediction", 0),
            )
            or 0
        )
        ai_source_hint = str(
            tracker_row.get("ai_source") or pred.get("ai_source") or "none"
        ).lower()
        open_price = float(
            tracker_row.get(
                "open_price", pred.get("open_price", pred_price_at_prediction)
            )
            or pred_price_at_prediction
        )
        open_price = _rescale_logged_price(open_price, actual_price)
        strategy_price_at_open = _rescale_logged_price(
            strategy_price_at_open, actual_price
        )
        if ai_last_prediction > 0:
            ai_last_prediction = _rescale_logged_price(ai_last_prediction, actual_price)
        if (
            ai_last_prediction > 0
            and abs(ai_last_prediction - strategy_price_at_open) <= 0.01
            and ai_source_hint in {"", "none", "fallback", "strategy_fallback"}
        ):
            ai_last_prediction = 0.0
        if actual_price > 0 and open_price <= 0:
            open_price = pred_price_at_prediction
        if actual_price > 0 and strategy_price_at_open <= 0:
            strategy_price_at_open = 0
        if actual_price > 0 and ai_last_prediction <= 0:
            ai_last_prediction = 0
        if open_price > 0 and (
            strategy_price_at_open <= 0 or strategy_price_at_open > open_price * 5
        ):
            strategy_price_at_open = round(
                open_price * (1 + pred_return_pct / 100.0), 2
            )
        if strategy_price_at_open <= 0:
            strategy_price_at_open = (
                round(open_price, 2)
                if open_price > 0
                else round(float(pred.get("predicted_price", 0) or 0), 2)
            )
        if ai_last_prediction <= 0 or ai_last_prediction > pred_price_at_prediction * 5:
            ai_last_prediction = 0.0
        if open_price > 0:
            actual_return = (actual_price - open_price) / open_price
            actual_return_pct = actual_return * 100
            strategy_return_pct = (
                (strategy_price_at_open - open_price) / open_price * 100
                if strategy_price_at_open > 0
                else 0.0
            )
            ai_return_pct = (
                (ai_last_prediction - open_price) / open_price * 100
                if ai_last_prediction > 0
                else None
            )
        else:
            actual_return = 0.0
            actual_return_pct = 0.0
            strategy_return_pct = 0.0
            ai_return_pct = None
        beta_val = _safe_float(
            tracker_row.get("beta")
            or (pred.get("fundamentals", {}) if isinstance(pred, dict) else {}).get(
                "beta"
            )
            or ALPHA_DEFAULT_BETA
        )
        if not np.isfinite(beta_val) or abs(beta_val) < 1e-9:
            beta_val = ALPHA_DEFAULT_BETA

        if "strategy_return_pct" in tracker_row:
            tracker_strategy_ret = _safe_float(
                tracker_row.get("strategy_return_pct", strategy_return_pct)
            )
            if abs(tracker_strategy_ret) <= 200:
                strategy_return_pct = tracker_strategy_ret
        if (
            "ai_return_pct" in tracker_row
            and tracker_row.get("ai_return_pct") is not None
        ):
            tracker_ai_ret = _safe_float(tracker_row.get("ai_return_pct"))
            if abs(tracker_ai_ret) <= 200:
                ai_return_pct = tracker_ai_ret
        if ai_last_prediction <= 0:
            ai_return_pct = None
        elif ai_return_pct is not None and abs(ai_return_pct) > 200:
            ai_return_pct = None
        strategy_error_pct = actual_return_pct - strategy_return_pct
        actual_alpha = _compute_alpha_metrics(
            actual_return_pct,
            benchmark_return_pct,
            beta=beta_val,
        )
        strategy_alpha = _compute_alpha_metrics(
            strategy_return_pct,
            benchmark_return_pct,
            beta=beta_val,
        )

        strategy_predicted_at_open = _normalize_open_window_timestamp(
            tracker_row.get("strategy_predicted_at_open")
            or pred.get("strategy_predicted_at_open")
            or pred.get("timestamp"),
            date_hint=date_str,
            default_offset_minutes=5,
        )
        ai_predicted_at_open = _normalize_open_window_timestamp(
            tracker_row.get("ai_predicted_at_open")
            or pred.get("ai_predicted_at_open")
            or pred.get("timestamp"),
            date_hint=date_str,
            default_offset_minutes=7,
        )
        ai_last_prediction_at = (
            tracker_row.get("ai_last_prediction_at")
            or pred.get("ai_last_prediction_at")
            or pred.get("timestamp")
        )
        strategy_direction = tracker_row.get(
            "strategy_direction_at_open"
        ) or _direction_from_prices(open_price, strategy_price_at_open)
        ai_direction = tracker_row.get("ai_direction_last") or (
            _direction_from_prices(open_price, ai_last_prediction)
            if ai_last_prediction > 0
            else "N/A"
        )
        actual_dir = _direction_from_prices(open_price, actual_price)
        strategy_vs_actual = strategy_direction == actual_dir
        direction_comparison = bool(strategy_vs_actual)
        if tracker_row.get("direction_comparison") in (True, False):
            direction_comparison = bool(tracker_row.get("direction_comparison"))
        strategy_vs_ai_direction = tracker_row.get("strategy_vs_ai_direction")
        if strategy_vs_ai_direction not in (True, False):
            strategy_vs_ai_direction = (
                strategy_direction == ai_direction if ai_last_prediction > 0 else None
            )
        predicted_price = float(pred.get("predicted_price", 0) or 0)
        if predicted_price <= 0 or predicted_price > pred_price_at_prediction * 5:
            predicted_price = round(
                pred_price_at_prediction * (1 + pred_return_pct / 100.0), 2
            )
        predicted_price = _rescale_logged_price(predicted_price, actual_price)
        strategy_vs_actual_price_diff = (
            actual_price - strategy_price_at_open if strategy_price_at_open > 0 else 0.0
        )
        strategy_vs_actual_pct = (
            (strategy_vs_actual_price_diff / open_price * 100)
            if open_price > 0
            else 0.0
        )
        market_open_price = (
            round(open_price, 2)
            if open_price > 0
            else round(pred_price_at_prediction, 2)
        )
        strategy_price_display = (
            round(strategy_price_at_open, 2)
            if strategy_price_at_open > 0
            else market_open_price
        )

        results.append(
            {
                "ticker": ticker,
                "name": ticker_names.get(ticker, _clean_name(ticker)),
                "signal": pred.get("signal", "N/A"),
                "predicted_return_pct": round(pred_return_pct, 3),
                "predicted_price": round(predicted_price, 2),
                "actual_price": round(actual_price, 2),
                "close_price": round(actual_price, 2),
                "actual_close": round(actual_price, 2),
                "actual_return_pct": round(actual_return_pct, 3),
                "strategy_return_pct": round(strategy_return_pct, 3),
                "ai_return_pct": (
                    round(ai_return_pct, 3) if ai_return_pct is not None else None
                ),
                "market_open_price": market_open_price,
                "open_price": market_open_price,
                "strategy_vs_actual_price_diff": round(
                    strategy_vs_actual_price_diff, 2
                ),
                "strategy_vs_actual_pct": round(strategy_vs_actual_pct, 3),
                "strategy_vs_actual_error_pct": round(strategy_error_pct, 3),
                "direction_predicted": pred_dir,
                "direction_actual": actual_dir,
                "strategy_direction_at_open": strategy_direction,
                "ai_direction_last": ai_direction,
                "strategy_vs_ai_direction": strategy_vs_ai_direction,
                "direction_comparison": direction_comparison,
                "direction_correct": direction_comparison,
                "last_prediction_basis": "strategy_vs_actual",
                "strategy_price_at_open": strategy_price_display,
                "ai_last_prediction": (
                    round(ai_last_prediction, 2) if ai_last_prediction > 0 else None
                ),
                "strategy_predicted_at_open": strategy_predicted_at_open,
                "strategy_predicted_at_open_display": _format_ist_timestamp(
                    strategy_predicted_at_open
                ),
                "ai_predicted_at_open": ai_predicted_at_open,
                "ai_predicted_at_open_display": _format_ist_timestamp(
                    ai_predicted_at_open
                ),
                "ai_last_prediction_at": ai_last_prediction_at,
                "ai_last_prediction_at_display": _format_ist_timestamp(
                    ai_last_prediction_at
                ),
                "alpha_pct": round(actual_alpha["simplified_alpha_pct"], 3),
                "alpha_capm_pct": round(actual_alpha["capm_alpha_pct"], 3),
                "strategy_alpha_expected_pct": round(
                    strategy_alpha["simplified_alpha_pct"], 3
                ),
                "strategy_alpha_expected_capm_pct": round(
                    strategy_alpha["capm_alpha_pct"], 3
                ),
                "benchmark_alpha_pct": round(actual_alpha["simplified_alpha_pct"], 3),
                "benchmark_return_pct": round(benchmark_return_pct, 3),
                "beta_used": round(actual_alpha["beta_used"], 4),
                "risk_free_daily_pct": round(actual_alpha["risk_free_daily_pct"], 4),
                "confidence": pred.get("confidence", 50),
                "checked_at": tracker_row.get("checked_at"),
                "market_status": market_status,
            }
        )

    if not results:
        return jsonify({"error": "Could not fetch actuals", "date": date_str})

    # Summary
    total = len(results)
    hits = sum(1 for r in results if r["direction_comparison"])
    hit_rate = (hits / total * 100) if total > 0 else 0
    avg_alpha = float(np.mean([r["alpha_pct"] for r in results]))
    total_alpha = float(np.sum([r["alpha_pct"] for r in results]))
    avg_confidence = float(np.mean([r["confidence"] for r in results]))

    return jsonify(
        {
            "date": date_str,
            "total_predictions": total,
            "direction_hits": hits,
            "hit_rate_pct": round(hit_rate, 1),
            "avg_alpha_pct": round(avg_alpha, 3),
            "total_alpha_pct": round(total_alpha, 3),
            "benchmark_return_pct": round(benchmark_return_pct, 3),
            "avg_confidence": round(avg_confidence, 1),
            "schema_version": PredictionTracker.SCHEMA_VERSION,
            "market_status": market_status,
            "results": sorted(results, key=lambda x: abs(x["alpha_pct"]), reverse=True),
        }
    )


@app.route("/api/prediction-dates")
def api_prediction_dates():
    """List all dates that have prediction logs."""
    if not PREDICTION_LOG_DIR.exists():
        return jsonify([])
    dates = sorted(
        [f.stem for f in PREDICTION_LOG_DIR.glob("*.json")],
        reverse=True,
    )
    return jsonify(dates)


@app.route("/api/ensemble-weights")
def api_ensemble_weights():
    """Return current ensemble model weights."""
    if predictor and predictor.ensemble:
        weights = predictor.ensemble.learned_weights or {}
        return jsonify(weights)
    return jsonify({"error": "Ensemble not loaded"}), 503


# ---------------------------------------------------------------------------
# Routes — Groq AI Explainer
# ---------------------------------------------------------------------------
@app.route("/api/explain", methods=["POST"])
def api_explain():
    """Get AI explanation for a metric/indicator/greek."""
    data = request.get_json()
    if not data:
        return jsonify({"error": "No data provided"}), 400

    explain_type = data.get("type", "")
    metric = data.get("metric", "")
    value = data.get("value", "")
    ticker = data.get("ticker", "")
    stock_name = data.get("stock_name", "")

    try:
        if explain_type == "fundamental":
            explanation = explain_fundamental(metric, value, ticker, stock_name)
        elif explain_type == "greek":
            explanation = explain_greek(
                metric, float(value) if value else 0, "call", ticker, stock_name
            )
        elif explain_type == "indicator":
            explanation = explain_indicator(metric, value, None, ticker, stock_name)
        else:
            explanation = f"Unknown explanation type: {explain_type}"

        return jsonify({"explanation": explanation})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/strategy/<ticker>")
def api_strategy(ticker: str):
    """Get Groq AI strategy analysis for a ticker."""
    if not models_loaded or predictor is None:
        return jsonify({"error": "Models still loading..."}), 503

    try:
        # Get prediction data (from cache if available)
        pred = predictor.predict_single(ticker, use_cache=True)
        if pred is None:
            return jsonify({"error": f"No prediction for {ticker}"}), 404

        stock_name = ticker_names.get(ticker, _clean_name(ticker))

        # Get Groq strategy
        groq_strat = get_groq_strategy(ticker, stock_name, pred)

        # Get combined strategy
        combined = get_combined_strategy(ticker, stock_name, pred, groq_strat)

        return jsonify(
            {
                "groq_strategy": groq_strat,
                "combined_strategy": combined,
            }
        )
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/overview/<ticker>")
def api_stock_overview(ticker: str):
    """Get comprehensive stock overview with company info and sentiment."""
    if not models_loaded or predictor is None:
        return jsonify({"error": "Models still loading..."}), 503

    try:
        # Get prediction data (from cache if available)
        pred = predictor.predict_single(ticker, use_cache=True)
        stock_name = ticker_names.get(ticker, _clean_name(ticker))

        fundamentals = pred.get("fundamentals", {}) if pred else {}
        current_price = pred.get("current_price", 0) if pred else 0
        signal = pred.get("signal", "") if pred else ""

        # Get stock overview
        overview = get_stock_overview(
            ticker, stock_name, fundamentals, current_price, signal
        )

        # Get news sentiment
        sentiment = get_news_sentiment(ticker, stock_name)

        return jsonify(
            {
                "overview": overview,
                "sentiment": sentiment,
                "stock_name": stock_name,
                "ticker": ticker,
            }
        )
    except Exception as e:
        log.warning(f"Overview error for {ticker}: {e}")
        return jsonify({"error": str(e)}), 500


# ---------------------------------------------------------------------------
# Routes — Prediction Tracking
# ---------------------------------------------------------------------------
@app.route("/api/tracking/daily")
def api_tracking_daily():
    """Get daily prediction tracking summary."""
    date_str = request.args.get("date")
    summary = PredictionTracker.get_daily_summary(date_str)
    return jsonify(summary)


@app.route("/api/tracking/check")
def api_tracking_check():
    """Check outcomes for a date's predictions."""
    date_str = request.args.get("date")
    results = PredictionTracker.check_outcomes(date_str)
    return jsonify(results)


@app.route("/api/tracking/monthly")
def api_tracking_monthly():
    """Get monthly accuracy report."""
    year = request.args.get("year", type=int)
    month = request.args.get("month", type=int)
    report = PredictionTracker.get_monthly_report(year, month)
    return jsonify(report)


@app.route("/api/tracking/history")
def api_tracking_history():
    """Get running monthly accuracy history."""
    history = PredictionTracker.get_accuracy_history()
    return jsonify(history)


@app.route("/api/tracking/feedback")
def api_tracking_feedback():
    """Get training feedback data for model improvement."""
    feedback = PredictionTracker.get_training_feedback_data()
    return jsonify(feedback)


@app.route("/api/training-feedback")
def api_training_feedback():
    """Alias endpoint for retraining export in a stable schema."""
    feedback = PredictionTracker.get_training_feedback_data()
    return jsonify(feedback)


# ---------------------------------------------------------------------------
# Routes — Daily Analysis (Enhanced Dashboard)
# ---------------------------------------------------------------------------
# In-memory cache of opening prices captured at market open
_opening_prices: dict[str, dict] = {}
_opening_prices_date: str = ""
_daily_prediction_baseline: dict[str, dict] = {}
_daily_prediction_baseline_date: str = ""
_premarket_snapshot: dict = {}
_premarket_snapshot_lock = threading.Lock()
_prediction_snapshots: dict = {}
_prediction_snapshots_lock = threading.Lock()
MARKET_OPEN_HOUR = 9
MARKET_OPEN_MINUTE = 15

# Background-computed daily analysis cache (avoids blocking requests)
_daily_analysis_cache: dict = {}
_daily_analysis_cache_time: float = 0
_daily_analysis_lock = threading.Lock()
DAILY_ANALYSIS_TTL = max(
    30,
    int(os.environ.get("DAILY_ANALYSIS_TTL_SEC", "90")),
)  # seconds

# Price cache with TTL (shared across endpoints)
_price_cache: dict[str, dict] = {}
_price_cache_time: float = 0
PRICE_CACHE_TTL = max(10, int(os.environ.get("PRICE_CACHE_TTL_SEC", "30")))


def _get_prices_chunked(tickers: list[str], chunk_size: int = 25) -> dict:
    """Fetch prices in parallel-friendly chunks to avoid yfinance timeouts."""
    all_prices: dict = {}
    for i in range(0, len(tickers), chunk_size):
        if _fd_pressure_high():
            log.warning("Price chunking paused under high FD pressure")
            break
        chunk = tickers[i : i + chunk_size]
        try:
            prices = _get_live_prices_batch(chunk)
            all_prices.update(prices)
        except Exception as e:
            log.warning(f"Price chunk {i//chunk_size} failed: {e}")
    return all_prices


def _get_cached_prices() -> dict:
    """Return cached prices if fresh, otherwise refetch in chunks."""
    global _price_cache, _price_cache_time
    now = time.time()
    if _price_cache and (now - _price_cache_time) < PRICE_CACHE_TTL:
        return _price_cache
    if _fd_pressure_high() and _price_cache:
        return _price_cache
    prices = _get_prices_chunked(all_tickers)
    if prices:
        _price_cache = prices
        _price_cache_time = time.time()
    return _price_cache


def _capture_opening_prices():
    """Capture opening prices once at start of day (uses chunked fetch)."""
    global _opening_prices, _opening_prices_date
    today = datetime.now(IST).strftime("%Y-%m-%d")
    if _opening_prices_date == today and _opening_prices:
        return  # Already captured today

    log.info("Capturing opening prices for today...")
    prices = _get_cached_prices()
    captured_at = datetime.now(IST).isoformat()
    _opening_prices = {}
    for t, p in prices.items():
        _opening_prices[t] = {
            "open": p.get("open", p.get("price", 0)),
            "prev_close": p.get("prev_close", 0),
            "captured_at": captured_at,
        }
    _opening_prices_date = today
    log.info(f"Captured opening prices for {len(_opening_prices)} tickers")


def _market_open_dt(now: datetime) -> datetime:
    return now.replace(
        hour=MARKET_OPEN_HOUR,
        minute=MARKET_OPEN_MINUTE,
        second=0,
        microsecond=0,
    )


def _premarket_cutoff_dt(now: datetime) -> datetime:
    return _market_open_dt(now) - timedelta(minutes=PREMARKET_MAX_BUFFER_MINUTES)


def _open_window_bounds(trading_day: date) -> tuple[datetime, datetime]:
    start = datetime(
        trading_day.year,
        trading_day.month,
        trading_day.day,
        MARKET_OPEN_HOUR,
        MARKET_OPEN_MINUTE,
        tzinfo=IST,
    )
    end = start + timedelta(minutes=15)  # 09:15 -> 09:30 IST window
    return start, end


def _premarket_window_start_dt(now: datetime) -> datetime:
    return now.replace(
        hour=PREMARKET_WINDOW_START_HOUR,
        minute=PREMARKET_WINDOW_START_MINUTE,
        second=0,
        microsecond=0,
    )


def _premarket_window_end_dt(now: datetime) -> datetime:
    return now.replace(
        hour=MARKET_LOCK_START_HOUR,
        minute=MARKET_LOCK_START_MINUTE,
        second=0,
        microsecond=0,
    )


def _market_lock_start_dt(now: datetime) -> datetime:
    return _premarket_window_end_dt(now)


def _market_lock_end_dt(now: datetime) -> datetime:
    return now.replace(
        hour=MARKET_LOCK_END_HOUR,
        minute=MARKET_LOCK_END_MINUTE,
        second=0,
        microsecond=0,
    )


def _prediction_window_type(now: datetime | None = None) -> str:
    """
    Returns one of:
      - premarket_open (09:15–09:30 IST)
      - market_open_locked (09:30–15:30 IST)
      - after_hours_live (15:30 IST onwards and pre-09:15)
    """
    ts = now or datetime.now(IST)
    if ts.weekday() >= 5:
        return "after_hours_live"
    if _premarket_window_start_dt(ts) <= ts < _premarket_window_end_dt(ts):
        return "premarket_open"
    if _market_lock_start_dt(ts) <= ts < _market_lock_end_dt(ts):
        return "market_open_locked"
    return "after_hours_live"


def _is_market_trade_window(now: datetime | None = None) -> bool:
    """BUY/SELL simulation triggers are allowed only during locked market hours."""
    ts = now or datetime.now(IST)
    if ts.weekday() >= 5:
        return False
    return _prediction_window_type(ts) == "market_open_locked"


def _load_prediction_snapshots_from_disk() -> dict:
    if not PREDICTION_SNAPSHOTS_FILE.exists():
        return {}
    try:
        data = json.loads(PREDICTION_SNAPSHOTS_FILE.read_text())
    except Exception:
        return {}
    if not isinstance(data, dict):
        return {}
    return data


def _save_prediction_snapshots_to_disk(payload: dict) -> None:
    PREDICTION_SNAPSHOTS_FILE.parent.mkdir(parents=True, exist_ok=True)
    PREDICTION_SNAPSHOTS_FILE.write_text(json.dumps(payload, indent=2, default=str))


def _derive_snapshot_item_fields(item: dict, snapshot_type: str) -> dict:
    row = dict(item or {})
    strategy_price = _safe_float(row.get("strategy_price_at_open"))
    if strategy_price <= 0:
        strategy_price = _safe_float(row.get("predicted_price"))
    predicted_return_pct = _safe_float(
        row.get("predicted_return_pct", row.get("predicted_return", 0))
    )
    predicted_return_decimal = predicted_return_pct / 100.0
    row["snapshot_type"] = snapshot_type
    row["source"] = str(row.get("source") or "strategy_engine")
    row["predicted_price"] = round(strategy_price, 2) if strategy_price > 0 else None
    row["predicted_return_decimal"] = round(predicted_return_decimal, 6)
    return row


def _normalize_prediction_snapshot(
    snapshot: dict, snapshot_type: str | None = None
) -> dict:
    if not isinstance(snapshot, dict):
        return {}
    out = dict(snapshot)
    inferred_type = str(snapshot_type or out.get("snapshot_type") or "premarket_open")
    out["snapshot_type"] = inferred_type
    out["schema_version"] = int(out.get("schema_version", SNAPSHOT_SCHEMA_VERSION))
    out["captured_at"] = out.get("captured_at") or datetime.now(IST).isoformat()
    out["source"] = str(out.get("source") or "strategy_engine")
    items = out.get("items", [])
    normalized_items = []
    for raw in items if isinstance(items, list) else []:
        if not isinstance(raw, dict):
            continue
        row = _derive_snapshot_item_fields(raw, inferred_type)
        normalized_items.append(row)
    out["items"] = normalized_items
    return out


def _to_snapshot_store(today: str, snapshots: dict[str, dict]) -> dict:
    return {
        "schema_version": SNAPSHOT_SCHEMA_VERSION,
        "date": today,
        "snapshots": snapshots,
    }


def _latest_available_snapshot(snapshots: dict[str, dict]) -> dict:
    for key in ("after_hours_live", "market_open_locked", "premarket_open"):
        snap = snapshots.get(key) if isinstance(snapshots, dict) else None
        if isinstance(snap, dict) and snap.get("items"):
            return snap
    return {}


def _build_prediction_snapshot(snapshot_type: str, force_live: bool = False) -> dict:
    now = datetime.now(IST)
    base = _normalize_premarket_snapshot(_build_premarket_snapshot())
    base["snapshot_type"] = snapshot_type
    base["schema_version"] = SNAPSHOT_SCHEMA_VERSION
    base["source"] = "strategy_engine"

    if snapshot_type == "premarket_open":
        capture_ts = _premarket_window_start_dt(now)
    elif snapshot_type == "market_open_locked":
        capture_ts = _market_lock_start_dt(now)
    else:
        capture_ts = now

    base["captured_at"] = capture_ts.isoformat()
    if force_live:
        base["captured_at_actual"] = now.isoformat()

    base["items"] = [
        _derive_snapshot_item_fields(item, snapshot_type)
        for item in base.get("items", [])
    ]
    return base


def _get_prediction_snapshot(
    *,
    force: bool = False,
    use_latest_stored: bool = False,
) -> dict:
    """
    Window-aware snapshot resolver:
      - premarket_open: 09:15–09:30 IST (frozen)
      - market_open_locked: 09:30–15:30 IST (frozen strategy values)
      - after_hours_live: manual refresh can update values
    """
    global _prediction_snapshots

    now = datetime.now(IST)
    today = now.strftime("%Y-%m-%d")
    target_type = _prediction_window_type(now)

    with _prediction_snapshots_lock:
        disk = _load_prediction_snapshots_from_disk()
        if (
            isinstance(disk, dict)
            and disk.get("date") == today
            and isinstance(disk.get("snapshots"), dict)
        ):
            _prediction_snapshots = disk
        elif not _prediction_snapshots or _prediction_snapshots.get("date") != today:
            _prediction_snapshots = _to_snapshot_store(today, {})

        snapshots = dict(_prediction_snapshots.get("snapshots", {}))

    if use_latest_stored:
        latest = _latest_available_snapshot(snapshots)
        if latest:
            return _normalize_prediction_snapshot(latest)

    if (
        target_type in {"premarket_open", "market_open_locked"}
        and not force
        and target_type in snapshots
        and snapshots[target_type].get("items")
    ):
        return _normalize_prediction_snapshot(snapshots[target_type], target_type)

    if target_type == "market_open_locked":
        # Market-open snapshot must be first computed at/after 09:30 and then frozen.
        built = _build_prediction_snapshot("market_open_locked")
    elif target_type == "premarket_open":
        # Premarket snapshot is fixed for 09:15–09:30 once created.
        if not force and snapshots.get("premarket_open", {}).get("items"):
            built = _normalize_prediction_snapshot(
                snapshots.get("premarket_open", {}), "premarket_open"
            )
        else:
            built = _build_prediction_snapshot("premarket_open")
    else:
        # After-hours AI view is intentionally live unless caller explicitly asks
        # for latest stored snapshot.
        if (
            not force
            and use_latest_stored
            and snapshots.get("after_hours_live", {}).get("items")
        ):
            built = _normalize_prediction_snapshot(
                snapshots.get("after_hours_live", {}), "after_hours_live"
            )
        else:
            built = _build_prediction_snapshot("after_hours_live", force_live=True)

    with _prediction_snapshots_lock:
        snapshots = dict(_prediction_snapshots.get("snapshots", {}))
        snapshots[target_type] = built
        _prediction_snapshots = _to_snapshot_store(today, snapshots)
        _save_prediction_snapshots_to_disk(_prediction_snapshots)
        return _normalize_prediction_snapshot(built, target_type)


def _next_day_prediction_switch_dt(now: datetime) -> datetime:
    return now.replace(
        hour=NEXT_DAY_PREDICTION_HOUR,
        minute=NEXT_DAY_PREDICTION_MINUTE,
        second=0,
        microsecond=0,
    )


def _is_next_day_prediction_window(now: datetime) -> bool:
    return now >= _next_day_prediction_switch_dt(now)


def _next_trading_day(d: date) -> date:
    out = d + timedelta(days=1)
    while out.weekday() >= 5:
        out += timedelta(days=1)
    return out


def _direction_from_prices(base_price: float, predicted_price: float) -> str:
    if base_price <= 0 or predicted_price <= 0:
        return "FLAT"
    if predicted_price > base_price:
        return "UP"
    if predicted_price < base_price:
        return "DOWN"
    return "FLAT"


def _parse_iso_datetime(value: str | None) -> datetime | None:
    if not value:
        return None
    try:
        parsed = datetime.fromisoformat(str(value))
    except Exception:
        return None
    if parsed.tzinfo is None:
        parsed = parsed.replace(tzinfo=IST)
    return parsed.astimezone(IST)


def _format_ist_timestamp(value: str | None) -> str | None:
    parsed = _parse_iso_datetime(value)
    if not parsed:
        return value if value else None
    return parsed.strftime("%d %b %Y, %I:%M:%S %p IST")


def _normalize_open_window_timestamp(
    value: str | None,
    *,
    date_hint: str | None = None,
    default_offset_minutes: int = 5,
    preserve_live_today: bool = False,
) -> str:
    """
    Clamp/derive an "at-open" timestamp into 09:15–09:30 IST for a trading day.
    """
    parsed = _parse_iso_datetime(value)
    if parsed is not None:
        trading_day = parsed.date()
    elif date_hint:
        try:
            trading_day = datetime.strptime(date_hint, "%Y-%m-%d").date()
        except Exception:
            trading_day = datetime.now(IST).date()
    else:
        trading_day = datetime.now(IST).date()

    start, end = _open_window_bounds(trading_day)
    offset = int(np.clip(default_offset_minutes, 0, 15))

    if parsed is None:
        return (start + timedelta(minutes=offset)).isoformat()
    now_ist = datetime.now(IST)
    # Preserve real-time timestamps for current-day live windows
    # (pre-09:15 and post-15:30) when explicitly requested.
    if (
        preserve_live_today
        and parsed.date() == now_ist.date()
        and _prediction_window_type(now_ist) == "after_hours_live"
    ):
        return parsed.isoformat()
    if parsed < start:
        return start.isoformat()
    if parsed > end:
        return end.isoformat()
    return parsed.isoformat()


def _get_premarket_row_for_ticker(ticker: str, date_str: str | None = None) -> dict:
    """
    Best-effort read of cached premarket row for a ticker (no live fetch).
    """
    target_date = date_str or datetime.now(IST).strftime("%Y-%m-%d")
    try:
        window_snapshot = _get_prediction_snapshot(use_latest_stored=True)
        if (
            window_snapshot
            and window_snapshot.get("date") == target_date
            and window_snapshot.get("items")
        ):
            for row in window_snapshot.get("items", []):
                if row.get("ticker") == ticker:
                    return row
    except Exception:
        pass

    snapshot: dict = {}
    with _premarket_snapshot_lock:
        if (
            _premarket_snapshot
            and _premarket_snapshot.get("date") == target_date
            and _premarket_snapshot.get("items")
        ):
            snapshot = dict(_premarket_snapshot)

    if not snapshot:
        disk_snapshot = _load_premarket_snapshot_from_disk()
        if (
            disk_snapshot
            and disk_snapshot.get("date") == target_date
            and disk_snapshot.get("items")
        ):
            snapshot = disk_snapshot

    snapshot = _normalize_premarket_snapshot(snapshot)

    for row in snapshot.get("items", []):
        if row.get("ticker") == ticker:
            return row
    return {}


def _load_premarket_snapshot_from_disk() -> dict:
    if not PREMARKET_OUTLOOK_FILE.exists():
        return {}
    try:
        data = json.loads(PREMARKET_OUTLOOK_FILE.read_text())
        if isinstance(data, dict):
            return data
    except Exception:
        pass
    return {}


def _save_premarket_snapshot_to_disk(snapshot: dict) -> None:
    PREMARKET_OUTLOOK_FILE.parent.mkdir(parents=True, exist_ok=True)
    PREMARKET_OUTLOOK_FILE.write_text(json.dumps(snapshot, indent=2, default=str))


def _normalize_premarket_snapshot(snapshot: dict) -> dict:
    if not isinstance(snapshot, dict):
        return {}
    out = dict(snapshot)
    snap_date = out.get("date") or datetime.now(IST).strftime("%Y-%m-%d")
    preserve_live_today = (
        _prediction_window_type(datetime.now(IST)) == "after_hours_live"
    )
    snap_captured_actual = out.get("captured_at_actual") or out.get("captured_at")
    snap_captured = _normalize_open_window_timestamp(
        out.get("captured_at") or snap_captured_actual,
        date_hint=snap_date,
        default_offset_minutes=20,
        preserve_live_today=preserve_live_today,
    )
    out["date"] = snap_date
    out["captured_at"] = snap_captured
    out["captured_at_actual"] = snap_captured_actual or snap_captured
    out["buffer_minutes"] = int(out.get("buffer_minutes", PREMARKET_MAX_BUFFER_MINUTES))

    parsed_actual = _parse_iso_datetime(out["captured_at_actual"])
    snapshot_type = str(out.get("snapshot_type", "")).strip()
    if snapshot_type not in {"market_open_live", "market_open_backfilled"}:
        if parsed_actual is not None:
            open_start, open_end = _open_window_bounds(parsed_actual.date())
            snapshot_type = (
                "market_open_backfilled"
                if parsed_actual > open_end
                else "market_open_live"
            )
            out["capture_cutoff"] = (
                out.get("capture_cutoff")
                or (open_start - timedelta(minutes=out["buffer_minutes"])).isoformat()
            )
        else:
            snapshot_type = "market_open"
    out["snapshot_type"] = snapshot_type

    if out.get("captured_within_buffer") not in (True, False):
        if parsed_actual is not None:
            out["captured_within_buffer"] = bool(
                parsed_actual <= _premarket_cutoff_dt(parsed_actual)
            )
        else:
            out["captured_within_buffer"] = True
    if not out.get("capture_note"):
        out["capture_note"] = (
            "Captured in pre-open buffer"
            if out["captured_within_buffer"]
            else "Backfilled from later session using market-open window timestamps"
        )

    rows = []
    for raw in out.get("items", []):
        if not isinstance(raw, dict):
            continue
        row = dict(raw)
        row_captured_actual = (
            row.get("captured_at_actual")
            or row.get("captured_at")
            or out["captured_at_actual"]
        )
        row_captured = _normalize_open_window_timestamp(
            row.get("captured_at") or row_captured_actual or snap_captured,
            date_hint=snap_date,
            default_offset_minutes=20,
            preserve_live_today=preserve_live_today,
        )
        row["captured_at"] = row_captured
        row["captured_at_actual"] = row_captured_actual or out["captured_at_actual"]
        row["strategy_predicted_at_open"] = _normalize_open_window_timestamp(
            row.get("strategy_predicted_at_open") or row_captured,
            date_hint=snap_date,
            default_offset_minutes=5,
            preserve_live_today=preserve_live_today,
        )
        row["ai_predicted_at_open"] = _normalize_open_window_timestamp(
            row.get("ai_predicted_at_open") or row_captured,
            date_hint=snap_date,
            default_offset_minutes=7,
            preserve_live_today=preserve_live_today,
        )
        try:
            ai_px = float(row.get("ai_predicted_price") or 0)
        except Exception:
            ai_px = 0.0
        row["strategy_source"] = row.get("strategy_source") or "ensemble_models"
        row["ai_source"] = row.get("ai_source") or (
            "groq_cache" if ai_px > 0 else "none"
        )
        if ai_px <= 0:
            row["ai_direction"] = "N/A"
        if row.get("strategy_vs_ai_direction") not in (True, False):
            if ai_px > 0:
                row["strategy_vs_ai_direction"] = (
                    str(row.get("strategy_direction", "FLAT")).upper()
                    == str(row.get("ai_direction", "N/A")).upper()
                )
            else:
                row["strategy_vs_ai_direction"] = None
        rows.append(row)
    out["items"] = rows
    return out


def _build_premarket_snapshot(tickers: list[str] | None = None) -> dict:
    """
    Build a premarket snapshot payload with strategy-open and AI-now predictions.
    Internal prices are numeric INR, returns are %.
    """
    if predictor is None:
        return {
            "date": datetime.now(IST).strftime("%Y-%m-%d"),
            "captured_at": datetime.now(IST).isoformat(),
            "items": [],
        }

    today = datetime.now(IST).strftime("%Y-%m-%d")
    scan_tickers = (
        tickers or tickers_by_sector.get("large_cap", [])[:PREMARKET_DEFAULT_TICKERS]
    )
    prices = _get_live_prices_batch(scan_tickers)
    captured_at_actual = datetime.now(IST).isoformat()
    captured_at_open_window = _normalize_open_window_timestamp(
        captured_at_actual,
        date_hint=today,
        default_offset_minutes=20,
        preserve_live_today=True,
    )
    rows: list[dict] = []

    for ticker in scan_tickers:
        quote = prices.get(ticker, {})
        current_price = float(quote.get("price", 0) or 0)
        open_price = float(quote.get("open", 0) or 0)
        if open_price <= 0:
            open_price = float(
                _opening_prices.get(ticker, {}).get("open", current_price)
                or current_price
            )
        if current_price <= 0:
            continue

        try:
            pred = predictor.predict_single(ticker, use_cache=True) or {}
        except Exception:
            pred = {}

        predicted_return_pct = _normalize_predicted_return_pct(
            pred.get("predicted_return", 0)
        )
        strategy_price_at_open = (
            round(open_price * (1 + predicted_return_pct / 100.0), 2)
            if open_price > 0
            else _safe_float(pred.get("predicted_price", 0))
        )
        ai_meta = _resolve_ai_forecast_price(
            ticker,
            open_price=open_price,
            strategy_price=strategy_price_at_open,
            current_price=current_price,
            allow_generate=False,
        )
        ai_predicted_price = _safe_float(ai_meta.get("price"))
        ai_direction = (
            _direction_from_prices(current_price, ai_predicted_price)
            if ai_predicted_price > 0
            else "N/A"
        )
        strategy_direction = _direction_from_prices(open_price, strategy_price_at_open)

        row = {
            "ticker": ticker,
            "name": ticker_names.get(ticker, _clean_name(ticker)),
            "current_price": round(current_price, 2),
            "open_price": round(open_price, 2) if open_price > 0 else None,
            "strategy_price_at_open": (
                round(strategy_price_at_open, 2) if strategy_price_at_open > 0 else None
            ),
            "ai_predicted_price": (
                round(ai_predicted_price, 2) if ai_predicted_price > 0 else None
            ),
            "strategy_direction": strategy_direction,
            "ai_direction": ai_direction,
            "predicted_return_pct": round(predicted_return_pct, 4),
            "predicted_return_decimal": round(predicted_return_pct / 100.0, 6),
            "signal": str(pred.get("signal", "HOLD")),
            "risk_reward": round(float(pred.get("risk_reward", 0) or 0), 3),
            "confidence": round(float(pred.get("confidence", 0) or 0), 2),
            "model_agreement": round(float(pred.get("model_agreement", 0) or 0), 2),
            "atr_pct": round(float(pred.get("atr_pct", 0) or 0), 3),
            "liquidity_factor": round(
                float(pred.get("liquidity_factor", 1.0) or 1.0), 3
            ),
            "weighted_sentiment_raw": round(
                float(pred.get("weighted_sentiment_raw", 0) or 0),
                4,
            ),
            "sentiment_articles": int(float(pred.get("sentiment_articles", 0) or 0)),
            "sentiment_weight_used": round(
                float(pred.get("sentiment_weight_used", 0) or 0),
                4,
            ),
            "sentiment_decay_weights": pred.get("sentiment_decay_weights", {}),
            "captured_at": captured_at_open_window,
            "captured_at_actual": captured_at_actual,
            "strategy_predicted_at_open": _normalize_open_window_timestamp(
                captured_at_open_window,
                date_hint=today,
                default_offset_minutes=5,
                preserve_live_today=True,
            ),
            "ai_predicted_at_open": _normalize_open_window_timestamp(
                ai_meta.get("generated_at_iso") or captured_at_open_window,
                date_hint=today,
                default_offset_minutes=7,
                preserve_live_today=True,
            ),
            "strategy_source": "ensemble_models",
            "ai_source": ai_meta.get("source", "none"),
        }
        rows.append(row)

        # Track for expected-vs-actual feedback.
        try:
            tracked = dict(pred)
            tracked.update(
                {
                    "open_price": row["open_price"] or row["current_price"],
                    "current_price": row["current_price"],
                    "strategy_price_at_open": row["strategy_price_at_open"]
                    or row["current_price"],
                    "ai_last_prediction": row["ai_predicted_price"] or 0,
                    "strategy_predicted_at_open": row["strategy_predicted_at_open"],
                    "ai_predicted_at_open": row["ai_predicted_at_open"],
                    "ai_last_prediction_at": ai_meta.get("generated_at_iso")
                    or captured_at_actual,
                    "strategy_direction_at_open": strategy_direction,
                    "ai_direction_last": ai_direction,
                    "snapshot_type": "premarket_open",
                    "strategy_vs_ai_direction": (
                        strategy_direction == ai_direction
                        if row["ai_predicted_price"] is not None
                        else None
                    ),
                }
            )
            PredictionTracker.record_prediction(ticker, tracked)
        except Exception as e:
            log.warning("Premarket tracking record failed for %s: %s", ticker, e)

    return _normalize_premarket_snapshot(
        {
            "date": today,
            "captured_at": captured_at_open_window,
            "captured_at_actual": captured_at_actual,
            "snapshot_type": "market_open",
            "items": rows,
        }
    )


def _capture_premarket_snapshot_if_due(force: bool = False) -> dict:
    """
    Capture premarket snapshot once per day, preferably before open-buffer cutoff.
    """
    global _premarket_snapshot
    now = datetime.now(IST)
    today = now.strftime("%Y-%m-%d")
    cutoff = _premarket_cutoff_dt(now)

    with _premarket_snapshot_lock:
        if (
            not force
            and _premarket_snapshot
            and _premarket_snapshot.get("date") == today
            and _premarket_snapshot.get("items")
        ):
            return _normalize_premarket_snapshot(dict(_premarket_snapshot))

        disk_snapshot = _load_premarket_snapshot_from_disk()
        if (
            not force
            and disk_snapshot
            and disk_snapshot.get("date") == today
            and disk_snapshot.get("items")
        ):
            _premarket_snapshot = _normalize_premarket_snapshot(dict(disk_snapshot))
            return dict(_premarket_snapshot)

    # Build outside lock (may call network/model inference).
    snapshot = _normalize_premarket_snapshot(_build_premarket_snapshot())
    open_window_end = _market_open_dt(now) + timedelta(minutes=15)
    captured_actual = snapshot.get("captured_at_actual") or now.isoformat()
    if _parse_iso_datetime(snapshot.get("captured_at")) is None:
        snapshot["captured_at"] = _normalize_open_window_timestamp(
            captured_actual,
            date_hint=today,
            default_offset_minutes=20,
        )
    snapshot["captured_at_actual"] = captured_actual
    snapshot["capture_cutoff"] = cutoff.isoformat()
    snapshot["buffer_minutes"] = PREMARKET_MAX_BUFFER_MINUTES
    snapshot["captured_within_buffer"] = now <= cutoff
    snapshot["snapshot_type"] = (
        "market_open_live" if now <= open_window_end else "market_open_backfilled"
    )
    snapshot["capture_note"] = (
        "Captured in pre-open buffer"
        if now <= cutoff
        else "Backfilled from later session using market-open window timestamps"
    )
    if not snapshot["captured_within_buffer"]:
        log.warning(
            "Premarket snapshot captured after cutoff. now=%s cutoff=%s buffer=%sm",
            now.isoformat(),
            cutoff.isoformat(),
            PREMARKET_MAX_BUFFER_MINUTES,
        )

    with _premarket_snapshot_lock:
        _premarket_snapshot = snapshot
        _save_premarket_snapshot_to_disk(snapshot)
        return dict(_premarket_snapshot)


def _normalize_predicted_return_pct(value) -> float:
    """Normalize return to a safe percent range [-50, 50]."""
    try:
        ret_pct = float(value)
    except (TypeError, ValueError):
        return 0.0

    if not np.isfinite(ret_pct):
        return 0.0

    if abs(ret_pct) > 500:
        ret_pct = ret_pct / 100.0

    return float(np.clip(ret_pct, -50.0, 50.0))


def _safe_float(value) -> float:
    try:
        out = float(value)
        if np.isfinite(out):
            return out
    except Exception:
        pass
    return 0.0


def _display_price_or_none(value, *, digits: int = 2) -> float | None:
    out = _safe_float(value)
    if out <= 0:
        return None
    return round(out, digits)


def _display_qty_or_none(value, *, digits: int = 4) -> float | None:
    out = _safe_float(value)
    if out <= 0:
        return None
    return round(out, digits)


def _as_bool(value, default: bool = False) -> bool:
    if value is None:
        return default
    if isinstance(value, bool):
        return value
    text = str(value).strip().lower()
    if text in {"1", "true", "yes", "y", "on"}:
        return True
    if text in {"0", "false", "no", "n", "off"}:
        return False
    return default


def _load_groq_forecast_cache() -> None:
    """Load persisted Groq price-forecast cache once per process."""
    global _groq_forecast_cache_loaded, _groq_forecast_cache, _groq_forecast_cache_time
    with _groq_forecast_cache_lock:
        if _groq_forecast_cache_loaded:
            return
        _groq_forecast_cache_loaded = True
        if not GROQ_FORECAST_CACHE_FILE.exists():
            return
        try:
            raw = json.loads(GROQ_FORECAST_CACHE_FILE.read_text())
        except Exception as exc:
            log.debug("Failed to read Groq forecast cache file: %s", exc)
            return

        if isinstance(raw, dict) and isinstance(raw.get("items"), dict):
            items = raw.get("items", {})
        elif isinstance(raw, dict):
            # Backward compatibility with legacy map format.
            items = raw
        else:
            items = {}

        loaded_count = 0
        now_ts = time.time()
        for ticker, entry in items.items():
            payload = None
            cached_at = 0.0
            if isinstance(entry, dict) and "payload" in entry:
                payload = entry.get("payload")
                cached_at = _safe_float(entry.get("cached_at"))
            else:
                payload = entry
            if not isinstance(payload, dict):
                continue
            ai_price = _safe_float(payload.get("ai_predicted_price"))
            if ai_price <= 0:
                continue
            if cached_at <= 0:
                generated_at_iso = payload.get("generated_at_iso")
                try:
                    cached_at = datetime.fromisoformat(
                        str(generated_at_iso)
                    ).timestamp()
                except Exception:
                    cached_at = now_ts
            t = str(ticker or "").upper()
            if not t:
                continue
            _groq_forecast_cache[t] = payload
            _groq_forecast_cache_time[t] = cached_at
            loaded_count += 1
        if loaded_count:
            log.info("Loaded %d cached Groq forecasts from disk", loaded_count)


def _persist_groq_forecast_cache_unlocked() -> None:
    items = {}
    for ticker, payload in _groq_forecast_cache.items():
        if not isinstance(payload, dict):
            continue
        items[ticker] = {
            "payload": payload,
            "cached_at": float(_groq_forecast_cache_time.get(ticker, time.time())),
        }
    data = {
        "schema_version": 1,
        "saved_at": datetime.now(IST).isoformat(),
        "items": items,
    }
    GROQ_FORECAST_CACHE_FILE.write_text(json.dumps(data, indent=2, default=str))


def _set_groq_forecast_cache_entry(ticker: str, payload: dict) -> None:
    """Set in-memory + persisted Groq forecast cache for one ticker."""
    _load_groq_forecast_cache()
    t = str(ticker or "").upper()
    if not t:
        return
    with _groq_forecast_cache_lock:
        _groq_forecast_cache[t] = dict(payload)
        _groq_forecast_cache_time[t] = time.time()
        try:
            _persist_groq_forecast_cache_unlocked()
        except Exception as exc:
            log.debug("Failed to persist Groq forecast cache for %s: %s", t, exc)


def _get_recent_legit_news_context(ticker: str, max_items: int = 6) -> str:
    """
    Pull recent Yahoo Finance-linked headlines with publisher names.
    Used to ground Groq prompts in identifiable sources.
    """
    key = str(ticker or "").strip().upper()
    if not key:
        return ""
    now_ts = time.time()
    with _news_context_cache_lock:
        cached = _news_context_cache.get(key)
    if isinstance(cached, dict):
        if (now_ts - float(cached.get("ts", 0) or 0)) < NEWS_CONTEXT_TTL:
            txt = str(cached.get("text", "") or "")
            if txt:
                return txt

    try:
        import yfinance as yf
    except Exception:
        return ""

    if _fd_pressure_high():
        return ""

    if not _yf_download_semaphore.acquire(timeout=YF_CALL_TIMEOUT_SEC):
        return ""
    try:
        news_items = yf.Ticker(key).news or []
    except Exception:
        return ""
    finally:
        _yf_download_semaphore.release()

    lines = []
    for row in news_items:
        if not isinstance(row, dict):
            continue
        title = str(row.get("title", "") or "").strip()
        if not title:
            continue
        publisher = str(
            row.get("publisher")
            or row.get("provider")
            or row.get("source")
            or "unknown"
        ).strip()
        published = row.get("providerPublishTime")
        try:
            pub_ts = datetime.fromtimestamp(float(published), tz=IST).strftime(
                "%Y-%m-%d %H:%M"
            )
        except Exception:
            pub_ts = ""
        lines.append(
            f"- [{publisher}] {title}" + (f" ({pub_ts} IST)" if pub_ts else "")
        )
        if len(lines) >= max_items:
            break
    text = "\n".join(lines)
    with _news_context_cache_lock:
        _news_context_cache[key] = {"ts": now_ts, "text": text}
    return text


def _get_cached_ai_forecast_meta(ticker: str) -> dict:
    _load_groq_forecast_cache()
    now = time.time()
    t = str(ticker or "").upper()
    with _groq_forecast_cache_lock:
        cached = _groq_forecast_cache.get(t, {})
        cached_at = _groq_forecast_cache_time.get(t, 0)
    if not cached:
        return {
            "available": False,
            "price": 0.0,
            "source": "none",
            "generated_at_iso": None,
        }
    ai_price = _safe_float(cached.get("ai_predicted_price"))
    if (now - cached_at) >= GROQ_FORECAST_TTL:
        return {
            "available": ai_price > 0,
            "price": round(ai_price, 2) if ai_price > 0 else 0.0,
            "source": "stale_cache" if ai_price > 0 else "stale",
            "generated_at_iso": cached.get("generated_at_iso"),
        }
    return {
        "available": ai_price > 0,
        "price": round(ai_price, 2) if ai_price > 0 else 0.0,
        "source": str(cached.get("ai_source", "groq")),
        "generated_at_iso": cached.get("generated_at_iso"),
    }


def _resolve_ai_forecast_price(
    ticker: str,
    *,
    open_price: float,
    strategy_price: float,
    current_price: float,
    allow_generate: bool = False,
) -> dict:
    """
    Resolve AI forecast price.
    - Uses cached Groq forecast first.
    - Optionally generates a fresh Groq forecast (small ticker sets only).
    - Never synthesizes AI values from strategy values.
    """
    cached_meta = _get_cached_ai_forecast_meta(ticker)
    if cached_meta.get("available"):
        return cached_meta

    if not allow_generate or not os.environ.get("GROQ_API_KEY"):
        return cached_meta

    strategy_px = _safe_float(strategy_price)
    open_px = _safe_float(open_price)
    current_px = _safe_float(current_price)
    if strategy_px <= 0:
        strategy_px = open_px if open_px > 0 else current_px
    if current_px <= 0:
        current_px = open_px if open_px > 0 else strategy_px
    if strategy_px <= 0 or current_px <= 0:
        return cached_meta

    stock_name = ticker_names.get(ticker, _clean_name(ticker))
    try:
        sentiment = get_news_sentiment(ticker, stock_name)
        legit_news_ctx = _get_recent_legit_news_context(ticker)
        if legit_news_ctx:
            sentiment = (
                (sentiment or "") + "\n\nRecent source headlines:\n" + legit_news_ctx
            )
        forecast = get_groq_price_forecast(
            ticker=ticker,
            stock_name=stock_name,
            open_price=open_px,
            strategy_predicted_price=strategy_px,
            current_price=current_px,
            sentiment_text=sentiment or "",
        )
        ai_price = _safe_float(forecast.get("ai_predicted_price"))
        if ai_price <= 0:
            return cached_meta
        gap = abs(ai_price - strategy_px) / max(strategy_px, 1e-9)
        if gap > AI_STRATEGY_GAP_WARN:
            log.warning(
                "AI/strategy price gap warning for %s: ai=%s strategy=%s gap=%.2f%%",
                ticker,
                round(ai_price, 2),
                round(strategy_px, 2),
                gap * 100.0,
            )
        if gap > AI_STRATEGY_GAP_BLOCK:
            log.warning(
                "Discarding AI forecast for %s due to excessive gap %.2f%%",
                ticker,
                gap * 100.0,
            )
            return {
                "available": False,
                "price": 0.0,
                "source": "filtered_outlier",
                "generated_at_iso": None,
            }

        generated_at_iso = datetime.now(IST).isoformat()
        payload = {
            "ticker": ticker,
            "name": stock_name,
            "strategy_predicted_price": round(strategy_px, 2),
            "current_price": round(current_px, 2),
            "ai_predicted_price": round(ai_price, 2),
            "open_price": round(open_px, 2) if open_px > 0 else None,
            "open_to_ai_predicted_pct": (
                round(((ai_price - open_px) / open_px * 100), 3)
                if open_px > 0
                else None
            ),
            "outlook": forecast.get("outlook", "Neutral"),
            "rationale": forecast.get("rationale", ""),
            "news_sentiment": sentiment,
            "ai_available": True,
            "ai_source": forecast.get("source", "groq"),
            "generated_at_iso": generated_at_iso,
            "generated_at": _format_ist_timestamp(generated_at_iso),
        }
        _set_groq_forecast_cache_entry(ticker, payload)
        return {
            "available": True,
            "price": round(ai_price, 2),
            "source": payload["ai_source"],
            "generated_at_iso": generated_at_iso,
        }
    except Exception as exc:
        log.debug("Groq AI forecast generation failed for %s: %s", ticker, exc)
        return cached_meta


def _get_prediction_for_ticker(
    ticker: str,
    fallback: dict | None = None,
    allow_live_refresh: bool = False,
    baseline_price: float = 0.0,
    freeze_baseline: bool = False,
) -> dict:
    """Get a robust prediction object from log or predictor cache/live output."""
    global _daily_prediction_baseline, _daily_prediction_baseline_date
    today = datetime.now(IST).strftime("%Y-%m-%d")
    if _daily_prediction_baseline_date != today:
        _daily_prediction_baseline = {}
        _daily_prediction_baseline_date = today

    if freeze_baseline and ticker in _daily_prediction_baseline:
        return dict(_daily_prediction_baseline[ticker])

    pred = dict(fallback or {})

    needs_refresh = (
        not pred
        or pred.get("predicted_price", 0) <= 0
        or abs(pred.get("predicted_return", 0) or 0) > 50
        or (
            abs(_safe_float(pred.get("confidence")) - 50.0) < 1e-9
            and abs(_safe_float(pred.get("model_agreement")) - 50.0) < 1e-9
            and not pred.get("model_predictions")
        )
    )

    if needs_refresh and allow_live_refresh and predictor is not None:
        try:
            refreshed = predictor.predict_single(ticker, use_cache=True)
            if refreshed:
                pred = refreshed
        except Exception:
            pass

    if not pred:
        pred = {
            "predicted_return": 0.0,
            "predicted_price": baseline_price,
            "signal": "HOLD",
            "confidence": 50.0,
            "model_agreement": 50.0,
            "risk_reward": 1.0,
            "model_predictions": {},
        }

    pred_return = _normalize_predicted_return_pct(pred.get("predicted_return", 0))
    pred["predicted_return"] = pred_return
    if freeze_baseline:
        _daily_prediction_baseline[ticker] = dict(pred)
    return pred


def _build_daily_analysis() -> dict:
    """Build the full daily analysis payload (called by background thread)."""
    global _daily_prediction_baseline, _daily_prediction_baseline_date
    _capture_opening_prices()
    current_prices = _get_cached_prices()
    now_ist = datetime.now(IST)
    market = get_market_status()
    market_status = market.get("status", "")
    window_type = _prediction_window_type(now_ist)
    freeze_baseline = window_type == "market_open_locked"
    today_str = now_ist.strftime("%Y-%m-%d")
    if _daily_prediction_baseline_date != today_str:
        _daily_prediction_baseline = {}
        _daily_prediction_baseline_date = today_str
    elif not freeze_baseline and _daily_prediction_baseline:
        # Outside lock window, allow prices/predictions to fluctuate on refresh.
        _daily_prediction_baseline = {}
    next_day_mode = (
        _is_next_day_prediction_window(now_ist) or market_status == "weekend"
    )
    prediction_mode = "next_day_after_close" if next_day_mode else "market_open_window"
    predicted_for_date = (
        _next_trading_day(now_ist.date()).strftime("%Y-%m-%d")
        if next_day_mode
        else now_ist.strftime("%Y-%m-%d")
    )
    prediction_generated_at = now_ist.isoformat()
    price_label = (
        "Close Price"
        if market_status in ("after_hours", "weekend")
        else "Current Price"
    )

    # Get all predictions (from cache)
    log_file = PREDICTION_LOG_DIR / f"{today_str}.json"
    predictions = {}
    if log_file.exists():
        try:
            with open(log_file) as f:
                predictions = json.load(f)
        except (json.JSONDecodeError, OSError):
            predictions = {}

    stocks = []
    prediction_count = 0
    refresh_budget = max(
        0,
        int(os.environ.get("DAILY_ANALYSIS_REFRESH_BUDGET", "40")),
    )
    for ticker in all_tickers:
        if ticker.startswith("^") or ticker in ("USDINR=X", "GC=F", "CL=F"):
            continue

        curr = current_prices.get(ticker, {})
        opening = _opening_prices.get(ticker, {})

        current_price = curr.get("price", 0)
        open_price = opening.get("open", curr.get("open", 0))
        prev_close = opening.get("prev_close", curr.get("prev_close", 0))
        if open_price <= 0:
            open_price = current_price
        pred = _get_prediction_for_ticker(
            ticker,
            predictions.get(ticker, {}),
            allow_live_refresh=refresh_budget > 0,
            baseline_price=open_price if open_price > 0 else current_price,
            freeze_baseline=freeze_baseline,
        )
        if refresh_budget > 0 and ticker not in predictions:
            refresh_budget -= 1
        predicted_return = _normalize_predicted_return_pct(
            pred.get("predicted_return", 0)
        )
        signal = pred.get("signal", "HOLD")
        confidence = float(pred.get("confidence", 50) or 50)
        premarket_row = _get_premarket_row_for_ticker(ticker, today_str)
        strategy_price_at_open = float(
            premarket_row.get(
                "strategy_price_at_open",
                pred.get("strategy_price_at_open", 0),
            )
            or 0
        )
        if strategy_price_at_open <= 0 and open_price > 0:
            strategy_price_at_open = round(
                open_price * (1 + predicted_return / 100.0), 2
            )
        if strategy_price_at_open <= 0:
            strategy_price_at_open = float(current_price or 0)

        ai_meta_open = _resolve_ai_forecast_price(
            ticker,
            open_price=float(open_price or 0),
            strategy_price=float(strategy_price_at_open or 0),
            current_price=float(current_price or 0),
            allow_generate=False,
        )
        ai_price_at_open = float(
            premarket_row.get("ai_predicted_price") or ai_meta_open.get("price") or 0
        )
        ai_source_open = str(
            premarket_row.get("ai_source") or ai_meta_open.get("source", "none")
        )

        strategy_predicted_at_open = _normalize_open_window_timestamp(
            premarket_row.get("strategy_predicted_at_open")
            or premarket_row.get("captured_at")
            or pred.get("strategy_predicted_at_open")
            or pred.get("timestamp"),
            date_hint=today_str,
            default_offset_minutes=5,
        )
        ai_predicted_at_open = _normalize_open_window_timestamp(
            premarket_row.get("ai_predicted_at_open")
            or premarket_row.get("captured_at")
            or pred.get("ai_predicted_at_open")
            or pred.get("timestamp"),
            date_hint=today_str,
            default_offset_minutes=7,
        )

        if next_day_mode:
            reference_price = float(current_price or open_price or 0)
            strategy_predicted_price = (
                round(reference_price * (1 + predicted_return / 100.0), 2)
                if reference_price > 0
                else strategy_price_at_open
            )
            ai_meta_now = _resolve_ai_forecast_price(
                ticker,
                open_price=float(open_price or 0),
                strategy_price=float(strategy_predicted_price or 0),
                current_price=float(current_price or 0),
                allow_generate=False,
            )
            ai_predicted_price = float(ai_meta_now.get("price") or 0)
            strategy_predicted_at = prediction_generated_at
            ai_predicted_at = ai_meta_now.get("generated_at_iso")
            ai_source = str(ai_meta_now.get("source", ai_source_open))
            prediction_context = "next_day"
        else:
            reference_price = float(open_price or current_price or 0)
            strategy_predicted_price = float(strategy_price_at_open or 0)
            ai_predicted_price = float(ai_price_at_open or 0)
            strategy_predicted_at = strategy_predicted_at_open
            ai_predicted_at = ai_meta_open.get("generated_at_iso") or (
                ai_predicted_at_open if ai_predicted_price > 0 else None
            )
            ai_source = ai_source_open
            prediction_context = "market_open"

        if current_price <= 0 or open_price <= 0:
            continue

        prediction_count += 1

        # Calculate percentage changes
        open_to_current_pct = (
            ((current_price - open_price) / open_price * 100) if open_price > 0 else 0
        )
        open_to_predicted_pct = (
            ((strategy_predicted_price - open_price) / open_price * 100)
            if open_price > 0 and strategy_predicted_price > 0
            else 0
        )
        open_to_ai_predicted_pct = (
            ((ai_predicted_price - open_price) / open_price * 100)
            if open_price > 0 and ai_predicted_price > 0
            else None
        )
        reference_to_strategy_pct = (
            ((strategy_predicted_price - reference_price) / reference_price * 100)
            if reference_price > 0 and strategy_predicted_price > 0
            else 0
        )
        reference_to_ai_pct = (
            ((ai_predicted_price - reference_price) / reference_price * 100)
            if reference_price > 0 and ai_predicted_price > 0
            else None
        )
        close_to_strategy_pct = (
            ((strategy_predicted_price - current_price) / current_price * 100)
            if current_price > 0 and strategy_predicted_price > 0
            else 0
        )
        close_to_ai_pct = (
            ((ai_predicted_price - current_price) / current_price * 100)
            if current_price > 0 and ai_predicted_price > 0
            else None
        )
        prev_to_current_pct = (
            ((current_price - prev_close) / prev_close * 100) if prev_close > 0 else 0
        )

        # Composite score for ranking (strategy-only numerics; AI only explains).
        model_agreement = float(pred.get("model_agreement", 50) or 50)
        risk_reward = pred.get("risk_reward", 1.0) if pred.get("risk_reward") else 1.0
        liq_factor = _safe_float(pred.get("liquidity_factor"))
        liq_factor = liq_factor if liq_factor > 0 else 1.0
        ens_weights = pred.get("ensemble_weights", {})
        diversity = 0.3
        if isinstance(ens_weights, dict) and ens_weights:
            vals = np.array(
                [abs(_safe_float(v)) for v in ens_weights.values()],
                dtype=float,
            )
            if vals.size > 0 and np.isfinite(vals).all():
                s = float(vals.sum())
                if s > 0:
                    vals = vals / s
                    concentration = float(vals.max())
                    diversity = max(0.05, 1.0 - concentration)

        score = (
            abs(predicted_return / 100.0)
            * (max(0.0, min(100.0, confidence)) / 100.0)
            * (max(0.0, min(100.0, model_agreement)) / 100.0)
            * max(0.6, min(1.5, liq_factor))
            * (1.0 + min(max(_safe_float(risk_reward), 0.0), 5.0) / 5.0)
            * (0.7 + diversity)
            * 10000.0
        )

        model_preds = pred.get("model_predictions", {})

        stocks.append(
            {
                "ticker": ticker,
                "name": ticker_names.get(ticker, _clean_name(ticker)),
                "open_price": round(open_price, 2),
                "predicted_price": round(strategy_predicted_price, 2),
                "strategy_predicted_price": round(strategy_predicted_price, 2),
                "ai_predicted_price": (
                    round(ai_predicted_price, 2) if ai_predicted_price > 0 else None
                ),
                "strategy_price_at_open": round(strategy_price_at_open, 2),
                "ai_price_at_open": (
                    round(ai_price_at_open, 2) if ai_price_at_open > 0 else None
                ),
                "current_price": round(current_price, 2),
                "prev_close": round(prev_close, 2),
                "display_price_label": price_label,
                "open_to_current_pct": round(open_to_current_pct, 3),
                "open_to_predicted_pct": round(open_to_predicted_pct, 3),
                "open_to_ai_predicted_pct": (
                    round(open_to_ai_predicted_pct, 3)
                    if open_to_ai_predicted_pct is not None
                    else None
                ),
                "reference_price": round(reference_price, 2),
                "reference_to_strategy_pct": round(reference_to_strategy_pct, 3),
                "reference_to_ai_pct": (
                    round(reference_to_ai_pct, 3)
                    if reference_to_ai_pct is not None
                    else None
                ),
                "close_to_strategy_pct": round(close_to_strategy_pct, 3),
                "close_to_ai_pct": (
                    round(close_to_ai_pct, 3) if close_to_ai_pct is not None else None
                ),
                "prediction_mode": prediction_mode,
                "prediction_context": prediction_context,
                "predicted_for_date": predicted_for_date,
                "strategy_predicted_at": strategy_predicted_at,
                "ai_predicted_at": ai_predicted_at,
                "strategy_predicted_at_open": strategy_predicted_at_open,
                "ai_predicted_at_open": ai_predicted_at_open,
                "strategy_source": "ensemble_models",
                "ai_source": ai_source,
                "prev_to_current_pct": round(prev_to_current_pct, 3),
                "predicted_return": round(predicted_return, 3),
                "signal": signal,
                "confidence": round(confidence, 1),
                "model_agreement": round(model_agreement, 1),
                "risk_reward": round(risk_reward, 1) if risk_reward else None,
                "liquidity_factor": round(liq_factor, 3),
                "volume": curr.get("volume", 0),
                "high": curr.get("high", 0),
                "low": curr.get("low", 0),
                "change": curr.get("change", 0),
                "change_pct": curr.get("change_pct", 0),
                "composite_score": round(score, 2),
                "model_predictions": model_preds,
            }
        )

    stocks.sort(key=lambda x: x["composite_score"], reverse=True)
    top_10 = stocks[:10]

    gainers = [s for s in stocks if s["change_pct"] > 0]
    losers = [s for s in stocks if s["change_pct"] < 0]

    return {
        "date": today_str,
        "market_status": market_status,
        "prediction_mode": prediction_mode,
        "predicted_for_date": predicted_for_date,
        "prediction_generated_at": prediction_generated_at,
        "total_stocks": len(stocks),
        "prediction_coverage": {
            "with_predictions": prediction_count,
            "coverage_pct": round((prediction_count / max(len(stocks), 1)) * 100, 1),
        },
        "market_summary": {
            "gainers": len(gainers),
            "losers": len(losers),
            "unchanged": len(stocks) - len(gainers) - len(losers),
            "avg_change_pct": round(
                float(np.mean([s["change_pct"] for s in stocks])) if stocks else 0, 3
            ),
        },
        "top_10": top_10,
        "all_stocks": stocks,
        "cached_at": time.time(),
    }


def _daily_analysis_background_loop():
    """Background thread: recompute daily analysis every DAILY_ANALYSIS_TTL seconds."""
    global _daily_analysis_cache, _daily_analysis_cache_time
    while True:
        if not models_loaded or not all_tickers:
            time.sleep(5)
            continue
        if _fd_pressure_high():
            log.warning("Daily analysis refresh skipped under high FD pressure")
            time.sleep(min(30, DAILY_ANALYSIS_TTL))
            continue
        try:
            _get_prediction_snapshot(force=False, use_latest_stored=True)
            result = _build_daily_analysis()
            with _daily_analysis_lock:
                _daily_analysis_cache = result
                _daily_analysis_cache_time = time.time()
            log.info(
                f"Daily analysis cache refreshed — {result['total_stocks']} stocks"
            )
        except Exception as e:
            log.error(f"Daily analysis background error: {e}")
        time.sleep(DAILY_ANALYSIS_TTL)


def _run_scheduled_precompute(tag: str) -> None:
    """Warm critical snapshots around open/close so UI loads instantly."""
    global _daily_analysis_cache, _daily_analysis_cache_time
    if not models_loaded or predictor is None:
        return
    now = datetime.now(IST)
    token = set_groq_request_endpoint(f"scheduler:{tag}")
    try:
        if tag == "open_0928":
            _capture_opening_prices()
            _capture_premarket_snapshot_if_due(force=True)
            _get_prediction_snapshot(force=True, use_latest_stored=False)
        elif tag == "close_1531":
            _get_prediction_snapshot(force=True, use_latest_stored=False)

        result = _build_daily_analysis()
        with _daily_analysis_lock:
            _daily_analysis_cache = result
            _daily_analysis_cache_time = time.time()
        log.info("Scheduled precompute complete (%s) at %s", tag, now.isoformat())
    except Exception as exc:
        log.warning("Scheduled precompute failed (%s): %s", tag, exc)
    finally:
        reset_groq_request_endpoint(token)


def _scheduled_precompute_loop():
    """Cron-style scheduler: precompute at 09:28 and 15:31 IST (trading days)."""
    executed: set[str] = set()
    while True:
        now = datetime.now(IST)
        if now.weekday() < 5:
            slot = None
            if (
                now.hour == PRECOMPUTE_OPEN_HOUR
                and now.minute == PRECOMPUTE_OPEN_MINUTE
            ):
                slot = "open_0928"
            elif (
                now.hour == PRECOMPUTE_CLOSE_HOUR
                and now.minute == PRECOMPUTE_CLOSE_MINUTE
            ):
                slot = "close_1531"
            if slot:
                run_key = f"{now.strftime('%Y-%m-%d %H:%M')}::{slot}"
                if run_key not in executed:
                    executed.add(run_key)
                    _run_scheduled_precompute(slot)

        today_prefix = now.strftime("%Y-%m-%d")
        executed = {k for k in executed if k.startswith(today_prefix)}
        time.sleep(12)


def _auto_simulation_background_loop():
    """
    Background auto-simulation loop.
    Runs AUTO_CHECK logic during market trade window (09:30-15:30 IST).
    """
    while True:
        try:
            if not AUTO_SIMULATION_ENABLED:
                time.sleep(10)
                continue
            if _fd_pressure_high():
                time.sleep(min(20, AUTO_SIMULATION_INTERVAL_SEC))
                continue
            if not _is_market_trade_window():
                time.sleep(min(30, AUTO_SIMULATION_INTERVAL_SEC))
                continue

            state = _read_portfolio_sim_state()
            result = _run_sim_auto_check(
                state,
                auto_buy_enabled=AUTO_SIMULATION_AUTO_BUY,
                source="strategy_auto_loop",
            )
            _write_portfolio_sim_state(result["state"])
            if result["triggered_count"] > 0 or result["trailing_updates"]:
                log.info(
                    "Auto simulation: sells=%s buys=%s trailing_updates=%s",
                    len(result["events"]),
                    len(result["auto_buy_events"]),
                    len(result["trailing_updates"]),
                )
        except Exception as exc:
            log.warning("Auto simulation loop error: %s", exc)
        time.sleep(AUTO_SIMULATION_INTERVAL_SEC)


def _send_daily_report_email_async(report: dict) -> None:
    if not TRADE_EMAIL_ENABLED:
        return

    def _send():
        report_date = str(report.get("date", ""))
        subject = f"[Trading Bot] Daily Simulation Report {report_date}"
        body = (
            f"Daily simulation summary ({report_date})\n\n"
            f"Total transactions: {report.get('total_transactions', 0)}\n"
            f"Buy transactions: {report.get('buy_transactions', 0)}\n"
            f"Sell transactions: {report.get('sell_transactions', 0)}\n"
            f"Buy transaction cost: ₹{report.get('buy_transaction_cost', 0)}\n"
            f"Sell transaction cost: ₹{report.get('sell_transaction_cost', 0)}\n"
            f"Total transaction cost: ₹{report.get('total_transaction_cost', 0)}\n"
            f"Total paid for buys: ₹{report.get('total_paid_for_buys', 0)}\n"
            f"Total received from sells: ₹{report.get('total_received_from_sells', 0)}\n"
            f"Net cash flow: ₹{report.get('net_cash_flow', 0)}\n"
            f"Realized PnL: ₹{report.get('realized_pnl', 0)}\n"
            f"Report file: {report.get('report_file', '')}\n"
        )
        ok, msg = _send_email(subject, body)
        if not ok:
            log.warning("Daily report email failed: %s", msg)

    threading.Thread(target=_send, daemon=True).start()


def _daily_trade_report_scheduler_loop():
    """
    Generate one daily report after 15:45 IST and optionally email it.
    """
    while True:
        try:
            now = datetime.now(IST)
            today = now.strftime("%Y-%m-%d")
            due = now.weekday() < 5 and (
                now.hour > TRADE_DAILY_REPORT_HOUR_IST
                or (
                    now.hour == TRADE_DAILY_REPORT_HOUR_IST
                    and now.minute >= TRADE_DAILY_REPORT_MINUTE_IST
                )
            )
            if due:
                with _daily_trade_report_lock:
                    if today not in _daily_trade_report_dates_done:
                        report = _write_daily_trade_report(today)
                        _daily_trade_report_dates_done.add(today)
                        log.info(
                            "Daily trade report generated for %s at %s",
                            today,
                            now.isoformat(),
                        )
                        _send_daily_report_email_async(report)

            # Keep only today's marker in memory.
            with _daily_trade_report_lock:
                _daily_trade_report_dates_done.intersection_update({today})
        except Exception as exc:
            log.warning("Daily trade report scheduler error: %s", exc)
        time.sleep(20)


@app.route("/api/daily-analysis")
def api_daily_analysis():
    """
    Comprehensive daily analysis — served from background-computed cache.
    Never blocks on live price fetching; returns cached data instantly.
    """
    if not models_loaded or predictor is None:
        return jsonify({"error": "Models still loading..."}), 503

    with _daily_analysis_lock:
        if _daily_analysis_cache:
            return jsonify(_daily_analysis_cache)

    # Cache not ready yet — return a loading indicator
    return (
        jsonify(
            {"error": "Analysis is being computed, please retry in ~15 seconds..."}
        ),
        202,
    )


@app.route("/api/price-tracker/<ticker>")
def api_price_tracker(ticker: str):
    """
    Lightweight endpoint for individual stock price tracking.
    Uses cached prices first, falls back to single-ticker fetch.
    """
    if not models_loaded or predictor is None:
        return jsonify({"error": "Models still loading..."}), 503
    refresh_ai = request.args.get("refresh_ai", "").lower() in ("true", "1", "yes")

    # Try cached prices first (from background analysis), then live fetch
    curr = _price_cache.get(ticker, {})
    if not curr:
        prices = _get_live_prices_batch([ticker])
        curr = prices.get(ticker, {})
    if not curr:
        # Last resort: pull from daily analysis cache
        with _daily_analysis_lock:
            for s in _daily_analysis_cache.get("all_stocks", []):
                if s["ticker"] == ticker:
                    return jsonify(s)
        return jsonify({"error": f"Price not available for {ticker}"}), 404

    # Get opening price from cache
    opening = _opening_prices.get(ticker, {})
    open_price = opening.get("open", curr.get("open", 0))
    prev_close = opening.get("prev_close", curr.get("prev_close", 0))
    open_price_captured_at = opening.get("captured_at")

    # Get prediction from today's log
    today_str = datetime.now(IST).strftime("%Y-%m-%d")
    log_file = PREDICTION_LOG_DIR / f"{today_str}.json"
    pred = {}
    if log_file.exists():
        try:
            with open(log_file) as f:
                all_preds = json.load(f)
                pred = all_preds.get(ticker, {})
        except (json.JSONDecodeError, OSError):
            pass

    current_price = float(curr.get("price", 0) or 0)
    predicted_return = _normalize_predicted_return_pct(pred.get("predicted_return", 0))
    strategy_predicted_price = (
        round(open_price * (1 + predicted_return / 100.0), 2)
        if open_price > 0
        else pred.get("predicted_price", 0)
    )
    premarket_row = _get_premarket_row_for_ticker(ticker, today_str)
    strategy_price_at_open = float(
        premarket_row.get(
            "strategy_price_at_open",
            pred.get("strategy_price_at_open", strategy_predicted_price),
        )
        or 0
    )
    strategy_price_source = (
        "market_open_snapshot"
        if _safe_float(premarket_row.get("strategy_price_at_open")) > 0
        else "strategy_model"
    )
    if strategy_price_at_open <= 0 and open_price > 0 and abs(predicted_return) <= 50:
        strategy_price_at_open = round(open_price * (1 + predicted_return / 100.0), 2)
        strategy_price_source = "derived_from_open_return"
    if strategy_price_at_open <= 0:
        strategy_price_at_open = 0.0
        strategy_price_source = "unavailable"
    ai_meta_open = _resolve_ai_forecast_price(
        ticker,
        open_price=float(open_price or 0),
        strategy_price=float(strategy_price_at_open or 0),
        current_price=float(current_price or 0),
        # Keep this endpoint lightweight; explicit AI refresh is optional.
        allow_generate=refresh_ai and bool(os.environ.get("GROQ_API_KEY")),
    )
    ai_predicted_price = float(ai_meta_open.get("price") or 0)
    ai_open_price = float(
        premarket_row.get("ai_predicted_price") or ai_predicted_price or 0
    )
    signal = pred.get("signal", "")
    confidence = pred.get("confidence", 0)
    market = get_market_status()
    market_status = market.get("status", "")
    now_ist = datetime.now(IST)
    prediction_window = _prediction_window_type(now_ist)
    next_day_mode = (
        _is_next_day_prediction_window(now_ist) or market_status == "weekend"
    )
    prediction_mode = "next_day_after_close" if next_day_mode else "market_open_window"
    predicted_for_date = (
        _next_trading_day(now_ist.date()).strftime("%Y-%m-%d")
        if next_day_mode
        else now_ist.strftime("%Y-%m-%d")
    )
    is_market_closed = market_status in ("after_hours", "weekend")
    display_price_label = "Close Price" if is_market_closed else "Current Price"
    strategy_predicted_at_open = _normalize_open_window_timestamp(
        premarket_row.get("strategy_predicted_at_open")
        or premarket_row.get("captured_at")
        or pred.get("strategy_predicted_at_open")
        or pred.get("timestamp"),
        date_hint=today_str,
        default_offset_minutes=5,
    )
    ai_predicted_at_open = _normalize_open_window_timestamp(
        premarket_row.get("ai_predicted_at_open")
        or premarket_row.get("captured_at")
        or pred.get("ai_predicted_at_open")
        or pred.get("timestamp"),
        date_hint=today_str,
        default_offset_minutes=7,
    )
    current_strategy_predicted_at = (
        pred.get("timestamp")
        or pred.get("ai_last_prediction_at")
        or datetime.now(IST).isoformat()
    )
    current_ai_predicted_at = (
        ai_meta_open.get("generated_at_iso")
        or pred.get("ai_last_prediction_at")
        or (
            ai_predicted_at_open
            if (ai_open_price > 0 or ai_predicted_price > 0)
            else pred.get("timestamp")
        )
        or datetime.now(IST).isoformat()
    )
    live_strategy_return = predicted_return
    if prediction_window == "after_hours_live":
        try:
            # Cache-first to prevent expensive model reruns on 3s polling.
            live_pred = predictor.predict_single(ticker, use_cache=True) or {}
        except Exception:
            live_pred = {}
        live_strategy_return = _normalize_predicted_return_pct(
            live_pred.get("predicted_return", predicted_return)
        )
        signal = live_pred.get("signal", signal)
        confidence = live_pred.get("confidence", confidence)
        current_strategy_predicted_at = (
            live_pred.get("timestamp")
            or live_pred.get("generated_at_iso")
            or now_ist.isoformat()
        )

    if open_price <= 0:
        open_price = current_price
    if ai_open_price <= 0:
        ai_open_price = ai_predicted_price if ai_predicted_price > 0 else 0.0

    sim_positions = dict(_read_portfolio_sim_state().get("open_positions", {}))
    sim_pos = dict(sim_positions.get(ticker, {}) or {})
    entry_range_source = "none"
    entry_range_low = _safe_float(sim_pos.get("entry_range_low"))
    entry_range_high = _safe_float(sim_pos.get("entry_range_high"))
    if entry_range_low > 0 and entry_range_high > 0:
        entry_range_source = "sim_position"
    else:
        entry_range_low = _safe_float(premarket_row.get("entry_range_low"))
        entry_range_high = _safe_float(premarket_row.get("entry_range_high"))
        if entry_range_low > 0 and entry_range_high > 0:
            entry_range_source = "market_open_snapshot"
        elif strategy_price_at_open > 0:
            entry_range_low, entry_range_high = _entry_price_range(
                strategy_price_at_open
            )
            entry_range_source = "strategy_derived"
        else:
            entry_range_low = 0.0
            entry_range_high = 0.0
            entry_range_source = "unavailable"
    if (
        entry_range_high > 0
        and entry_range_low > 0
        and entry_range_high < entry_range_low
    ):
        entry_range_low, entry_range_high = entry_range_high, entry_range_low

    current_strategy_ref = current_price if current_price > 0 else open_price
    current_strategy_price = (
        round(current_strategy_ref * (1 + live_strategy_return / 100.0), 2)
        if current_strategy_ref > 0
        else strategy_price_at_open
    )
    ai_meta_live = {
        "available": ai_open_price > 0,
        "price": float(ai_open_price or 0),
        "source": str(
            premarket_row.get("ai_source") or ai_meta_open.get("source", "none")
        ),
        "generated_at_iso": ai_meta_open.get("generated_at_iso")
        or ai_predicted_at_open,
    }
    if prediction_window == "after_hours_live" or ai_open_price <= 0:
        ai_meta_live = _resolve_ai_forecast_price(
            ticker,
            open_price=float(open_price or 0),
            strategy_price=float(current_strategy_price or strategy_price_at_open or 0),
            current_price=float(current_price or 0),
            allow_generate=refresh_ai and bool(os.environ.get("GROQ_API_KEY")),
        )
    current_ai_price = float(ai_meta_live.get("price") or ai_open_price or 0)
    if prediction_window == "after_hours_live":
        current_ai_predicted_at = (
            ai_meta_live.get("generated_at_iso") or now_ist.isoformat()
        )

    if next_day_mode:
        reference_price = current_price if current_price > 0 else open_price
        next_day_strategy_price = (
            round(reference_price * (1 + live_strategy_return / 100.0), 2)
            if reference_price > 0
            else strategy_price_at_open
        )
        next_day_ai_price = (
            current_ai_price
            if current_ai_price > 0
            else float(ai_meta_open.get("price") or 0)
        )
        next_day_predicted_at = now_ist.isoformat()
        current_strategy_predicted_at = next_day_predicted_at
        current_ai_predicted_at = ai_meta_live.get("generated_at_iso") or None
        ai_source = str(ai_meta_live.get("source", ai_meta_open.get("source", "none")))
    else:
        reference_price = open_price
        if prediction_window == "after_hours_live":
            next_day_strategy_price = current_strategy_price
            next_day_ai_price = current_ai_price
            next_day_predicted_at = current_strategy_predicted_at
            ai_source = str(
                ai_meta_live.get("source", ai_meta_open.get("source", "none"))
            )
        else:
            next_day_strategy_price = strategy_price_at_open
            next_day_ai_price = ai_open_price
            next_day_predicted_at = strategy_predicted_at_open
            ai_source = str(
                premarket_row.get("ai_source") or ai_meta_open.get("source", "none")
            )

    open_to_current_pct = (
        ((current_price - open_price) / open_price * 100) if open_price > 0 else 0
    )
    open_to_predicted_pct = (
        ((strategy_price_at_open - open_price) / open_price * 100)
        if open_price > 0 and strategy_price_at_open > 0
        else 0
    )

    open_to_ai_predicted_pct = (
        ((ai_open_price - open_price) / open_price * 100)
        if open_price > 0 and ai_open_price > 0
        else None
    )
    close_to_strategy_pct = (
        ((next_day_strategy_price - current_price) / current_price * 100)
        if current_price > 0 and next_day_strategy_price > 0
        else None
    )
    close_to_ai_pct = (
        ((next_day_ai_price - current_price) / current_price * 100)
        if current_price > 0 and next_day_ai_price > 0
        else None
    )

    display_strategy_price = (
        next_day_strategy_price
        if prediction_window == "after_hours_live"
        else strategy_price_at_open
    )
    display_ai_price = (
        next_day_ai_price if prediction_window == "after_hours_live" else ai_open_price
    )
    strategy_display_predicted_at = (
        current_strategy_predicted_at
        if prediction_window == "after_hours_live"
        else strategy_predicted_at_open
    )
    ai_display_predicted_at = (
        current_ai_predicted_at
        if prediction_window == "after_hours_live"
        else ai_predicted_at_open
    )
    display_predicted_return = (
        live_strategy_return
        if prediction_window == "after_hours_live"
        else predicted_return
    )

    snapshot_now_iso = datetime.now(IST).isoformat()
    open_price_display = _display_price_or_none(open_price)
    market_open_price = _display_price_or_none(open_price)
    strategy_price_at_open_display = _display_price_or_none(strategy_price_at_open)
    return jsonify(
        {
            "ticker": ticker,
            "name": ticker_names.get(ticker, _clean_name(ticker)),
            "open_price": open_price_display,
            "market_open_price": market_open_price,
            "open_price_captured_at": open_price_captured_at,
            "open_price_captured_at_display": _format_ist_timestamp(
                open_price_captured_at
            ),
            "predicted_price": _display_price_or_none(display_strategy_price),
            "strategy_predicted_price": _display_price_or_none(display_strategy_price),
            "strategy_price_at_open": strategy_price_at_open_display,
            "strategy_price_source": strategy_price_source,
            "ai_predicted_price": (
                round(display_ai_price, 2) if display_ai_price > 0 else None
            ),
            "current_strategy_predicted_price": _display_price_or_none(
                next_day_strategy_price
            ),
            "current_ai_predicted_price": (
                round(next_day_ai_price, 2) if next_day_ai_price > 0 else None
            ),
            "current_price": _display_price_or_none(current_price),
            "close_price": (
                _display_price_or_none(current_price) if is_market_closed else None
            ),
            "entry_range_low": _display_price_or_none(entry_range_low),
            "entry_range_high": _display_price_or_none(entry_range_high),
            "entry_range_source": entry_range_source,
            "display_price_label": display_price_label,
            "market_status": market_status,
            "prediction_window": prediction_window,
            "prediction_mode": prediction_mode,
            "predicted_for_date": predicted_for_date,
            "prediction_reference_price": (
                round(reference_price, 2) if reference_price > 0 else None
            ),
            "next_day_strategy_predicted_price": (
                _display_price_or_none(next_day_strategy_price)
            ),
            "next_day_ai_predicted_price": (
                round(next_day_ai_price, 2) if next_day_ai_price > 0 else None
            ),
            "next_day_predicted_at": next_day_predicted_at,
            "next_day_predicted_at_display": _format_ist_timestamp(
                next_day_predicted_at
            ),
            "strategy_predicted_at_open": strategy_predicted_at_open,
            "strategy_predicted_at_open_display": _format_ist_timestamp(
                strategy_predicted_at_open
            ),
            "strategy_display_predicted_at": strategy_display_predicted_at,
            "strategy_display_predicted_at_display": _format_ist_timestamp(
                strategy_display_predicted_at
            ),
            "ai_predicted_at_open": ai_predicted_at_open,
            "ai_predicted_at_open_display": _format_ist_timestamp(ai_predicted_at_open),
            "ai_display_predicted_at": ai_display_predicted_at,
            "ai_display_predicted_at_display": _format_ist_timestamp(
                ai_display_predicted_at
            ),
            "current_strategy_predicted_at": current_strategy_predicted_at,
            "current_strategy_predicted_at_display": _format_ist_timestamp(
                current_strategy_predicted_at
            ),
            "current_ai_predicted_at": current_ai_predicted_at,
            "current_ai_predicted_at_display": _format_ist_timestamp(
                current_ai_predicted_at
            ),
            "current_snapshot_at": snapshot_now_iso,
            "current_snapshot_at_display": _format_ist_timestamp(snapshot_now_iso),
            "prev_close": _display_price_or_none(prev_close),
            "open_to_current_pct": round(open_to_current_pct, 3),
            "open_to_predicted_pct": round(open_to_predicted_pct, 3),
            "open_to_ai_predicted_pct": (
                round(open_to_ai_predicted_pct, 3)
                if open_to_ai_predicted_pct is not None
                else None
            ),
            "close_to_strategy_pct": (
                round(close_to_strategy_pct, 3)
                if close_to_strategy_pct is not None
                else None
            ),
            "close_to_ai_pct": (
                round(close_to_ai_pct, 3) if close_to_ai_pct is not None else None
            ),
            "predicted_return": round(display_predicted_return, 3),
            "signal": signal,
            "confidence": round(confidence, 1),
            "ai_forecast_available": bool(ai_open_price > 0 or next_day_ai_price > 0),
            "ai_forecast_source": (
                ai_source if (ai_open_price > 0 or next_day_ai_price > 0) else "none"
            ),
            "strategy_source": "ensemble_models",
            "volume": curr.get("volume", 0),
            "high": _display_price_or_none(curr.get("high", 0)),
            "low": _display_price_or_none(curr.get("low", 0)),
            "open": _display_price_or_none(curr.get("open", 0)),
            "change": curr.get("change", 0),
            "change_pct": curr.get("change_pct", 0),
        }
    )


@app.route("/api/news/<ticker>")
def api_news(ticker: str):
    """Get real-world news and sentiment for a specific stock."""
    try:
        stock_name = ticker_names.get(ticker, _clean_name(ticker))
        sentiment = get_news_sentiment(ticker, stock_name)
        return jsonify(
            {
                "ticker": ticker,
                "name": stock_name,
                "sentiment": sentiment,
                "generated_at": datetime.now(IST).strftime("%d %b %Y, %I:%M %p IST"),
            }
        )
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/groq-price-forecast/<ticker>")
def api_groq_price_forecast(ticker: str):
    """Return Groq-based AI price forecast using strategy + current context."""
    if not models_loaded or predictor is None:
        return jsonify({"error": "Models still loading..."}), 503

    _load_groq_forecast_cache()
    now = time.time()
    t = str(ticker or "").upper()
    with _groq_forecast_cache_lock:
        cached_payload = _groq_forecast_cache.get(t)
        cached_age = now - _groq_forecast_cache_time.get(t, 0)
    force = request.args.get("force", "").lower() in ("true", "1", "yes")
    if cached_payload and cached_age < GROQ_FORECAST_TTL:
        cached_ai = _safe_float(cached_payload.get("ai_predicted_price"))
        # Keep returning valid cached values, but do not pin an unavailable response
        # when a Groq key is available for regeneration.
        if cached_ai > 0 or not os.environ.get("GROQ_API_KEY"):
            return jsonify(cached_payload)
    if cached_payload and not force:
        cached_ai = _safe_float(cached_payload.get("ai_predicted_price"))
        if cached_ai > 0:
            stale_payload = dict(cached_payload)
            stale_payload["ai_source"] = str(stale_payload.get("ai_source") or "groq")
            stale_payload["ai_source"] = f"{stale_payload['ai_source']}:stale_cache"
            stale_payload["stale"] = True
            return jsonify(stale_payload)

    try:
        tracker = api_price_tracker(ticker)
        if isinstance(tracker, tuple):
            tracker_payload = tracker[0].get_json()
            status = tracker[1]
            if status != 200:
                return jsonify(tracker_payload), status
        else:
            tracker_payload = tracker.get_json()

        if tracker_payload.get("error"):
            return jsonify(tracker_payload), 404

        stock_name = ticker_names.get(ticker, _clean_name(ticker))
        sentiment = get_news_sentiment(ticker, stock_name)
        forecast = get_groq_price_forecast(
            ticker=ticker,
            stock_name=stock_name,
            open_price=float(tracker_payload.get("open_price") or 0),
            strategy_predicted_price=float(
                tracker_payload.get("strategy_predicted_price")
                or tracker_payload.get("predicted_price")
                or 0
            ),
            current_price=float(tracker_payload.get("current_price") or 0),
            sentiment_text=sentiment or "",
        )
        ai_raw = forecast.get("ai_predicted_price")
        try:
            ai_price = float(ai_raw) if ai_raw not in (None, "") else 0.0
        except Exception:
            ai_price = 0.0
        ai_available = ai_price > 0
        open_price = float(tracker_payload.get("open_price") or 0)
        open_to_ai_pct = (
            ((ai_price - open_price) / open_price * 100)
            if open_price > 0 and ai_available
            else None
        )

        generated_at_iso = datetime.now(IST).isoformat()
        payload = {
            "ticker": ticker,
            "name": stock_name,
            "strategy_predicted_price": tracker_payload.get("strategy_predicted_price"),
            "current_price": tracker_payload.get("current_price"),
            "ai_predicted_price": round(ai_price, 2) if ai_available else None,
            "open_price": tracker_payload.get("open_price"),
            "open_to_ai_predicted_pct": (
                round(open_to_ai_pct, 3) if open_to_ai_pct is not None else None
            ),
            "outlook": forecast.get("outlook", "Neutral"),
            "rationale": forecast.get("rationale", ""),
            "news_sentiment": sentiment,
            "ai_available": ai_available,
            "ai_source": forecast.get("source", "groq"),
            "generated_at_iso": generated_at_iso,
            "generated_at": _format_ist_timestamp(generated_at_iso),
        }
        _set_groq_forecast_cache_entry(ticker, payload)
        return jsonify(payload)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/score-explain/<ticker>")
def api_score_explain(ticker: str):
    """Return strategy-derived numeric score plus optional Groq explanation."""
    if not models_loaded or predictor is None:
        return jsonify({"error": "Models still loading..."}), 503

    try:
        pred = predictor.predict_single(ticker, use_cache=True)
        if not pred:
            return jsonify({"error": f"Prediction unavailable for {ticker}"}), 404
        score = float(LivePredictor._score_prediction(pred))
        context = (
            f"Ticker={ticker}, signal={pred.get('signal')}, "
            f"predicted_return_pct={pred.get('predicted_return')}, "
            f"confidence={pred.get('confidence')}, "
            f"model_agreement={pred.get('model_agreement')}, "
            f"risk_reward={pred.get('risk_reward')}."
        )
        explanation = explain_risk_term("Strategy Composite Score", context)
        return jsonify(
            {
                "ticker": ticker,
                "score": round(score, 4),
                "signal": pred.get("signal", "HOLD"),
                "source": "strategy_metrics",
                "explanation": explanation,
            }
        )
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/explain-risk-term")
def api_explain_risk_term():
    """Get Groq explanation for a risk metric/term."""
    term = request.args.get("term", "").strip()
    context = request.args.get("context", "").strip()
    if not term:
        return jsonify({"error": "term is required"}), 400
    try:
        explanation = explain_risk_term(term, context)
        return jsonify({"term": term, "explanation": explanation})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/debug/prediction-status/<ticker>")
def api_debug_prediction_status(ticker: str):
    """Debug cache/snapshot state for a ticker."""
    if predictor is None:
        return jsonify({"error": "Predictor not initialized"}), 503

    try:
        now = datetime.now(IST)
        window = _prediction_window_type(now)
        snapshot = _get_prediction_snapshot(use_latest_stored=True)
        snapshot_row = next(
            (row for row in snapshot.get("items", []) if row.get("ticker") == ticker),
            {},
        )
        cache_entry = predictor.cache.get(ticker)
        live_quote = _get_live_prices_batch([ticker]).get(ticker, {})

        formula_ok = None
        if cache_entry:
            cp = _safe_float(cache_entry.get("current_price"))
            ret_pct = _safe_float(cache_entry.get("predicted_return"))
            px = _safe_float(cache_entry.get("predicted_price"))
            if cp > 0 and px > 0:
                expected = round(cp * (1 + ret_pct / 100.0), 2)
                formula_ok = abs(px - expected) <= 0.02

        return jsonify(
            {
                "ticker": ticker,
                "timestamp": now.isoformat(),
                "market_status": get_market_status().get("status"),
                "prediction_window": window,
                "cache_ttl_seconds": getattr(predictor.cache, "ttl_seconds", None),
                "cache_hit": bool(cache_entry),
                "cache_entry": cache_entry,
                "snapshot_type": snapshot.get("snapshot_type"),
                "snapshot_captured_at": snapshot.get("captured_at"),
                "snapshot_row": snapshot_row or None,
                "live_quote": live_quote,
                "predicted_price_formula_ok": formula_ok,
            }
        )
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/delisted-tickers")
def api_delisted_tickers():
    """
    Return tracked unavailable tickers.
    Query params:
      - min_hits (default 1)
      - status (optional: watchlist|delisted_candidate|recovered)
    """
    try:
        min_hits = int(request.args.get("min_hits", 1))
    except Exception:
        min_hits = 1
    status_filter = request.args.get("status", "").strip().lower()

    _prune_delisted_registry()
    with _delisted_lock:
        registry = _load_delisted_registry_unlocked()

    rows = []
    for row in registry.values():
        hits = int(row.get("hit_count", 0) or 0)
        status = str(row.get("status", "")).lower()
        if hits < min_hits:
            continue
        if status_filter and status != status_filter:
            continue
        rows.append(row)

    rows.sort(
        key=lambda r: (int(r.get("hit_count", 0) or 0), r.get("last_seen", "")),
        reverse=True,
    )
    return jsonify(
        {"count": len(rows), "tickers": rows, "file": str(DELISTED_TICKERS_FILE)}
    )


@app.route("/api/delisted-tickers/export.csv")
def api_delisted_tickers_export_csv():
    """Export unavailable/delisted tracking sheet as CSV."""
    _prune_delisted_registry()
    with _delisted_lock:
        registry = _load_delisted_registry_unlocked()

    fieldnames = [
        "ticker",
        "first_seen",
        "last_seen",
        "hit_count",
        "last_reason",
        "last_source",
        "status",
    ]
    out = StringIO()
    writer = csv.DictWriter(out, fieldnames=fieldnames)
    writer.writeheader()
    for ticker in sorted(registry):
        writer.writerow(registry[ticker])

    resp = make_response(out.getvalue())
    resp.headers["Content-Type"] = "text/csv; charset=utf-8"
    resp.headers["Content-Disposition"] = "attachment; filename=delisted_tickers.csv"
    return resp


@app.route("/api/portfolio", methods=["GET", "POST", "DELETE"])
def api_portfolio():
    """Portfolio positions derived from trade ledger; backward compatible POST adds BUY."""
    _sanitize_portfolio_storage()

    if request.method == "GET":
        summary = _portfolio_summary_from_trades(_read_portfolio_trades())
        entries = summary["positions"]
        ticker = request.args.get("ticker", "").strip()
        if ticker:
            entries = [e for e in entries if e.get("ticker") == ticker]
        return jsonify({"holdings": entries, "count": len(entries)})

    if request.method == "DELETE":
        ticker = request.args.get("ticker", "").strip()
        if not ticker:
            _write_portfolio([])
            _write_portfolio_trades([])
            return jsonify({"ok": True, "holdings": [], "count": 0})
        trades = [e for e in _read_portfolio_trades() if e.get("ticker") != ticker]
        _write_portfolio_trades(trades)
        summary = _portfolio_summary_from_trades(trades)
        return jsonify(
            {
                "ok": True,
                "holdings": summary["positions"],
                "count": summary["position_count"],
            }
        )

    payload = request.get_json(silent=True) or {}
    ticker = str(payload.get("ticker", "")).strip().upper()
    qty = payload.get("quantity")
    entry_price = payload.get("entry_price")
    if not ticker:
        return jsonify({"error": "ticker is required"}), 400
    if not _is_tradeable_ticker(ticker):
        return jsonify({"error": f"{ticker} is not in configured ticker universe"}), 400
    try:
        qty = float(qty)
        entry_price = float(entry_price)
    except (TypeError, ValueError):
        return jsonify({"error": "quantity and entry_price must be numeric"}), 400
    if qty <= 0 or entry_price <= 0:
        return jsonify({"error": "quantity and entry_price must be > 0"}), 400

    # Backward-compatible path: this endpoint creates a BUY trade.
    trades = _read_portfolio_trades()
    trades.append(
        _new_trade_entry(ticker=ticker, side="BUY", qty=qty, price=entry_price)
    )
    _write_portfolio_trades(trades)
    summary = _portfolio_summary_from_trades(trades)
    _write_portfolio(
        [
            {
                "ticker": row["ticker"],
                "name": row["name"],
                "quantity": row["quantity"],
                "entry_price": row["avg_buy_price"],
                "updated_at": datetime.now(IST).isoformat(),
            }
            for row in summary["positions"]
        ]
    )
    return jsonify(
        {
            "ok": True,
            "holdings": summary["positions"],
            "count": summary["position_count"],
        }
    )


@app.route("/api/portfolio/trade", methods=["POST"])
def api_portfolio_trade():
    """Record BUY/SELL trade and return updated P&L summary."""
    _sanitize_portfolio_storage()
    payload = request.get_json(silent=True) or {}
    ticker = str(payload.get("ticker", "")).strip().upper()
    side = str(payload.get("side", "BUY")).strip().upper()
    qty = payload.get("quantity")
    price = payload.get("price")
    if side not in {"BUY", "SELL"}:
        return jsonify({"error": "side must be BUY or SELL"}), 400
    if not ticker:
        return jsonify({"error": "ticker is required"}), 400
    if not _is_tradeable_ticker(ticker):
        return jsonify({"error": f"{ticker} is not in configured ticker universe"}), 400
    try:
        qty = float(qty)
        price = float(price)
    except (TypeError, ValueError):
        return jsonify({"error": "quantity and price must be numeric"}), 400
    if qty <= 0 or price <= 0:
        return jsonify({"error": "quantity and price must be > 0"}), 400

    trades = _read_portfolio_trades()
    summary_before = _portfolio_summary_from_trades(trades, include_live_prices=False)
    pos_before = {p["ticker"]: p for p in summary_before["positions"]}
    if side == "SELL":
        open_qty = float(pos_before.get(ticker, {}).get("quantity", 0) or 0)
        if open_qty + 1e-9 < qty:
            return (
                jsonify({"error": f"Cannot SELL {qty}; open quantity is {open_qty}"}),
                400,
            )

    trades.append(_new_trade_entry(ticker=ticker, side=side, qty=qty, price=price))
    _write_portfolio_trades(trades)
    summary = _portfolio_summary_from_trades(trades)
    _write_portfolio(
        [
            {
                "ticker": row["ticker"],
                "name": row["name"],
                "quantity": row["quantity"],
                "entry_price": row["avg_buy_price"],
                "updated_at": datetime.now(IST).isoformat(),
            }
            for row in summary["positions"]
        ]
    )
    return jsonify({"ok": True, "summary": summary})


@app.route("/api/portfolio/trade/<trade_id>", methods=["DELETE", "PATCH"])
def api_portfolio_trade_edit(trade_id: str):
    """Edit or delete a trade row by ID."""
    _sanitize_portfolio_storage()
    trades = _read_portfolio_trades()
    idx = next(
        (i for i, tr in enumerate(trades) if str(tr.get("id", "")) == trade_id), None
    )
    if idx is None:
        return jsonify({"error": f"trade_id {trade_id} not found"}), 404

    if request.method == "DELETE":
        trades.pop(idx)
        _write_portfolio_trades(trades)
        summary = _portfolio_summary_from_trades(trades)
        _write_portfolio(
            [
                {
                    "ticker": row["ticker"],
                    "name": row["name"],
                    "quantity": row["quantity"],
                    "entry_price": row["avg_buy_price"],
                    "updated_at": datetime.now(IST).isoformat(),
                }
                for row in summary["positions"]
            ]
        )
        return jsonify({"ok": True, "summary": summary, "trade_count": len(trades)})

    payload = request.get_json(silent=True) or {}
    old = dict(trades[idx])

    ticker = str(payload.get("ticker", old.get("ticker", ""))).strip().upper()
    side = str(payload.get("side", old.get("side", "BUY"))).strip().upper()
    try:
        qty = float(payload.get("quantity", old.get("quantity", 0)) or 0)
        price = float(payload.get("price", old.get("price", 0)) or 0)
    except Exception:
        return jsonify({"error": "quantity and price must be numeric"}), 400

    if side not in {"BUY", "SELL"}:
        return jsonify({"error": "side must be BUY or SELL"}), 400
    if qty <= 0 or price <= 0:
        return jsonify({"error": "quantity and price must be > 0"}), 400
    if not _is_tradeable_ticker(ticker):
        return jsonify({"error": f"{ticker} is not in configured ticker universe"}), 400

    trades[idx].update(
        {
            "ticker": ticker,
            "name": ticker_names.get(ticker, _clean_name(ticker)),
            "side": side,
            "quantity": round(qty, 4),
            "price": round(price, 2),
            "edited_at": datetime.now(IST).isoformat(),
        }
    )
    valid, err = _validate_trade_sequence(trades)
    if not valid:
        trades[idx] = old
        return jsonify({"error": err}), 400

    _write_portfolio_trades(trades)
    summary = _portfolio_summary_from_trades(trades)
    _write_portfolio(
        [
            {
                "ticker": row["ticker"],
                "name": row["name"],
                "quantity": row["quantity"],
                "entry_price": row["avg_buy_price"],
                "updated_at": datetime.now(IST).isoformat(),
            }
            for row in summary["positions"]
        ]
    )
    return jsonify({"ok": True, "summary": summary, "trade_count": len(trades)})


@app.route("/api/portfolio/clean", methods=["POST"])
def api_portfolio_clean():
    """Force cleanup of invalid/dummy portfolio rows."""
    cleaned = _sanitize_portfolio_storage()
    return jsonify({"ok": True, **cleaned})


@app.route("/api/portfolio/summary")
def api_portfolio_summary():
    """Return portfolio summary with optional Groq suggestion."""
    _sanitize_portfolio_storage()
    suggest = request.args.get("suggest", "").lower() in ("1", "true", "yes")
    trades = _read_portfolio_trades()
    summary = _portfolio_summary_from_trades(trades)
    if suggest:
        try:
            summary["strategy_suggestion"] = portfolio_profit_suggestion(summary)
        except Exception as e:
            summary["strategy_suggestion"] = f"Suggestion unavailable: {e}"
    summary["trade_count"] = len(trades)
    return jsonify(summary)


@app.route("/api/portfolio/refresh")
def api_portfolio_refresh():
    """
    Atomic portfolio refresh payload for UI.
    Ensures holdings and summary are computed from the same trade snapshot and
    live price batch.
    """
    _sanitize_portfolio_storage()
    suggest = request.args.get("suggest", "").lower() in ("1", "true", "yes")
    ticker = request.args.get("ticker", "").strip().upper()
    limit = request.args.get("limit")

    trades = _read_portfolio_trades()
    if _ensure_trade_ids(trades):
        _write_portfolio_trades(trades)
    summary = _portfolio_summary_from_trades(trades)
    if suggest:
        try:
            summary["strategy_suggestion"] = portfolio_profit_suggestion(summary)
        except Exception as e:
            summary["strategy_suggestion"] = f"Suggestion unavailable: {e}"

    holdings = list(summary.get("positions", []))
    if ticker:
        holdings = [h for h in holdings if str(h.get("ticker", "")).upper() == ticker]

    view_trades = sorted(trades, key=lambda x: x.get("timestamp", ""), reverse=True)
    if ticker:
        view_trades = [
            t for t in view_trades if str(t.get("ticker", "")).upper() == ticker
        ]
    if limit:
        try:
            n = max(1, int(limit))
            view_trades = view_trades[:n]
        except ValueError:
            pass

    summary["trade_count"] = len(trades)
    return jsonify(
        {
            "summary": summary,
            "holdings": holdings,
            "trades": view_trades,
            "count": len(holdings),
            "trade_count": len(view_trades),
            "refreshed_at": datetime.now(IST).isoformat(),
        }
    )


@app.route("/api/portfolio/trades")
def api_portfolio_trades():
    """Return full trade history."""
    _sanitize_portfolio_storage()
    ticker = request.args.get("ticker", "").strip().upper()
    limit = request.args.get("limit")
    trades = _read_portfolio_trades()
    if _ensure_trade_ids(trades):
        _write_portfolio_trades(trades)
    if ticker:
        trades = [t for t in trades if str(t.get("ticker", "")).upper() == ticker]
    trades = sorted(trades, key=lambda x: x.get("timestamp", ""), reverse=True)
    if limit:
        try:
            n = max(1, int(limit))
            trades = trades[:n]
        except ValueError:
            pass
    return jsonify({"trades": trades, "count": len(trades)})


@app.route("/api/portfolio/export.csv")
def api_portfolio_export_csv():
    """Export trade ledger as CSV."""
    _sanitize_portfolio_storage()
    trades = sorted(_read_portfolio_trades(), key=lambda x: x.get("timestamp", ""))
    out = StringIO()
    writer = csv.writer(out)
    writer.writerow(
        ["id", "timestamp", "ticker", "name", "side", "quantity", "price", "notional"]
    )
    for tr in trades:
        qty = float(tr.get("quantity", 0) or 0)
        price = float(tr.get("price", 0) or 0)
        writer.writerow(
            [
                tr.get("id", ""),
                tr.get("timestamp", ""),
                tr.get("ticker", ""),
                tr.get("name", ""),
                tr.get("side", ""),
                round(qty, 4),
                round(price, 2),
                round(qty * price, 2),
            ]
        )
    resp = make_response(out.getvalue())
    resp.headers["Content-Type"] = "text/csv; charset=utf-8"
    resp.headers["Content-Disposition"] = "attachment; filename=portfolio_trades.csv"
    return resp


@app.route("/api/groq-ticker-suggestions")
def api_groq_ticker_suggestions():
    """Suggest shortlist of tickers and reasoning from current model outputs."""
    if not models_loaded or predictor is None:
        return jsonify({"error": "Models still loading..."}), 503
    try:
        n = int(request.args.get("n", 8))
        n = max(3, min(n, 20))
    except Exception:
        n = 8

    try:
        groups = predictor.predict_top_picks_grouped(
            sectors=[
                "large_cap",
                "banking",
                "mid_cap",
                "high_volatility",
                "commodities",
            ],
            top_n=n,
        )
        candidates: list[dict] = []
        for bucket in ("top_buy", "top_hold", "top_sell"):
            for p in groups.get(bucket, []):
                if float(p.get("current_price", 0) or 0) <= 0:
                    continue
                candidates.append(
                    {
                        "ticker": p.get("ticker"),
                        "name": ticker_names.get(
                            p.get("ticker", ""), _clean_name(p.get("ticker", ""))
                        ),
                        "signal": p.get("signal", "HOLD"),
                        "predicted_return": round(
                            float(p.get("predicted_return", 0) or 0), 3
                        ),
                        "confidence": round(float(p.get("confidence", 0) or 0), 1),
                        "model_agreement": round(
                            float(p.get("model_agreement", 0) or 0), 1
                        ),
                        "current_price": round(
                            float(p.get("current_price", 0) or 0), 2
                        ),
                    }
                )
        # Keep unique symbols and cap length.
        seen: set[str] = set()
        unique_candidates = []
        for c in candidates:
            t = c.get("ticker")
            if not t or t in seen:
                continue
            seen.add(t)
            unique_candidates.append(c)
            if len(unique_candidates) >= n:
                break

        recommendation = suggest_ticker_shortlist(unique_candidates)
        return jsonify(
            {
                "count": len(unique_candidates),
                "candidates": unique_candidates,
                "recommendation": recommendation,
                "generated_at": datetime.now(IST).strftime("%d %b %Y, %I:%M %p IST"),
            }
        )
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/groq-trade-review", methods=["POST"])
def api_groq_trade_review():
    """Review user-selected ticker + entry plan with model/news context."""
    if not models_loaded or predictor is None:
        return jsonify({"error": "Models still loading..."}), 503

    payload = request.get_json(silent=True) or {}
    ticker = str(payload.get("ticker", "")).strip().upper()
    if not ticker:
        return jsonify({"error": "ticker is required"}), 400
    if not _is_tradeable_ticker(ticker):
        return jsonify({"error": f"{ticker} is not in configured ticker universe"}), 400
    try:
        entry_price = float(payload.get("entry_price", 0) or 0)
        quantity = float(payload.get("quantity", 0) or 0)
    except Exception:
        return jsonify({"error": "entry_price and quantity must be numeric"}), 400
    if entry_price <= 0 or quantity <= 0:
        return jsonify({"error": "entry_price and quantity must be > 0"}), 400

    try:
        pred = predictor.predict_single(ticker, use_cache=True) or {}
        px = _get_live_prices_batch([ticker]).get(ticker, {})
        current_price = float(px.get("price", 0) or pred.get("current_price", 0) or 0)
        signal = pred.get("signal", "HOLD")
        predicted_return = float(pred.get("predicted_return", 0) or 0)
        confidence = float(pred.get("confidence", 0) or 0)
        agreement = float(pred.get("model_agreement", 0) or 0)
        stock_name = ticker_names.get(ticker, _clean_name(ticker))
        sentiment = get_news_sentiment(ticker, stock_name)
        review = review_trade_plan(
            ticker=ticker,
            stock_name=stock_name,
            entry_price=entry_price,
            quantity=quantity,
            current_price=current_price,
            signal=signal,
            predicted_return_pct=predicted_return,
            confidence=confidence,
            agreement=agreement,
            sentiment_text=sentiment,
        )
        return jsonify(
            {
                "ticker": ticker,
                "name": stock_name,
                "entry_price": round(entry_price, 2),
                "quantity": quantity,
                "current_price": round(current_price, 2) if current_price > 0 else None,
                "signal": signal,
                "predicted_return": round(predicted_return, 3),
                "confidence": round(confidence, 1),
                "model_agreement": round(agreement, 1),
                "review": review,
                "generated_at": datetime.now(IST).strftime("%d %b %Y, %I:%M %p IST"),
            }
        )
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/ai-risk-analysis")
def api_ai_risk_analysis():
    """AI-generated risk analysis page payload based on current portfolio summary."""
    _sanitize_portfolio_storage()
    summary = _portfolio_summary_from_trades(_read_portfolio_trades())
    try:
        analysis = ai_risk_assessment(summary)
    except Exception as e:
        analysis = f"AI risk analysis unavailable: {e}"
    return jsonify(
        {
            "summary": summary,
            "analysis": analysis,
            "generated_at": datetime.now(IST).strftime("%d %b %Y, %I:%M %p IST"),
        }
    )


@app.route("/api/explain-model")
def api_explain_model():
    """Explain model purpose and behavior."""
    model = request.args.get("model", "").strip()
    if not model:
        return jsonify({"error": "model is required"}), 400
    try:
        return jsonify({"model": model, "explanation": explain_model(model)})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/stock-chat/<ticker>", methods=["POST"])
def api_stock_chat(ticker: str):
    """Contextual stock Q&A assistant."""
    payload = request.get_json(silent=True) or {}
    question = str(payload.get("question", "")).strip()
    if not question:
        return jsonify({"error": "question is required"}), 400
    if not models_loaded or predictor is None:
        return jsonify({"error": "Models still loading..."}), 503
    try:
        pred = predictor.predict_single(ticker, use_cache=True) or {}
        indicators = pred.get("indicators", {})
        stock_name = ticker_names.get(ticker, _clean_name(ticker))
        answer = stock_chat_response(
            ticker=ticker,
            stock_name=stock_name,
            question=question,
            prediction_data=pred,
            indicator_snapshot=indicators,
        )
        return jsonify({"ticker": ticker, "answer": answer})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/feature-importance/<ticker>")
def api_feature_importance(ticker: str):
    """Return feature importance for a stock's prediction."""
    if not models_loaded or predictor is None:
        return jsonify({"error": "Models still loading..."}), 503

    try:
        pred = predictor.predict_single(ticker, use_cache=True)
        if pred is None:
            return jsonify({"error": f"No prediction for {ticker}"}), 404

        # Extract feature importances from tree-based models
        importances = {}
        if hasattr(predictor, "models") and predictor.models:
            for name, model in predictor.models.items():
                inner = getattr(model, "model", model)
                if hasattr(inner, "feature_importances_"):
                    fi = inner.feature_importances_
                    feat_names = getattr(inner, "feature_names_in_", None) or getattr(
                        model, "feature_names", None
                    )
                    if feat_names is not None:
                        # Pair names with importances, sort by importance desc
                        paired = sorted(
                            zip(feat_names, fi), key=lambda x: x[1], reverse=True
                        )[:20]
                        importances[name] = {
                            str(n): round(float(v), 4) for n, v in paired
                        }

        # Provide top indicators affecting price
        indicators = pred.get("indicators", {})
        fundamentals = pred.get("fundamentals", {})

        key_factors = []
        # RSI
        rsi = indicators.get("rsi")
        if rsi is not None:
            if rsi > 70:
                key_factors.append(
                    {
                        "factor": "RSI (Overbought)",
                        "value": f"{rsi:.1f}",
                        "impact": "bearish",
                        "weight": 0.15,
                    }
                )
            elif rsi < 30:
                key_factors.append(
                    {
                        "factor": "RSI (Oversold)",
                        "value": f"{rsi:.1f}",
                        "impact": "bullish",
                        "weight": 0.15,
                    }
                )
            else:
                key_factors.append(
                    {
                        "factor": "RSI",
                        "value": f"{rsi:.1f}",
                        "impact": "neutral",
                        "weight": 0.10,
                    }
                )

        # MACD
        macd = indicators.get("macd")
        macd_signal = indicators.get("macd_signal")
        if macd is not None and macd_signal is not None:
            if macd > macd_signal:
                key_factors.append(
                    {
                        "factor": "MACD Crossover",
                        "value": f"{macd:.3f}",
                        "impact": "bullish",
                        "weight": 0.12,
                    }
                )
            else:
                key_factors.append(
                    {
                        "factor": "MACD Crossover",
                        "value": f"{macd:.3f}",
                        "impact": "bearish",
                        "weight": 0.12,
                    }
                )

        # Volume
        vol_ratio = indicators.get("volume_ratio")
        if vol_ratio is not None:
            if vol_ratio > 1.5:
                key_factors.append(
                    {
                        "factor": "High Volume",
                        "value": f"{vol_ratio:.2f}x",
                        "impact": "bullish",
                        "weight": 0.10,
                    }
                )
            elif vol_ratio < 0.5:
                key_factors.append(
                    {
                        "factor": "Low Volume",
                        "value": f"{vol_ratio:.2f}x",
                        "impact": "bearish",
                        "weight": 0.08,
                    }
                )

        # ADX
        adx = indicators.get("adx")
        if adx is not None:
            key_factors.append(
                {
                    "factor": "ADX (Trend Strength)",
                    "value": f"{adx:.1f}",
                    "impact": "bullish" if adx > 25 else "neutral",
                    "weight": 0.08,
                }
            )

        # ATR
        atr = indicators.get("atr_pct")
        if atr is not None:
            key_factors.append(
                {
                    "factor": "Volatility (ATR%)",
                    "value": f"{atr*100:.2f}%",
                    "impact": "neutral",
                    "weight": 0.07,
                }
            )

        # Fundamentals
        pe = fundamentals.get("pe_ratio")
        if pe is not None and pe > 0:
            key_factors.append(
                {
                    "factor": "P/E Ratio",
                    "value": f"{pe:.1f}",
                    "impact": (
                        "bearish" if pe > 40 else "neutral" if pe > 20 else "bullish"
                    ),
                    "weight": 0.10,
                }
            )

        # Sort by weight
        key_factors.sort(key=lambda x: x["weight"], reverse=True)

        return jsonify(
            {
                "ticker": ticker,
                "key_factors": key_factors,
                "model_importances": importances,
            }
        )
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ---------------------------------------------------------------------------
# Helpers — JSON Sanitization for numpy types
# ---------------------------------------------------------------------------


def _sanitize(obj):
    """Recursively convert numpy types to Python-native for JSON serialization."""
    if isinstance(obj, dict):
        return {k: _sanitize(v) for k, v in obj.items()}
    if isinstance(obj, (list, tuple)):
        return [_sanitize(v) for v in obj]
    if isinstance(obj, (np.integer,)):
        return int(obj)
    if isinstance(obj, (np.floating,)):
        return float(obj)
    if isinstance(obj, (np.bool_,)):
        return bool(obj)
    if isinstance(obj, np.ndarray):
        return obj.tolist()
    return obj


# ---------------------------------------------------------------------------
# Routes — Portfolio Risk Analytics (Institutional Grade)
# ---------------------------------------------------------------------------


@app.route("/api/risk-analytics")
def api_risk_analytics():
    """
    Comprehensive portfolio-level risk analytics dashboard.
    Returns correlation matrix, sector exposure, risk metrics,
    Monte Carlo simulation, and statistical validation.
    """
    from src.backtest.metrics import (
        sharpe_ratio as calc_sharpe,
        sortino_ratio as calc_sortino,
        max_drawdown as calc_mdd,
        max_drawdown_duration as calc_mdd_dur,
        value_at_risk as calc_var,
        conditional_var as calc_cvar,
        parametric_var as calc_pvar,
        return_skewness,
        return_kurtosis,
        tail_ratio as calc_tail,
        jarque_bera_test,
        monte_carlo_backtest,
        information_ratio as calc_ir,
    )

    try:
        _sanitize_portfolio_storage()
        manual_rows = _read_portfolio()
        sim_rows = []
        try:
            sim_state = _read_portfolio_sim_state()
            for t, pos in dict(sim_state.get("open_positions", {})).items():
                ticker = str(t or "").strip().upper()
                qty = _safe_float((pos or {}).get("quantity"))
                entry = _safe_float((pos or {}).get("avg_entry_price"))
                if ticker and qty > 0 and entry > 0:
                    sim_rows.append(
                        {
                            "ticker": ticker,
                            "name": ticker_names.get(ticker, _clean_name(ticker)),
                            "quantity": qty,
                            "entry_price": entry,
                            "source": "simulation",
                        }
                    )
        except Exception:
            sim_rows = []

        merged_rows: dict[str, dict] = {}

        def _merge_row(row: dict, source: str) -> None:
            ticker = str(row.get("ticker", "")).strip().upper()
            qty = _safe_float(row.get("quantity"))
            entry = _safe_float(row.get("entry_price"))
            if not ticker or qty <= 0 or entry <= 0:
                return
            prev = merged_rows.get(ticker)
            if not prev:
                merged_rows[ticker] = {
                    "ticker": ticker,
                    "name": row.get("name")
                    or ticker_names.get(ticker, _clean_name(ticker)),
                    "quantity": qty,
                    "entry_price": entry,
                    "source": source,
                }
                return
            prev_qty = _safe_float(prev.get("quantity"))
            prev_px = _safe_float(prev.get("entry_price"))
            total_qty = prev_qty + qty
            avg_px = (
                (prev_qty * prev_px + qty * entry) / total_qty
                if total_qty > 0
                else max(prev_px, entry)
            )
            prev["quantity"] = round(total_qty, 6)
            prev["entry_price"] = round(avg_px, 6)
            prev["source"] = "merged"
            merged_rows[ticker] = prev

        for row in manual_rows:
            _merge_row(row if isinstance(row, dict) else {}, "manual")
        for row in sim_rows:
            _merge_row(row, "simulation")

        portfolio_rows = list(merged_rows.values())
        if not portfolio_rows:
            return (
                jsonify(
                    {
                        "error": "No portfolio holdings found. Add portfolio positions to run risk analytics.",
                        "portfolio_tickers": [],
                    }
                ),
                400,
            )

        custom_weights = {}
        portfolio_universe = []
        if portfolio_rows:
            for row in portfolio_rows:
                t = row.get("ticker")
                q = float(row.get("quantity", 0) or 0)
                p = float(row.get("entry_price", 0) or 0)
                if t and q > 0 and p > 0:
                    portfolio_universe.append(t)
                    custom_weights[t] = custom_weights.get(t, 0.0) + (q * p)
        portfolio_universe = sorted(set(portfolio_universe))
        if not portfolio_universe:
            return (
                jsonify(
                    {
                        "error": "No valid portfolio holdings found. Remove invalid tickers and add holdings again.",
                        "portfolio_tickers": [],
                    }
                ),
                400,
            )

        raw_requested = request.args.getlist("tickers")
        requested: list[str] = []
        for raw in raw_requested:
            requested.extend(
                [t.strip().upper() for t in str(raw).split(",") if t.strip()]
            )
        if not requested:
            tickers_param = request.args.get("tickers", "")
            if tickers_param:
                requested = [
                    t.strip().upper()
                    for t in str(tickers_param).split(",")
                    if t.strip()
                ]
        if requested:
            portfolio_tickers = [t for t in requested if t in portfolio_universe]
            ignored = [t for t in requested if t not in portfolio_universe]
        else:
            portfolio_tickers = list(portfolio_universe)
            ignored = []

        if not portfolio_tickers:
            return (
                jsonify(
                    {
                        "error": "Requested tickers are not in portfolio holdings.",
                        "portfolio_tickers": [],
                        "ignored_tickers": ignored,
                    }
                ),
                400,
            )

        # Download 6 months of history for robust risk calcs
        ticker_str = " ".join(portfolio_tickers + ["^NSEI"])
        data = _safe_yf_download(
            ticker_str,
            period="6mo",
            interval="1d",
            auto_adjust=True,
        )

        if data is None or data.empty:
            return jsonify({"error": "Unable to fetch historical data"}), 500

        # Extract closing prices
        if isinstance(data.columns, pd.MultiIndex):
            close = data["Close"].dropna(how="all")
        else:
            close = data[["Close"]].dropna()

        # Separate benchmark before filtering
        bench_close = close["^NSEI"] if "^NSEI" in close.columns else None
        port_close = close.drop(columns=["^NSEI"], errors="ignore")

        # Keep ticker history if it has at least a minimum observation count.
        min_obs = 20
        if len(port_close) > 0:
            port_close = port_close.dropna(axis=1, thresh=min_obs)

        # Recover missing symbols via single-ticker pulls (yfinance batch can skip some).
        missing_tickers = [t for t in portfolio_tickers if t not in port_close.columns]
        for t in missing_tickers:
            single = _safe_yf_download(
                t,
                period="6mo",
                interval="1d",
                auto_adjust=True,
            )
            if single is None or single.empty:
                continue
            ser = None
            if isinstance(single.columns, pd.MultiIndex):
                lvl0 = single.columns.get_level_values(0)
                if "Close" in lvl0:
                    obj = single["Close"]
                    if isinstance(obj, pd.DataFrame):
                        if t in obj.columns:
                            ser = obj[t]
                        elif obj.shape[1] == 1:
                            ser = obj.iloc[:, 0]
                    else:
                        ser = obj
            else:
                if "Close" in single.columns:
                    ser = single["Close"]
            if ser is None:
                continue
            ser = ser.dropna()
            if len(ser) < min_obs:
                continue
            port_close[t] = ser

        port_close = port_close.sort_index().dropna(axis=0, how="all")
        actual_period_days = len(port_close)

        # Recombine with benchmark for returns calculation
        if bench_close is not None:
            combined = pd.concat([port_close, bench_close.rename("^NSEI")], axis=1)
        else:
            combined = port_close
        returns = combined.pct_change().dropna()

        # Update portfolio_tickers to only surviving tickers
        surviving_tickers = [c for c in port_close.columns if c in portfolio_tickers]

        # --- Correlation Matrix (portfolio tickers only) ---
        port_returns_only = returns[[c for c in returns.columns if c != "^NSEI"]]
        corr_matrix = port_returns_only.corr()
        corr_data = {}
        for col in corr_matrix.columns:
            corr_data[col] = {
                c: round(float(v), 3) for c, v in corr_matrix[col].items()
            }

        # --- Sector Exposure ---
        sector_exposure: dict[str, dict] = {}
        total_invested_surviving = sum(
            custom_weights.get(t, 0.0) for t in surviving_tickers
        )
        for t in surviving_tickers:
            sec_key = str(ticker_to_sector.get(t, "other"))
            sec_label = SECTOR_DISPLAY.get(sec_key, sec_key.replace("_", " ").title())
            bucket = sector_exposure.setdefault(
                sec_label,
                {"count": 0, "tickers": [], "invested_value": 0.0, "weight_pct": 0.0},
            )
            bucket["count"] += 1
            bucket["tickers"].append(t)
            bucket["invested_value"] += float(custom_weights.get(t, 0.0) or 0.0)

        for sec_label, bucket in sector_exposure.items():
            invested = float(bucket.get("invested_value", 0.0) or 0.0)
            if total_invested_surviving > 0:
                bucket["weight_pct"] = round(
                    invested / total_invested_surviving * 100.0, 1
                )
            else:
                bucket["weight_pct"] = round(
                    bucket.get("count", 0) / max(len(surviving_tickers), 1) * 100.0,
                    1,
                )
            bucket["position_pct"] = round(
                bucket.get("count", 0) / max(len(surviving_tickers), 1) * 100.0,
                1,
            )
            bucket["invested_value"] = round(invested, 2)

        # --- Portfolio-level returns (weighted by user holdings if available) ---
        portfolio_cols = [
            c for c in returns.columns if c != "^NSEI" and c in surviving_tickers
        ]
        if not portfolio_cols:
            return jsonify({"error": "Insufficient data for analytics"}), 500

        if custom_weights:
            total_cost = sum(custom_weights.get(t, 0.0) for t in portfolio_cols)
            if total_cost > 0:
                normalized = {
                    t: custom_weights.get(t, 0.0) / total_cost for t in portfolio_cols
                }
            else:
                normalized = {t: 1.0 / len(portfolio_cols) for t in portfolio_cols}
            weighted = pd.Series(0.0, index=returns.index)
            for t in portfolio_cols:
                weighted = weighted + returns[t] * normalized.get(t, 0.0)
            port_returns = weighted
        else:
            port_returns = returns[portfolio_cols].mean(axis=1)
        if len(port_returns.dropna()) < 20:
            return (
                jsonify(
                    {
                        "error": "Insufficient portfolio return history for risk analytics (need at least 20 observations).",
                        "portfolio_tickers": surviving_tickers,
                    }
                ),
                400,
            )
        initial_capital = float(sum(custom_weights.get(t, 0.0) for t in portfolio_cols))
        if initial_capital <= 0:
            initial_capital = 100000.0
        port_equity = (1 + port_returns).cumprod() * initial_capital

        benchmark_returns = returns["^NSEI"] if "^NSEI" in returns.columns else None

        def _safe(fn, *args, default=0.0):
            """Safely compute a metric, returning default on NaN/error."""
            try:
                val = float(fn(*args))
                return default if (np.isnan(val) or np.isinf(val)) else round(val, 6)
            except Exception:
                return default

        # --- Risk Metrics ---
        risk_metrics = {
            "sharpe_ratio": _safe(calc_sharpe, port_returns),
            "sortino_ratio": _safe(calc_sortino, port_returns),
            "max_drawdown": _safe(calc_mdd, port_equity),
            "max_drawdown_duration_days": int(calc_mdd_dur(port_equity)),
            "daily_var_95": _safe(calc_var, port_returns, 0.95),
            "daily_cvar_95": _safe(calc_cvar, port_returns, 0.95),
            "parametric_var_cf_95": _safe(
                calc_pvar, port_returns, 0.95, "cornish_fisher"
            ),
            "skewness": _safe(return_skewness, port_returns),
            "excess_kurtosis": _safe(return_kurtosis, port_returns),
            "tail_ratio": _safe(calc_tail, port_returns, default=1.0),
            "annualized_volatility": round(float(port_returns.std() * np.sqrt(252)), 4),
        }

        # Add benchmark-relative metrics
        if benchmark_returns is not None and not benchmark_returns.empty:
            risk_metrics["information_ratio"] = _safe(
                calc_ir, port_returns, benchmark_returns
            )

        # --- Statistical Tests ---
        stat_tests = {
            "jarque_bera": jarque_bera_test(port_returns),
        }

        # --- Monte Carlo ---
        mc_result = monte_carlo_backtest(
            port_returns,
            n_simulations=500,
            initial_capital=initial_capital,
        )

        # --- Equity Curve (for chart) ---
        equity_data = [
            {"date": d.strftime("%Y-%m-%d"), "value": round(float(v), 2)}
            for d, v in port_equity.items()
        ]

        return jsonify(
            _sanitize(
                {
                    "portfolio_tickers": surviving_tickers,
                    "portfolio_holdings": [
                        r
                        for r in portfolio_rows
                        if r.get("ticker") in surviving_tickers
                    ],
                    "portfolio_source": (
                        "merged_manual_and_simulation"
                        if manual_rows and sim_rows
                        else ("simulation" if sim_rows else "manual")
                    ),
                    "initial_capital": round(initial_capital, 2),
                    "ignored_tickers": ignored,
                    "n_stocks": len(surviving_tickers),
                    "period": f"{actual_period_days} days",
                    "risk_metrics": risk_metrics,
                    "correlation_matrix": corr_data,
                    "sector_exposure": sector_exposure,
                    "statistical_tests": stat_tests,
                    "monte_carlo": mc_result,
                    "equity_curve": equity_data,
                }
            )
        )

    except Exception as e:
        log.exception("Risk analytics error")
        return jsonify({"error": str(e)}), 500


@app.route("/risk")
def risk_dashboard():
    """Portfolio risk analytics dashboard page."""
    return render_template("risk.html")


@app.route("/ai-risk")
def ai_risk_dashboard():
    """AI-only risk interpretation page."""
    return render_template("ai_risk.html")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
if __name__ == "__main__":
    load_tickers()
    log.info(
        f"Loaded {len(all_tickers)} tickers across {len(tickers_by_sector)} sectors"
    )
    cleaned = _sanitize_portfolio_storage()
    if cleaned.get("changed"):
        log.info(
            "Portfolio sanitized at startup: removed_trades=%s removed_holdings=%s remaining_trades=%s",
            cleaned.get("removed_trade_rows"),
            cleaned.get("removed_holding_rows"),
            cleaned.get("remaining_trades"),
        )
    pruned = _prune_delisted_registry()
    if pruned.get("changed"):
        log.info(
            "Delisted registry pruned at startup: removed=%s remaining=%s",
            pruned.get("removed"),
            pruned.get("remaining"),
        )
    _warm_sim_trade_cache_from_log()
    if TRADE_EMAIL_ENABLED and not _smtp_ready():
        log.warning(
            "TRADE_EMAIL_ENABLED is true but SMTP is not fully configured. "
            "Set SMTP_HOST/SMTP_PORT/SMTP_FROM (and credentials if needed)."
        )

    # Load models in background thread
    model_thread = threading.Thread(target=load_models_background, daemon=True)
    model_thread.start()

    # Start daily analysis background computation thread
    analysis_thread = threading.Thread(
        target=_daily_analysis_background_loop, daemon=True
    )
    analysis_thread.start()
    log.info("Daily analysis background thread started")

    precompute_thread = threading.Thread(target=_scheduled_precompute_loop, daemon=True)
    precompute_thread.start()
    log.info(
        "Scheduled precompute thread started (%02d:%02d and %02d:%02d IST)",
        PRECOMPUTE_OPEN_HOUR,
        PRECOMPUTE_OPEN_MINUTE,
        PRECOMPUTE_CLOSE_HOUR,
        PRECOMPUTE_CLOSE_MINUTE,
    )

    auto_sim_thread = threading.Thread(
        target=_auto_simulation_background_loop,
        daemon=True,
    )
    auto_sim_thread.start()
    log.info(
        "Auto simulation thread started (enabled=%s, auto_buy=%s, interval=%ss)",
        AUTO_SIMULATION_ENABLED,
        AUTO_SIMULATION_AUTO_BUY,
        AUTO_SIMULATION_INTERVAL_SEC,
    )

    report_thread = threading.Thread(
        target=_daily_trade_report_scheduler_loop,
        daemon=True,
    )
    report_thread.start()
    log.info(
        "Daily trade report scheduler started (%02d:%02d IST)",
        TRADE_DAILY_REPORT_HOUR_IST,
        TRADE_DAILY_REPORT_MINUTE_IST,
    )

    port = int(os.environ.get("FLASK_PORT", 5001))
    threaded_mode = os.environ.get("FLASK_THREADED", "1").lower() in (
        "1",
        "true",
        "yes",
    )
    log.info(f"Starting Flask server at http://localhost:{port}")
    app.run(host="0.0.0.0", port=port, debug=False, threaded=threaded_mode)
